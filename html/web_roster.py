"""HTML engine: web roster — adjust template and export to HTML.

This engine is completely independent from the Excel export engines.
It loads a template workbook, adjusts it for the target month, fills
personnel data, and exports the result as HTML.
"""
from __future__ import annotations

import calendar
import re
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import range_boundaries

from ..common.data_source import resolve_template
from ..common.shift_info import _get_holidays_for_month, WEEKDAY_CH
from .. import excel_row_ops, excel_column_ops
from ..common.col_adjust import update_all_formulas_for_col_change


# ── Constants ──
DAY_START_COL = 4
RANK_COL = 2
NAME_COL = 3
TEMPLATE_DAYS = 31

_TITLE_RE = re.compile(r"^(\d+)\.\s")
_RANK_RE = re.compile(r"^[A-Z]{1,2}\d+$")
_CROSS_SHEET_DATE_RE = re.compile(r"^=[^!'\"\s]+![A-Z]+\d+$")


# ── Inlined from cleaning_shortfall_v2 (to keep engines independent) ──

def _find_date_start(ws):
    """Find the row and column of the first date cell in the worksheet."""
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and hasattr(v, "year"):
                return r, c
            if isinstance(v, str) and _CROSS_SHEET_DATE_RE.match(v):
                return r, c
    return None, None


def col_num_to_letter(n: int) -> str:
    """Convert 1-based column number to Excel column letter(s)."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _delete_one_column(ws, col: int) -> None:
    """Delete a single column using safe helpers (no phantom cells)."""
    update_all_formulas_for_col_change(ws, col, -1, is_delete=True)
    excel_column_ops.safe_delete_cols(ws, col, 1)
    excel_column_ops.fix_merged_cells(ws, col)
    excel_column_ops.adjust_conditional_formatting(ws, col)


def _adjust_fifth_week_summary_column(ws, days_in_month: int) -> None:
    """Delete or rename the '29-31號時數' summary column based on month length."""
    fifth_week_col = None
    for c in range(1, ws.max_column + 1):
        for r in range(1, min(ws.max_row, 15) + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "29-31" in v:
                fifth_week_col = c
                break
        if fifth_week_col:
            break

    if not fifth_week_col:
        return

    if days_in_month <= 28:
        _delete_one_column(ws, fifth_week_col)
    elif days_in_month == 30:
        for r in range(1, min(ws.max_row, 15) + 1):
            cell = ws.cell(r, fifth_week_col)
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value.replace("31", "30")


def _adjust_day_columns(ws, days_in_month: int) -> int:
    """Adjust day columns from template's 31 days to target month days."""
    date_row, first_day_col = _find_date_start(ws)
    if not first_day_col:
        return ws.max_column

    if days_in_month >= TEMPLATE_DAYS:
        return first_day_col + TEMPLATE_DAYS - 1

    for day in [31, 30, 29]:
        if days_in_month < day:
            col = first_day_col + day - 1
            _delete_one_column(ws, col)

    _adjust_fifth_week_summary_column(ws, days_in_month)
    return first_day_col + days_in_month - 1


def _update_dates(ws, month: str, date_row: int | None = None, day_start_col: int | None = None) -> None:
    """Update date row to target month (first cell = date, rest = +1 chain)."""
    if date_row is None or day_start_col is None:
        date_row, day_start_col = _find_date_start(ws)
    if not date_row:
        return

    year, month_num = int(month[:4]), int(month[5:7])
    days_in_month = calendar.monthrange(year, month_num)[1]

    ws.cell(date_row, day_start_col, value=date(year, month_num, 1))

    weekday_row = date_row + 2
    if weekday_row <= ws.max_row:
        for c in range(day_start_col, day_start_col + days_in_month):
            v = ws.cell(weekday_row, c).value
            if isinstance(v, str) and "TEXT" in v:
                col_letter = col_num_to_letter(c)
                ws.cell(weekday_row, c,
                        value=f'=SUBSTITUTE(TEXT({col_letter}{date_row},"aaa"),"週","")')


# ── Core roster processing ──

def _html_escape(s: str) -> str:
    return (
        s.replace("&", "&amp;")
         .replace("<", "&lt;")
         .replace(">", "&gt;")
         .replace('"', "&quot;")
    )


# Office default theme colour mapping (index → RRGGBB)
_THEME_COLORS = {
    0: "000000",   # dk1
    1: "FFFFFF",   # lt1
    2: "44546A",   # dk2
    3: "E7E6E6",   # lt2
    4: "4472C4",   # accent1
    5: "ED7D31",   # accent2
    6: "A5A5A5",   # accent3
    7: "FFC000",   # accent4
    8: "5B9BD5",   # accent5
    9: "70AD47",   # accent6
}


def _apply_tint(rgb_hex: str, tint: float) -> str:
    """Apply Excel tint to an RRGGBB colour."""
    r = int(rgb_hex[0:2], 16)
    g = int(rgb_hex[2:4], 16)
    b = int(rgb_hex[4:6], 16)

    if tint > 0:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    elif tint < 0:
        r = int(r * (1 + tint))
        g = int(g * (1 + tint))
        b = int(b * (1 + tint))

    r = max(0, min(255, r))
    g = max(0, min(255, g))
    b = max(0, min(255, b))
    return f"{r:02X}{g:02X}{b:02X}"


def _color_to_hex(color) -> str | None:
    """Convert openpyxl Color to CSS #RRGGBB string.

    Handles indexed, rgb, and theme color types.  Returns None for
    transparent / auto / unresolvable colours.
    """
    if color is None:
        return None

    # Indexed colours (most common in xlsx templates)
    if getattr(color, "type", None) == "indexed":
        idx = color.value
        # idx == 64 means "auto" – use default (None)
        if idx is not None and 0 <= idx < len(COLOR_INDEX):
            rgb = COLOR_INDEX[idx]
            if rgb and isinstance(rgb, str) and len(rgb) == 8 and rgb != "00000000":
                return f"#{rgb[2:]}"
        return None

    # Direct RGB (8-char AARRGGBB or 6-char RRGGBB)
    raw = None
    if getattr(color, "type", None) == "rgb":
        raw = color.value
    elif getattr(color, "rgb", None):
        raw = color.rgb
        # openpyxl rgb descriptor may raise on read; guard against error text
        if isinstance(raw, str) and "must be of type" in raw:
            raw = None
    if raw and isinstance(raw, str):
        raw = raw.strip()
        if len(raw) == 8 and raw != "00000000":
            return f"#{raw[2:]}"
        if len(raw) == 6:
            return f"#{raw}"

    # Theme colours
    if getattr(color, "type", None) == "theme":
        theme_idx = getattr(color, "theme", None)
        if theme_idx is not None and theme_idx in _THEME_COLORS:
            base = _THEME_COLORS[theme_idx]
            tint = getattr(color, "tint", 0.0) or 0.0
            if tint:
                return f"#{_apply_tint(base, tint)}"
            return f"#{base}"
        return None

    return None


# Excel border style → CSS width mapping
_BORDER_WIDTH_MAP = {
    "hair": "0.5px",
    "thin": "1px",
    "medium": "2px",
    "thick": "3px",
    "double": "3px",
    "dashed": "1px",
    "dotted": "1px",
    "mediumDashDot": "2px",
    "mediumDashDotDot": "2px",
    "slantDashDot": "1px",
}


# Excel border style → CSS border-style
_BORDER_STYLE_MAP = {
    "double": "double",
    "dashed": "dashed",
    "dotted": "dotted",
    "mediumDashDot": "dashed",
    "mediumDashDotDot": "dotted",
    "slantDashDot": "dashed",
}


def _shift_from_sheet_name(sheet_name: str) -> str:
    mapping = {"早": "A", "中": "B", "夜": "C"}
    return mapping.get(sheet_name, "A")


def _fetch_data(context: Dict[str, Any], shift: str) -> Dict[str, Any]:
    from ...api.shortfall_engine import api_shortfall_engine
    return api_shortfall_engine(
        shift=shift,
        project_id=context["project_id"],
        category_id=context["category_id"],
        month=context["month"],
    )


def _detect_roster_regions(ws) -> List[Dict[str, Any]]:
    regions = []
    for row in range(1, ws.max_row + 1):
        b_val = ws.cell(row=row, column=RANK_COL).value
        if b_val and isinstance(b_val, str) and _TITLE_RE.match(b_val):
            title_row = row
            data_start = None
            for r in range(title_row + 1, ws.max_row + 1):
                b = ws.cell(row=r, column=RANK_COL).value
                if b and isinstance(b, str) and _RANK_RE.match(b.strip()):
                    data_start = r
                    break
            if data_start is None:
                continue
            data_end = data_start
            for r in range(data_start + 1, ws.max_row + 1):
                b = ws.cell(row=r, column=RANK_COL).value
                if b and isinstance(b, str) and _RANK_RE.match(b.strip()):
                    data_end = r
                else:
                    break
            regions.append({
                "index": len(regions) + 1,
                "title_row": title_row,
                "data_start": data_start,
                "data_end": data_end,
            })
    return regions


def _shrink_region(ws, region: Dict[str, Any], target_count: int) -> int:
    current = region["data_end"] - region["data_start"] + 1
    to_delete = current - target_count
    if to_delete <= 0:
        return 0
    for _ in range(to_delete):
        row_to_delete = region["data_end"]
        excel_row_ops.delete_row_and_fixup(ws, row_to_delete)
        region["data_end"] -= 1
    return to_delete


def _safe_write(ws, row: int, col: int, value) -> bool:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


def _fill_data_row(
    ws, row: int, row_data: Dict[str, Any], day_start_col: int, days_in_month: int
) -> None:
    _safe_write(ws, row, RANK_COL, row_data.get("rank_seq") or row_data.get("employee_no") or "")
    _safe_write(ws, row, NAME_COL, row_data.get("name") or "")
    cells = {c["day"]: c for c in row_data.get("cells", [])}
    for d in range(1, days_in_month + 1):
        col = day_start_col + d - 1
        c = cells.get(d)
        if not c:
            _safe_write(ws, row, col, "")
            continue
        v = c.get("value")
        if v is None:
            v = c.get("code")
        try:
            num = float(v)
            _safe_write(ws, row, col, int(num) if num == int(num) else num)
        except (TypeError, ValueError):
            _safe_write(ws, row, col, str(v) if v else "")


def _adjust_number_formats(ws, day_start_col: int, days_in_month: int) -> None:
    for row in range(1, ws.max_row + 1):
        for col in range(day_start_col, day_start_col + days_in_month):
            cell = ws.cell(row, col)
            v = cell.value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                if isinstance(v, int) or v == int(v):
                    cell.number_format = "0"
                else:
                    cell.number_format = "0.00"


# ── Conditional formatting helpers ──

_CF_SEARCH_RE = re.compile(r'SEARCH\s*\(\s*"([^"]+)"')


def _parse_cf_formula_value(formula_list):
    """Parse cellIs formula value, e.g. ['"NC"'] -> 'NC'."""
    if not formula_list or not formula_list[0]:
        return None
    val = str(formula_list[0]).strip()
    if val.startswith('"') and val.endswith('"'):
        return val[1:-1]
    return val


def _parse_cf_contains_text(formula_list):
    """Extract search text from containsText formula."""
    if not formula_list or not formula_list[0]:
        return None
    m = _CF_SEARCH_RE.search(str(formula_list[0]))
    if m:
        return m.group(1)
    return None


def _cell_in_cf_range(sqref, row: int, col: int) -> bool:
    """Check if a cell (row, col) is in a conditional formatting sqref string."""
    sqref_str = str(sqref) if not isinstance(sqref, str) else sqref
    for part in sqref_str.split():
        min_col, min_row, max_col, max_row = range_boundaries(part)
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return True
    return False


def _cell_matches_cf_rule(cell_value, rule) -> bool:
    """Check if a cell value matches a conditional formatting rule."""
    rule_type = rule.type
    formula = rule.formula

    if rule_type == "cellIs":
        operator = getattr(rule, "operator", None)
        expected = _parse_cf_formula_value(formula)
        if expected is None:
            return False

        cell_str = str(cell_value) if cell_value is not None else ""

        if operator == "equal":
            return cell_str == expected
        elif operator == "notEqual":
            return cell_str != expected
        elif operator in ("greaterThan", "lessThan", "greaterThanOrEqual", "lessThanOrEqual"):
            try:
                cell_num = float(cell_value) if cell_value is not None else 0
                expected_num = float(expected)
                if operator == "greaterThan":
                    return cell_num > expected_num
                elif operator == "lessThan":
                    return cell_num < expected_num
                elif operator == "greaterThanOrEqual":
                    return cell_num >= expected_num
                elif operator == "lessThanOrEqual":
                    return cell_num <= expected_num
            except (ValueError, TypeError):
                return False
        return False

    elif rule_type == "containsText":
        search_text = _parse_cf_contains_text(formula)
        if search_text is None:
            return False
        cell_str = str(cell_value) if cell_value is not None else ""
        return search_text in cell_str

    # Other types not supported yet
    return False


def _get_conditional_bg_color(ws, row: int, col: int) -> str | None:
    """Get the conditional formatting background colour for a cell."""
    matching_rules = []

    for cf in ws.conditional_formatting:
        # Check if cell is in the conditional formatting range
        if not _cell_in_cf_range(cf.sqref, row, col):
            continue

        for rule in cf.rules:
            cell_value = ws.cell(row, col).value
            if _cell_matches_cf_rule(cell_value, rule):
                priority = getattr(rule, "priority", 999999)
                matching_rules.append((priority, rule))

    if not matching_rules:
        return None

    # Sort by priority (lower = higher priority)
    matching_rules.sort(key=lambda x: x[0])
    _, rule = matching_rules[0]

    if not rule.dxf or not rule.dxf.fill:
        return None

    # Conditional format dxfs often use bgColor for the background
    bg_color = None
    if rule.dxf.fill.bgColor:
        bg_color = _color_to_hex(rule.dxf.fill.bgColor)

    # Fallback to fgColor
    if not bg_color and rule.dxf.fill.fgColor:
        bg_color = _color_to_hex(rule.dxf.fill.fgColor)

    return bg_color


# ── HTML export helpers ──

def _cell_to_css(cell, conditional_bg: str | None = None) -> str:
    """Convert openpyxl cell styles to a CSS style string."""
    styles: List[str] = []

    # ── Font ──
    if cell.font:
        if cell.font.name:
            styles.append(f"font-family:{cell.font.name}")
        if cell.font.size:
            styles.append(f"font-size:{cell.font.size}pt")
        if cell.font.bold:
            styles.append("font-weight:bold")
        if cell.font.italic:
            styles.append("font-style:italic")
        if cell.font.underline and cell.font.underline != "none":
            styles.append("text-decoration:underline")
        color_hex = _color_to_hex(cell.font.color)
        if color_hex:
            styles.append(f"color:{color_hex}")

    # ── Fill (background) ──
    # Conditional formatting bg takes precedence over normal fill
    bg_color = conditional_bg
    if not bg_color and cell.fill:
        # Prefer fgColor; fallback to start_color for older openpyxl
        if cell.fill.fgColor:
            bg_color = _color_to_hex(cell.fill.fgColor)
        if not bg_color and hasattr(cell.fill, "start_color") and cell.fill.start_color:
            bg_color = _color_to_hex(cell.fill.start_color)
    if bg_color:
        styles.append(f"background-color:{bg_color}")

    # ── Border ──
    if cell.border:
        for side_name, side in [
            ("top", cell.border.top),
            ("bottom", cell.border.bottom),
            ("left", cell.border.left),
            ("right", cell.border.right),
        ]:
            if side and side.style:
                width = _BORDER_WIDTH_MAP.get(side.style, "1px")
                bstyle = _BORDER_STYLE_MAP.get(side.style, "solid")
                color_hex = _color_to_hex(side.color)
                color_str = color_hex if color_hex else "#000000"
                styles.append(f"border-{side_name}:{width} {bstyle} {color_str}")

    # ── Alignment ──
    if cell.alignment:
        if cell.alignment.horizontal:
            styles.append(f"text-align:{cell.alignment.horizontal}")
        if cell.alignment.vertical:
            v = cell.alignment.vertical
            # openpyxl uses "center" for both; CSS needs explicit mapping
            if v == "center":
                styles.append("vertical-align:middle")
            else:
                styles.append(f"vertical-align:{v}")
        # Text wrap
        if cell.alignment.wrapText:
            styles.append("white-space:normal;word-wrap:break-word")

    return ";".join(styles)


def _format_cell_value(value, computed_value=None) -> str:
    """Format a cell value for HTML display.

    Args:
        value: The raw cell value (may be a formula string).
        computed_value: The computed/cached value from data_only=True load.
            If provided and not None, it takes precedence over ``value``.
    """
    # Prefer the pre-computed (data_only) value when available
    if computed_value is not None:
        value = computed_value

    if value is None:
        return ""
    if hasattr(value, "year"):
        # date/datetime – show day number for date rows
        return str(value.day)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Format numbers consistently with Excel
        if isinstance(value, int) or value == int(value):
            return str(int(value))
        return f"{value:.2f}"
    return _html_escape(str(value))


def _find_weekday_row(ws, date_row):
    for r in range(date_row + 1, min(ws.max_row, date_row + 5) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "TEXT" in v and "aaa" in v:
                return r
    return None


def _get_print_area_bounds(ws) -> tuple[int, int, int, int] | None:
    """解析 print_area，返回 (min_col, min_row, max_col, max_row)。

    如果没有设置 print_area，返回 None（表示使用整张表）。
    """
    pa = ws.print_area
    if not pa:
        return None
    # openpyxl 可能返回字符串或列表
    if isinstance(pa, list):
        pa = pa[0]
    if "!" in pa:
        pa = pa.split("!")[1]
    return range_boundaries(pa)


def _is_row_visible(ws, row: int, bounds: tuple[int, int, int, int] | None) -> bool:
    """行是否在 print_area 内且未隐藏。"""
    if bounds:
        _, min_row, _, max_row = bounds
        if not (min_row <= row <= max_row):
            return False
    return not ws.row_dimensions[row].hidden


def _is_col_visible(ws, col: int, bounds: tuple[int, int, int, int] | None) -> bool:
    """列是否在 print_area 内且未隐藏。"""
    if bounds:
        min_col, _, max_col, _ = bounds
        if not (min_col <= col <= max_col):
            return False
    letter = col_num_to_letter(col)
    dim = ws.column_dimensions.get(letter)
    return not (dim and dim.hidden)


def _visible_rows_cols(ws, bounds: tuple[int, int, int, int] | None):
    """返回可见的行列列表和映射。

    Returns:
        visible_rows: List[int] — 可见行号（1-based，已排序）
        visible_cols: List[int] — 可见列号（1-based，已排序）
        row_map: Dict[int, int] — Excel 行号 → HTML 行索引（1-based）
        col_map: Dict[int, int] — Excel 列号 → HTML 列索引（1-based）
    """
    if bounds:
        min_col, min_row, max_col, max_row = bounds
    else:
        min_col, min_row, max_col, max_row = 1, 1, ws.max_column, ws.max_row

    visible_rows = [
        r for r in range(min_row, max_row + 1)
        if not ws.row_dimensions[r].hidden
    ]
    visible_cols = [
        c for c in range(min_col, max_col + 1)
        if not _is_col_hidden(ws, c)
    ]

    row_map = {r: i + 1 for i, r in enumerate(visible_rows)}
    col_map = {c: i + 1 for i, c in enumerate(visible_cols)}

    return visible_rows, visible_cols, row_map, col_map


def _is_col_hidden(ws, col: int) -> bool:
    """检查列是否隐藏。"""
    letter = col_num_to_letter(col)
    dim = ws.column_dimensions.get(letter)
    return bool(dim and dim.hidden)


def sheet_to_html(ws, month: str, value_ws=None) -> str:
    """将单个 worksheet 转换为 HTML，只导出 print_area 内的可见行列。

    Args:
        ws: Worksheet loaded with data_only=False (styles, formulas, merged cells).
        month: Target month (YYYY-MM) for holiday/ weekday rendering.
        value_ws: Optional worksheet loaded with data_only=True; its cell
            values (pre-computed formula results) are preferred over the
            raw formula strings in ``ws``.

    Features:
      - 只渲染 print_area 范围内的单元格
      - 跳过隐藏的行和列
      - 保留合并单元格（rowspan/colspan 根据可见行列重新计算）
      - 保留单元格样式（字体、颜色、边框、对齐）
      - 公式单元格显示计算值
    """
    year, month_num = int(month[:4]), int(month[5:7])
    days_in_month = calendar.monthrange(year, month_num)[1]
    holidays = _get_holidays_for_month(year, month_num)

    bounds = _get_print_area_bounds(ws)
    visible_rows, visible_cols, row_map, col_map = _visible_rows_cols(ws, bounds)

    if not visible_rows or not visible_cols:
        return ""

    # ── 处理合并单元格 ──
    # 只保留在 print_area 内且包含可见单元格的合并区域
    merged_map: Dict[tuple[int, int], tuple[int, int]] = {}
    merged_covered: set[tuple[int, int]] = set()

    for mr in ws.merged_cells.ranges:
        # 检查合并区域是否与 print_area 相交
        if bounds:
            min_col, min_row, max_col, max_row = bounds
            if mr.max_row < min_row or mr.min_row > max_row:
                continue
            if mr.max_col < min_col or mr.min_col > max_col:
                continue

        # 计算可见范围内的行列
        visible_in_merge_rows = [
            r for r in range(mr.min_row, mr.max_row + 1)
            if r in row_map
        ]
        visible_in_merge_cols = [
            c for c in range(mr.min_col, mr.max_col + 1)
            if c in col_map
        ]

        if not visible_in_merge_rows or not visible_in_merge_cols:
            continue

        leader_row = visible_in_merge_rows[0]
        leader_col = visible_in_merge_cols[0]
        rs = len(visible_in_merge_rows)
        cs = len(visible_in_merge_cols)

        merged_map[(leader_row, leader_col)] = (rs, cs)
        for r in visible_in_merge_rows:
            for c in visible_in_merge_cols:
                if (r, c) != (leader_row, leader_col):
                    merged_covered.add((r, c))

    date_row, day_start_col = _find_date_start(ws)
    weekday_row = _find_weekday_row(ws, date_row) if date_row else None

    html_parts = [
        "<!DOCTYPE html>",
        '<html>',
        '<head>',
        '<meta charset="utf-8">',
        f"<title>{_html_escape(ws.title)}</title>",
        "<style>",
        "body { font-family: Arial, 'Microsoft JhengHei', sans-serif; margin: 20px; }",
        "table { border-collapse: collapse; table-layout: fixed; width: 100%; }",
        "td, th { padding: 2px 4px; white-space: nowrap; overflow: hidden; box-sizing: border-box; }",
        '.holiday { color: #c00; font-size: 8pt; font-weight: bold; }',
        "</style>",
        "</head>",
        "<body>",
        f'<h2>{_html_escape(ws.title)}</h2>',
        "<table>",
    ]

    # ── Column widths ──
    for c in visible_cols:
        letter = col_num_to_letter(c)
        width = ws.column_dimensions[letter].width if letter in ws.column_dimensions else None
        if width:
            px = int(width * 7)
            html_parts.append(f'<col style="width:{px}px">')
        else:
            html_parts.append('<col>')

    for r in visible_rows:
        # Row height
        row_height = ws.row_dimensions[r].height if r in ws.row_dimensions else None
        row_style = ""
        if row_height:
            px = int(row_height * 1.33)
            row_style = f' style="height:{px}px"'
        html_parts.append(f"<tr{row_style}>")

        for c in visible_cols:
            if (r, c) in merged_covered:
                continue

            cell = ws.cell(r, c)
            cf_bg = _get_conditional_bg_color(ws, r, c)
            style = _cell_to_css(cell, conditional_bg=cf_bg)
            value = cell.value

            # Prefer pre-computed value from data_only workbook
            computed_value = None
            if value_ws:
                try:
                    computed_value = value_ws.cell(r, c).value
                except Exception:
                    pass

            if r == date_row and day_start_col and c >= day_start_col:
                day = c - day_start_col + 1
                if 1 <= day <= days_in_month:
                    val_str = str(day)
                    if day in holidays:
                        val_str += (
                            f'<br><span class="holiday">'
                            f'{_html_escape(holidays[day])}</span>'
                        )
                else:
                    val_str = _format_cell_value(value, computed_value)
            elif r == weekday_row and day_start_col and c >= day_start_col:
                day = c - day_start_col + 1
                if 1 <= day <= days_in_month:
                    val_str = WEEKDAY_CH[date(year, month_num, day).weekday()]
                else:
                    val_str = _format_cell_value(value, computed_value)
            else:
                val_str = _format_cell_value(value, computed_value)

            attrs = ""
            if (r, c) in merged_map:
                rs, cs = merged_map[(r, c)]
                if rs > 1:
                    attrs += f' rowspan="{rs}"'
                if cs > 1:
                    attrs += f' colspan="{cs}"'
            if style:
                attrs += f' style="{style}"'

            html_parts.append(f"<td{attrs}>{val_str}</td>")
        html_parts.append("</tr>")

    html_parts.append("</table>")
    html_parts.append("</body></html>")
    return "\n".join(html_parts)


# ── Main engine entry (mutates workbook in-place) ──

def _process_workbook(wb: Workbook, context: Dict[str, Any]) -> None:
    """Adjust the workbook in-place for web display.

    Steps per sheet (skips "Data" sheets):
      1. Adjust day columns to match month length
      2. Update date row to target month
      3. Detect roster regions
      4. Fetch API data for the shift
      5. Shrink regions to match actual headcount (delete from bottom)
      6. Fill retained rows with API data
      7. Normalise number formats
    """
    month = context["month"]
    year, month_num = int(month[:4]), int(month[5:7])
    days_in_month = calendar.monthrange(year, month_num)[1]

    for ws in wb.worksheets:
        if ws.title == "Data":
            continue

        # Phase 1: Adjust day columns and dates
        _adjust_day_columns(ws, days_in_month)
        _update_dates(ws, month)

        # Phase 2: Detect roster regions
        regions = _detect_roster_regions(ws)
        if not regions:
            continue

        # Phase 3: Fetch data for this shift
        shift = _shift_from_sheet_name(ws.title)
        try:
            data = _fetch_data(context, shift)
        except Exception:
            # DB / API not available — leave template data as-is
            continue

        if not data.get("has_data"):
            # No data — clear all data rows
            for region in regions:
                for r in range(region["data_start"], region["data_end"] + 1):
                    for col in range(1, ws.max_column + 1):
                        ws.cell(r, col).value = ""
            continue

        # Match API segments to regions by index
        segments = data.get("segments", [])
        api_by_index: Dict[int, Dict[str, Any]] = {}
        for i, seg in enumerate(segments):
            api_by_index[i + 1] = seg

        # Phase 4-6: Process regions bottom-up (shrink → re-detect → fill).
        for region in sorted(regions, key=lambda r: r["data_start"], reverse=True):
            idx = region["index"]
            seg = api_by_index.get(idx)
            if not seg:
                continue
            rows = seg.get("rows", [])
            target = len(rows)
            _shrink_region(ws, region, target)

            # Re-detect to get updated row numbers after this shrink
            fresh_regions = _detect_roster_regions(ws)
            fresh_region = next((r for r in fresh_regions if r["index"] == idx), None)
            if not fresh_region:
                continue

            # Phase 5: Fill data for this region
            date_row, day_start_col = _find_date_start(ws)
            for i, row_data in enumerate(rows):
                row = fresh_region["data_start"] + i
                _fill_data_row(ws, row, row_data, day_start_col, days_in_month)

        # Phase 6: Adjust number formats
        date_row, day_start_col = _find_date_start(ws)
        if day_start_col:
            _adjust_number_formats(ws, day_start_col, days_in_month)


# ── Public entry point (registered in HTML_ENGINES) ──

def run(
    shift: str,
    project_id: Optional[int] = None,
    category_id: Optional[int] = None,
    month: Optional[str] = None,
) -> Dict[str, Any]:
    """Load template, process workbook, return HTML.

    Args:
        shift: Ignored — web_roster processes all sheets in the workbook.
        project_id: Project ID for template lookup.
        category_id: Category ID for template lookup.
        month: Target month (YYYY-MM).

    Returns:
        {"html": <HTML string>, "month": month, "template_path": path}
    """
    effective_month = month or "2026-03"

    # Try to resolve template from DB
    excel_path, rel_path = resolve_template(
        project_id, category_id, effective_month, "cleaning_shortfall_v2"
    )
    if not excel_path:
        excel_path, rel_path = resolve_template(
            project_id, category_id, effective_month, "roster_shortfall"
        )

    # Fallback: local templates directory (for standalone testing)
    if not excel_path:
        tpl_dir = Path(__file__).resolve().parent.parent.parent / "engine" / "templates"
        candidates = [
            tpl_dir / "TY-2026.03- SG_SEC-Deploy Roster  Shortfall - template.xlsx",
            tpl_dir / "东汇-保安轮休表template.xlsx",
            tpl_dir / "东汇-保洁轮休表_template.xlsx",
        ]
        for c in candidates:
            if c.exists():
                excel_path = c
                rel_path = str(c.name)
                break

    if not excel_path or not excel_path.exists():
        return {
            "html": "<p>No template found</p>",
            "month": effective_month,
            "error": "template_not_found",
        }

    # Load workbook twice:
    #   - data_only=False  → styles, formulas, merged cells, structure
    #   - data_only=True   → cached formula values (computed by Excel)
    wb = load_workbook(str(excel_path), data_only=False, keep_links=False)
    wb_values = load_workbook(str(excel_path), data_only=True, keep_links=False)

    context = {
        "project_id": project_id,
        "category_id": category_id,
        "month": effective_month,
    }

    _process_workbook(wb, context)

    # Keep wb_values' column structure in sync with wb so that cell
    # coordinates line up when we read computed values in workbook_to_html.
    year, month_num = int(effective_month[:4]), int(effective_month[5:7])
    days_in_month = calendar.monthrange(year, month_num)[1]
    for vws in wb_values.worksheets:
        if vws.title == "Data":
            continue
        _adjust_day_columns(vws, days_in_month)
        _update_dates(vws, effective_month)

    # 每个 sheet 生成独立的 HTML
    sheets_html: Dict[str, str] = {}
    for ws in wb.worksheets:
        if ws.title == "Data":
            continue
        value_ws = None
        if wb_values:
            for vws in wb_values.worksheets:
                if vws.title == ws.title:
                    value_ws = vws
                    break
        sheets_html[ws.title] = sheet_to_html(ws, effective_month, value_ws=value_ws)

    return {
        "sheets": sheets_html,
        "month": effective_month,
        "template_path": str(rel_path) if rel_path else "",
        "_wb": wb,           # internal: processed workbook for testing
        "_wb_values": wb_values,  # internal: value workbook for testing
    }
