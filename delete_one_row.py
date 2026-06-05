#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Delete the 2nd data row from the m-th roster region in an Excel template.

Usage:
    python delete_one_row.py <input.xlsx> <output.xlsx> <region_index>

Example:
    python delete_one_row.py template.xlsx output.xlsx 2
    # Deletes the 2nd data row from region 2
"""

import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────────────────────────────────────
# Region Detection
# ──────────────────────────────────────────────────────────────────────────────

def detect_roster_regions(ws):
    """
    Detect all roster (personnel list) regions in a worksheet.

    A region is defined as:
      - Title row: B-column starts with "N. " (e.g. "1. 保安主管出勤時數")
      - Data rows: B-column contains a rank code like M01, A1, D01, etc.
      - Ends at the first non-data row below the data block.

    Returns a list of dicts:
        [{
            'index': 1-based region index,
            'title_row': row number of the title,
            'data_start': first data row,
            'data_end':   last data row,
            'title':      title text,
        }, ...]
    """
    regions = []
    rank_pattern = re.compile(r'^[A-Z]+\d+$')
    title_pattern = re.compile(r'^\d+\.\s+')

    for row in range(1, ws.max_row + 1):
        b_val = ws.cell(row=row, column=2).value
        if b_val and isinstance(b_val, str) and title_pattern.match(b_val):
            title_row = row

            # Find first data row (rank code in column B)
            data_start = None
            for r in range(title_row + 1, ws.max_row + 1):
                b = ws.cell(row=r, column=2).value
                if b and isinstance(b, str) and rank_pattern.match(b.strip()):
                    data_start = r
                    break

            if data_start is None:
                continue

            # Find data end: last consecutive row with a rank code in column B
            data_end = data_start
            for r in range(data_start + 1, ws.max_row + 1):
                b = ws.cell(row=r, column=2).value
                is_data = b and isinstance(b, str) and rank_pattern.match(b.strip())
                if is_data:
                    data_end = r
                else:
                    break

            regions.append({
                'index': len(regions) + 1,
                'title_row': title_row,
                'data_start': data_start,
                'data_end': data_end,
                'title': b_val,
            })

    return regions


# ──────────────────────────────────────────────────────────────────────────────
# Formula Adjustment (handles $A$1, $A1, A$1, A1, and ranges)
# ──────────────────────────────────────────────────────────────────────────────

def adjust_formula(formula, deleted_row):
    """
    Adjust all cell references in an Excel formula after a row is deleted.

    Rules (matching Excel behavior):
      - Relative row references (A5, A5:B10):  row > deleted_row → row-1
      - Absolute row references ($A$5, A$5):   row unchanged
      - Mixed refs ($A5):                      row relative → adjusted like A5
    """
    # Match a cell reference part: $?[A-Z]{1,3}$?\d+
    # Group 1 = column part (e.g. "A", "$A", "AM", "$AM")
    # Group 2 = "$" if row is absolute, else ""
    # Group 3 = row number
    cell_ref = re.compile(r'(?<![A-Z$])(\$?[A-Z]{1,3})(\$?)(\d+)')

    result = []
    last_end = 0

    for match in cell_ref.finditer(formula):
        result.append(formula[last_end:match.start()])

        col_part = match.group(1)      # e.g. "A", "$A"
        abs_row = match.group(2) == '$'
        row_num = int(match.group(3))

        if not abs_row and row_num > deleted_row:
            row_num -= 1

        result.append(f"{col_part}{'$' if abs_row else ''}{row_num}")
        last_end = match.end()

    result.append(formula[last_end:])
    return ''.join(result)


def adjust_merged_range(range_str, deleted_row):
    """
    Adjust row numbers in a merged cell range string after deleting a row.
    Example: 'B50:C50' with deleted_row=23 → 'B49:C49'
    """
    ref_pattern = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')

    def parse_ref(ref):
        m = ref_pattern.match(ref)
        if m:
            col_abs, col, row_abs, row = m.groups()
            row = int(row)
            if row > deleted_row:
                row -= 1
            return f"{col_abs}{col}{row_abs}{row}"
        return ref

    if ':' in range_str:
        start, end = range_str.split(':')
        return f"{parse_ref(start)}:{parse_ref(end)}"
    return parse_ref(range_str)


def safe_delete_rows(ws, idx, amount=1):
    """
    Delete rows, correctly mimicking Excel behavior.
    Replaces openpyxl's ws.delete_rows() which has an iter_rows bug that
    creates phantom empty cells which then overwrite valid data during shift.
    """
    max_row = max((row for row, col in ws._cells.keys()), default=0)

    # Process each row from top to bottom
    for r in range(idx, max_row + 1):
        source_r = r + amount
        if source_r > max_row:
            # No source row, just clear this row
            for key in list(ws._cells.keys()):
                if key[0] == r:
                    del ws._cells[key]
            continue

        # Get columns in target and source
        target_cols = {col for row, col in ws._cells.keys() if row == r}
        source_cols = {col for row, col in ws._cells.keys() if row == source_r}

        # Delete cells in target that don't exist in source
        for col in target_cols - source_cols:
            del ws._cells[(r, col)]

        # Move cells from source to target
        for col in source_cols:
            if (source_r, col) in ws._cells:
                cell = ws._cells[(source_r, col)]
                ws._cells[(r, col)] = cell
                cell.row = r
                del ws._cells[(source_r, col)]

    # Update row dimensions (height + style + other attrs)
    for r in range(idx, max_row + 1):
        source_r = r + amount
        dst_rd = ws.row_dimensions[r]
        if source_r in ws.row_dimensions:
            src_rd = ws.row_dimensions[source_r]
            dst_rd.height = src_rd.height
            if src_rd._style is not None:
                from copy import copy
                dst_rd._style = copy(src_rd._style)
            else:
                dst_rd._style = None
            dst_rd.hidden = src_rd.hidden
            dst_rd.outline_level = src_rd.outline_level
            dst_rd.collapsed = src_rd.collapsed
        else:
            dst_rd.height = None
            dst_rd._style = None
            dst_rd.hidden = False
            dst_rd.outline_level = 0
            dst_rd.collapsed = False

    # Update current_row
    ws._current_row = max(row for row, col in ws._cells.keys()) if ws._cells else 0


def adjust_print_area(ws, deleted_row):
    """Adjust print-area row numbers after a row is deleted."""
    if not ws.print_area or not ws._print_area:
        return
    for cr in ws._print_area.ranges:
        if cr.min_row > deleted_row:
            cr.min_row -= 1
        if cr.max_row > deleted_row:
            cr.max_row -= 1


def fix_merged_cells(ws, deleted_row):
    """
    Adjust merged cell ranges after a row deletion WITHOUT unmerge+merge.

    openpyxl's unmerge_cells + merge_cells has a side effect of copying
    the top-left cell's style to ALL merged cells, which corrupts styles
    (e.g. fills backgrounds white or overwrites custom colours).
    We directly mutate MergedCellRange objects in place.
    """
    to_remove = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row > deleted_row:
            # Entirely below the deleted row — shift up by 1
            mr.shift(-1, 0)
        elif mr.max_row < deleted_row:
            # Entirely above — no change
            pass
        else:
            # deleted_row is inside or at the boundary of this range
            if mr.min_row == mr.max_row == deleted_row:
                # Single-row range that is completely deleted
                to_remove.append(mr)
            else:
                # Reduce the range by 1 row at the bottom
                mr.max_row -= 1
                if mr.min_row > mr.max_row:
                    to_remove.append(mr)

    for mr in to_remove:
        ws.merged_cells.ranges.remove(mr)


def adjust_all_formulas(ws, deleted_row):
    """Scan entire worksheet and adjust every formula after a row deletion."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = cell.value
            if val and isinstance(val, str) and val.startswith('='):
                cell.value = adjust_formula(val, deleted_row)


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

def process_workbook(input_path, output_path, region_index):
    wb = openpyxl.load_workbook(input_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        regions = detect_roster_regions(ws)

        print(f"\n📋 Sheet '{sheet_name}' — detected {len(regions)} roster region(s):")
        for r in regions:
            print(f"   Region {r['index']}: rows {r['data_start']}-{r['data_end']} "
                  f"({r['data_end'] - r['data_start'] + 1} rows) — {r['title']}")

        region = next((r for r in regions if r['index'] == region_index), None)
        if region is None:
            print(f"   ⚠️  Region {region_index} not found, skipping.")
            continue

        row_to_delete = region['data_start'] + 1
        if row_to_delete > region['data_end']:
            print(f"   ⚠️  Region {region_index} has only 1 row, nothing to delete.")
            continue

        print(f"   🗑️  Deleting row {row_to_delete} (2nd data row of region {region_index})")

        # 1. Adjust formulas
        adjust_all_formulas(ws, row_to_delete)

        # 2. Physically remove the row (safe replacement for ws.delete_rows)
        safe_delete_rows(ws, row_to_delete, 1)

        # 3. Fix merged cells (direct mutation — no unmerge/merge side effects)
        fix_merged_cells(ws, row_to_delete)

        # 4. Adjust print area
        adjust_print_area(ws, row_to_delete)

        print(f"   ✅ Done. Original row {row_to_delete} deleted.")

    wb.save(output_path)
    print(f"\n💾 Saved to: {output_path}")
    return wb


def main():
    if len(sys.argv) >= 4:
        input_path = sys.argv[1]
        output_path = sys.argv[2]
        region_index = int(sys.argv[3])
    else:
        # Default test
        input_path = "templates/东汇-保洁轮休表_template.xlsx"
        output_path = "test_outputs_delete_rows/东汇-保洁_区域2删第2行.xlsx"
        region_index = 2
        print("No arguments provided; running default test case.")
        print(f"  Input:  {input_path}")
        print(f"  Output: {output_path}")
        print(f"  Region: {region_index}")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    process_workbook(input_path, output_path, region_index)


if __name__ == '__main__':
    main()
