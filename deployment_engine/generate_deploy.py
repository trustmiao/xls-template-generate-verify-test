# -*- coding: utf-8 -*-
"""Generate static deploy.xlsx from roster + salary data.

Uses excel_utils row-segment APIs to safely expand / shrink the template data
area while preserving merged cells, print area and conditional formatting.
"""
import sys
import os

# Allow import of excel_utils from sibling directory
if __name__ == "__main__":
    _engine_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _engine_dir not in sys.path:
        sys.path.insert(0, _engine_dir)

import openpyxl
from openpyxl import load_workbook

from excel_utils import (
    expand_segment_for_count,
    shrink_segment_for_count,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# 1. 从花名册提取人员 + 合计工时
# ============================================================
roster_wb = load_workbook(
    os.path.join(BASE_DIR, "东汇-保洁轮休表.xlsx"), data_only=True
)
roster_ws = roster_wb.active

hours_by_name = {}
id_by_name = {}

rows_to_check = list(range(10, 12)) + list(range(18, 35)) + list(range(41, 44))
for row in rows_to_check:
    id_val = roster_ws.cell(row, 2).value   # col B = 排位
    name   = roster_ws.cell(row, 3).value   # col C = 姓名
    hours  = roster_ws.cell(row, 36).value  # col AJ = 每月工作時數
    if name and id_val:
        if name not in hours_by_name:
            hours_by_name[name] = 0
            id_by_name[name] = str(id_val)
        hours_by_name[name] += float(hours) if hours else 0


def sort_key(n):
    s = id_by_name[n]
    if s.startswith('A'):
        try:
            return int(s[1:])
        except ValueError:
            pass
    return 99


sorted_cleaners = sorted(hours_by_name.keys(), key=sort_key)
print(f"花名册人数: {len(sorted_cleaners)}")

# ============================================================
# 2. 从薪俸結算書提取薪资数据 (2026.04, TUW dept)
# ============================================================
salary_wb = load_workbook(
    os.path.join(BASE_DIR, "薪俸結算書 (大元沙田坳Master).xlsx"), data_only=True
)
salary_ws = salary_wb["Salary Detail Report (2026.04)"]

salary_data = {}
for row in salary_ws.iter_rows(min_row=3, values_only=True):
    ch = row[5]   # col F = Chinese Name
    if ch in sorted_cleaners:
        salary_data[ch] = {
            'english':  row[4],   # col E
            'position': row[6],   # col G
            'mwage':    row[15],  # col P  每月工資
            'sh':       row[28],  # col AC 法定假日薪酬
            'grat':     row[32],  # col AG 酬金
            'typh':     row[33],  # col AH 颱風薪酬
            'total':    row[52],  # col BA 薪金結算書總計
        }

not_found = [n for n in sorted_cleaners if n not in salary_data]
if not_found:
    print(f"警告: 薪资未找到: {not_found}")

# ============================================================
# 3. 加载模板，动态调整数据区行数，写入静态数据
# ============================================================
template_wb = load_workbook(
    os.path.join(BASE_DIR, "Deploy Roster  Shortfall.xlsx")
)
ws = template_wb["Deployment"]

DATA_START = 12      # 第一数据行（模板格式行）
TPL_COUNT = 113      # 模板原始数据行数 (12-124)
needed = len(sorted_cleaners)

# 动态调整数据区行数（expand/shrink 会保护 merged cell / print area / CF）
if needed > TPL_COUNT:
    expand_segment_for_count(ws, DATA_START, TPL_COUNT, needed)
elif needed < TPL_COUNT:
    shrink_segment_for_count(ws, DATA_START, TPL_COUNT, needed)

new_last = DATA_START + needed - 1

# 写入保洁数据，全部使用 M 序列
total_hours = 0
for i, name in enumerate(sorted_cleaners):
    r = DATA_START + i
    sal = salary_data.get(name, {})

    hours      = hours_by_name[name]
    total_wage = sal.get('total') or 0
    mwage      = sal.get('mwage') or 0
    sh         = sal.get('sh')    or 0
    typh       = sal.get('typh')  or 0
    grat       = sal.get('grat')  or 0
    eng        = sal.get('english', '')
    position   = sal.get('position', 'CLEANER')
    avg_hourly = round(total_wage / hours, 2) if (total_wage and hours) else None
    total_hours += hours

    # 覆盖写入静态值（原有 VLOOKUP/INDEX 公式被替换）
    ws.cell(r, 1).value  = 'TUW'
    ws.cell(r, 2).value  = f"M{i+1:02d}"
    ws.cell(r, 4).value  = position
    ws.cell(r, 5).value  = eng
    ws.cell(r, 6).value  = name
    ws.cell(r, 7).value  = mwage
    ws.cell(r, 8).value  = sh
    ws.cell(r, 9).value  = typh
    ws.cell(r, 10).value = grat
    ws.cell(r, 11).value = hours
    ws.cell(r, 12).value = total_wage
    ws.cell(r, 13).value = avg_hourly

    print(f"  M{i+1:02d} {name:<6} {position:<10} {hours}h  工资={total_wage}")

# 更新合计工时公式（row 在数据区下方两行：空行 + 合计行）
total_row = new_last + 2
ws.cell(total_row, 11).value = f"=SUM(K{DATA_START}:K{new_last})"

# ============================================================
# 4. 保存
# ============================================================
output_path = os.path.join(BASE_DIR, "deploy.xlsx")
template_wb.save(output_path)
print(f"\n已生成: {output_path}")
print(f"总工时: {total_hours}h，共 {len(sorted_cleaners)} 人")
