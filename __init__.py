"""Unified engine execution entrypoint.

  execute_html()  — run an HTML engine (returns JSON)
  execute_excel() — run an Excel engine (returns file metadata)
"""
from __future__ import annotations

import json
import shutil
import uuid
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import load_workbook

from ..config import DOC_DIR, OUTPUT_DIR
from ..models import create_generated_file, get_generated_file, get_output_engine
from ..utils.excel_cf import trim_cf_to_print_area
from ..utils.excel_verify import summarize_for_metadata, verify_generated_xlsx
from .common.data_source import build_context, get_engine_config, resolve_template_for_engine
from .registry import EXCEL_ENGINES, HTML_ENGINES


def execute_html(
    engine_name: str,
    shift: str,
    project_id: Optional[int] = None,
    category_id: Optional[int] = None,
    month: Optional[str] = None,
) -> Dict[str, Any]:
    """Run an HTML engine and return frontend JSON."""
    engine = HTML_ENGINES.get(engine_name)
    if not engine:
        raise ValueError(f"Unknown HTML engine: {engine_name}")
    return engine(shift=shift, project_id=project_id, category_id=category_id, month=month)


def execute_excel(
    engine_id: int,
    month: str,
    created_by: Optional[str] = None,
) -> Dict[str, Any]:
    """Run an Excel engine and return generated file metadata.

    Steps (mirrors the old output_engines.execute_engine):
      1. Load engine config + template
      2. Build data context
      3. Run engine (builtin or json_mapping)
      4. Save to OUTPUT_DIR
      5. Record in generated_files table
    """
    engine_config = get_engine_config(engine_id)
    if not engine_config:
        raise ValueError(f"Engine {engine_id} not found")

    project_id = engine_config["project_id"]
    category_id = engine_config["category_id"]
    engine_type = engine_config.get("engine_type")
    builtin_key = engine_config.get("builtin_key")

    # Resolve template
    template_path, rel_path = resolve_template_for_engine(engine_config, month)
    if not template_path:
        raise FileNotFoundError(
            f"No template found for engine {engine_id} (month={month})"
        )

    wb = load_workbook(str(template_path), data_only=False)

    # Build context
    context = build_context(project_id, category_id, month)
    context["template_path"] = str(template_path)

    # Execute
    if engine_type == "builtin":
        if not builtin_key or builtin_key not in EXCEL_ENGINES:
            raise ValueError(f"Unknown builtin engine: {builtin_key}")
        EXCEL_ENGINES[builtin_key](wb, context)
    elif engine_type == "json_mapping":
        from .excel import json_mapper_apply
        rules = json.loads(engine_config.get("json_rules") or "{}")
        json_mapper_apply(wb, context, rules)
    else:
        raise ValueError(f"Unknown engine_type: {engine_type}")

    # Save
    file_uuid = str(uuid.uuid4())
    safe_name = f"{engine_config['name']}_{month}_{file_uuid[:8]}.xlsx"
    safe_name = safe_name.replace(" ", "_")
    out_path = OUTPUT_DIR / safe_name
    tmp_path = Path("/tmp") / safe_name
    wb.save(str(tmp_path))

    try:
        trim_cf_to_print_area(tmp_path)
    except Exception:
        pass

    shutil.move(str(tmp_path), str(out_path))

    file_size = out_path.stat().st_size
    rel_path = str(out_path.relative_to(OUTPUT_DIR)).replace("\\", "/")

    try:
        xlsx_checks = verify_generated_xlsx(out_path)
    except Exception as e:
        xlsx_checks = {
            "ok": False, "summary": {}, "details": {},
            "error": f"verify failed: {type(e).__name__}: {e}",
        }

    gen = create_generated_file(
        engine_id=engine_id,
        project_id=project_id,
        category_id=category_id,
        month=month,
        file_name=safe_name,
        file_path=rel_path,
        file_size=file_size,
        created_by=created_by,
        metadata_json=json.dumps({
            "engine_name": engine_config["name"],
            "engine_type": engine_config["engine_type"],
            "xlsx_checks": summarize_for_metadata(xlsx_checks),
        }),
    )

    return {
        "file_id": gen["id"],
        "file_name": safe_name,
        "download_url": f"/api/output-files/{gen['id']}/download",
        "preview_url": f"/api/output-files/{gen['id']}/preview",
        "file_size": file_size,
        "xlsx_checks": xlsx_checks,
    }


# Preview helpers (re-exported by output_engines for backward compat)

def preview_file(file_id: int) -> str:
    """Return an HTML table preview of a generated Excel file."""
    from openpyxl import load_workbook as _load_workbook
    gen = get_generated_file(file_id)
    if not gen:
        raise FileNotFoundError(f"Generated file {file_id} not found")
    file_path = OUTPUT_DIR / gen["file_path"]
    if not file_path.exists():
        raise FileNotFoundError(f"File not found on disk: {file_path}")
    wb = _load_workbook(str(file_path), data_only=True)
    ws = wb.worksheets[0]
    html = "<html><head><meta charset='utf-8'><style>"
    html += "table{border-collapse:collapse;font-family:Arial,sans-serif;font-size:12px;width:100%;}"
    html += "td,th{border:1px solid #ccc;padding:4px 8px;text-align:center;}"
    html += "th{background:#f0f0f0;}"
    html += "tr:nth-child(even){background:#fafafa;}"
    html += "@media print{body{margin:0;}table{width:100%;}td,th{font-size:10px;padding:2px 4px;}"
    html += ".no-print{display:none !important;}}"
    html += "</style></head><body>"
    html += "<div class='no-print' style='padding:8px;background:#f8fafc;border-bottom:1px solid #e2e8f0;margin-bottom:8px;'>"
    html += f"<button onclick='window.print()' style='padding:6px 12px;cursor:pointer;'>打印</button>"
    html += f" <span style='margin-left:12px;color:#64748b;'>{gen['file_name']}</span>"
    html += "</div>"
    html += "<table>"
    for r in range(1, ws.max_row + 1):
        html += "<tr>"
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            cell_html = str(val) if val is not None else ""
            is_th = r <= 2 or (c <= 3 and val)
            tag = "th" if is_th else "td"
            html += f"<{tag}>{cell_html}</{tag}>"
        html += "</tr>"
    html += "</table></body></html>"
    return html


def get_generated_file_path(file_id: int) -> Optional[Path]:
    gen = get_generated_file(file_id)
    if not gen:
        return None
    p = OUTPUT_DIR / gen["file_path"]
    return p if p.exists() else None


# Re-exports for backward compat
from .registry import BUILTIN_ENGINES, EXCEL_ENGINES, HTML_ENGINES

__all__ = [
    "execute_html",
    "execute_excel",
    "BUILTIN_ENGINES",
    "EXCEL_ENGINES",
    "HTML_ENGINES",
]
