#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Detailed verification of issue #6 outputs after refactoring.

Checks the 大元保安 2026-02 output for:
  - Row heights
  - No orphaned cells / phantom backgrounds
  - Merged cell counts and ranges
  - Conditional formatting counts and key ranges
  - Print area
"""

from pathlib import Path
from openpyxl import load_workbook

OUTPUT = Path("test_outputs_issue6") / "2026-02_TY-大元保安_adjusted.xlsx"


def main():
    wb = load_workbook(OUTPUT, data_only=False)
    ws = wb["早"]

    print("=" * 60)
    print(f"Verification: {OUTPUT.name} → sheet '{ws.title}'")
    print("=" * 60)

    # 1. Row heights (key rows)
    print("\n📏 Row heights (key rows):")
    for r in [12, 13, 30, 31, 32, 33]:
        h = ws.row_dimensions[r].height
        print(f"   Row {r:2d}: height = {h}")

    # 2. Check row 31 for orphaned cells
    print("\n🧹 Row 31 (should be empty after deletions):")
    row31_cells = [(c, ws.cell(31, c).value) for c in range(1, ws.max_column + 1)
                   if ws.cell(31, c).value is not None]
    if row31_cells:
        print(f"   ⚠️  Found {len(row31_cells)} non-empty cells: {row31_cells[:10]}...")
    else:
        print("   ✅ Row 31 is completely empty")

    # 3. Check I32 / I33 content
    print("\n📝 I32 / I33 (should have shifted-up text):")
    print(f"   I32 = {ws['I32'].value!r}")
    print(f"   I33 = {ws['I33'].value!r}")

    # 4. Merged cells
    print(f"\n🔗 Merged cells: {len(ws.merged_cells.ranges)} ranges")
    for mr in sorted(ws.merged_cells.ranges, key=lambda r: (r.min_row, r.min_col)):
        print(f"   {mr}")

    # 5. Conditional formatting
    print(f"\n🎨 Conditional formatting: {len(ws.conditional_formatting)} rules")
    for cf in ws.conditional_formatting:
        cells_str = str(cf.cells)
        rule_types = [r.type for r in cf.rules]
        print(f"   {cells_str} → types={rule_types}")

    # 6. Print area
    print(f"\n📄 Print area: {ws.print_area}")

    # 7. Check specific conditional formatting on I33
    print("\n🔍 I33 conditional formatting:")
    i33_cf = []
    for cf in ws.conditional_formatting:
        for cr in cf.cells:
            if cr.min_row <= 33 <= cr.max_row and cr.min_col <= 9 <= cr.max_col:
                i33_cf.append((str(cf.cells), [r.type for r in cf.rules]))
    if i33_cf:
        for cells, types in i33_cf:
            print(f"   {cells} → {types}")
    else:
        print("   (none)")

    # 8. Quick sanity: max_row should be ~33 (was 73, deleted 40 rows)
    print(f"\n📊 Max row: {ws.max_row}")

    print("\n" + "=" * 60)
    print("Verification complete")
    print("=" * 60)


if __name__ == "__main__":
    main()
