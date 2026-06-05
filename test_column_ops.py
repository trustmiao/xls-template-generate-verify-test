#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test suite for excel_column_ops.py — safe column deletion.

Covers:
  - Phantom-cell avoidance (no empty cells overwriting valid data)
  - Formula adjustment (all reference styles)
  - Merged-cell range adjustment
  - Print-area adjustment
  - Conditional-formatting adjustment
  - Column-width copying
"""

from openpyxl import Workbook
from openpyxl.formatting.formatting import ConditionalFormatting
from openpyxl.styles import PatternFill
from openpyxl.worksheet.cell_range import MultiCellRange

from excel_column_ops import (
    adjust_formula,
    safe_delete_cols,
    fix_merged_cells,
    adjust_print_area,
    adjust_conditional_formatting,
    delete_col_and_fixup,
)


def test_adjust_formula():
    """Verify formula adjustment for all reference styles."""
    cases = [
        # (input, deleted_col, expected)
        ("=A1", 1, "=#REF!"),           # col A deleted
        ("=A1", 2, "=A1"),               # col B deleted, A unchanged
        ("=B1", 1, "=A1"),               # col A deleted, B→A
        ("=$A$1", 1, "=$A$1"),           # absolute col — unchanged
        ("=$A1", 1, "=$A1"),             # absolute col — unchanged
        ("=A$1", 1, "=#REF!"),           # relative col deleted
        ("=SUM(A1:C1)", 2, "=SUM(A1:B1)"),   # range shrinks
        ("=SUM(A1:A5)", 1, "=SUM(#REF!:#REF!)"),  # single-col range, col num becomes 0
        ("=SUM(B1:D1)+A1", 1, "=SUM(A1:C1)+#REF!"),
        ("=SUM(B1:D1)+E1", 2, "=SUM(A1:C1)+D1"),
    ]
    for formula, deleted_col, expected in cases:
        result = adjust_formula(formula, deleted_col)
        assert result == expected, f"adjust_formula({formula!r}, {deleted_col}) = {result!r}, expected {expected!r}"
    print("✅ adjust_formula: all cases pass")


def test_safe_delete_cols_no_phantom():
    """
    Verify safe_delete_cols does NOT create phantom cells.

    Scenario:
      - Row 1 has data in columns A and C (B is empty)
      - Row 2 has data in columns A, B, C
      - Delete column B
      - Row 1, col C should NOT be overwritten by a phantom empty cell
    """
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "A1-data"
    ws["C1"] = "C1-data"
    ws["A2"] = "A2-data"
    ws["B2"] = "B2-data"
    ws["C2"] = "C2-data"

    safe_delete_cols(ws, 2, 1)   # delete column B

    # After deletion:
    # A1 should keep A1-data
    # B1 should have C1-data (shifted left from C1)
    # A2 should keep A2-data
    # B2 should have C2-data (shifted left from C2)
    assert ws["A1"].value == "A1-data", f"A1 = {ws['A1'].value!r}"
    assert ws["B1"].value == "C1-data", f"B1 = {ws['B1'].value!r}"
    assert ws["A2"].value == "A2-data", f"A2 = {ws['A2'].value!r}"
    assert ws["B2"].value == "C2-data", f"B2 = {ws['B2'].value!r}"
    print("✅ safe_delete_cols: no phantom cells")


def test_safe_delete_cols_style_preserved():
    """Verify cell styles (fill, border) move with the cell."""
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws["B2"] = "styled"
    ws["B2"].font = Font(bold=True)
    ws["B2"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    thin = Side(style="thin")
    ws["B2"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

    safe_delete_cols(ws, 1, 1)   # delete column A; B2 → A2

    cell = ws["A2"]
    assert cell.value == "styled"
    assert cell.font.bold is True
    assert cell.fill.start_color.rgb == "00FF0000"
    assert cell.border.left.style == "thin"
    print("✅ safe_delete_cols: cell style preserved after shift")


def test_fix_merged_cells():
    """Verify merged-cell ranges adjust correctly after column deletion."""
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("B2:D4")   # cols 2-4
    ws.merge_cells("E2:E4")   # col 5 only
    ws.merge_cells("A1:A3")   # col 1 only, should be untouched

    fix_merged_cells(ws, 3)   # delete column C

    ranges = {str(mr) for mr in ws.merged_cells.ranges}
    # B2:D4 → B2:C4 (cols 2-3)
    # E2:E4 → D2:D4 (col 5 → col 4)
    # A1:A3 unchanged
    assert "B2:C4" in ranges, f"Expected B2:C4 in {ranges}"
    assert "D2:D4" in ranges, f"Expected D2:D4 in {ranges}"
    assert "A1:A3" in ranges, f"Expected A1:A3 in {ranges}"
    print("✅ fix_merged_cells: ranges adjusted correctly")


def test_adjust_print_area():
    """Verify print area shrinks after column deletion."""
    wb = Workbook()
    ws = wb.active
    ws.print_area = "A1:D10"

    adjust_print_area(ws, 3)   # delete column C

    pa = ws.print_area
    assert "$A$1:$C$10" in pa, f"Print area = {pa!r}, expected '$A$1:$C$10'"
    print("✅ adjust_print_area: correctly shrunk")


def test_adjust_conditional_formatting():
    """Verify conditional formatting ranges adjust after column deletion."""
    wb = Workbook()
    ws = wb.active

    # Add a conditional formatting rule covering B2:D4
    cf = ConditionalFormatting()
    cf.cells = MultiCellRange("B2:D4")
    cf.rules = []
    ws.conditional_formatting._cf_rules[cf] = []

    adjust_conditional_formatting(ws, 2)   # delete column B

    ranges = [str(cf.cells) for cf in ws.conditional_formatting]
    assert "B2:C4" in ranges, f"Expected B2:C4 in {ranges}"
    print("✅ adjust_conditional_formatting: ranges shrunk correctly")


def test_column_width_copied():
    """Verify column width is copied leftward after deletion."""
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions["B"].width = 25.5
    ws.column_dimensions["C"].width = 15.0

    safe_delete_cols(ws, 1, 1)   # delete column A; B→A, C→B

    assert ws.column_dimensions["A"].width == 25.5, f"A width = {ws.column_dimensions['A'].width}"
    assert ws.column_dimensions["B"].width == 15.0, f"B width = {ws.column_dimensions['B'].width}"
    print("✅ safe_delete_cols: column widths copied")


def test_delete_col_and_fixup_integration():
    """End-to-end test of the high-level helper."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "header"
    ws["B1"] = "=A1"
    ws["C1"] = "=B1"
    ws.merge_cells("A2:C2")
    ws.print_area = "A1:C5"

    delete_col_and_fixup(ws, 2)   # delete column B

    assert ws["A1"].value == "header"
    # C1 (=B1) shifts left to B1; B reference == deleted_col → #REF!
    assert ws["B1"].value == "=#REF!", f"B1 formula = {ws['B1'].value!r}"
    # C1 is now empty (was D1)
    assert ws["C1"].value is None, f"C1 formula = {ws['C1'].value!r}"
    ranges = {str(mr) for mr in ws.merged_cells.ranges}
    assert "A2:B2" in ranges, f"Expected A2:B2 in {ranges}"
    assert "$A$1:$B$5" in ws.print_area, f"Print area = {ws.print_area!r}"
    print("✅ delete_col_and_fixup: integration test pass")


def main():
    print("=" * 60)
    print("excel_column_ops test suite")
    print("=" * 60)
    test_adjust_formula()
    test_safe_delete_cols_no_phantom()
    test_safe_delete_cols_style_preserved()
    test_fix_merged_cells()
    test_adjust_print_area()
    test_adjust_conditional_formatting()
    test_column_width_copied()
    test_delete_col_and_fixup_integration()
    print("=" * 60)
    print("All tests passed ✅")
    print("=" * 60)


if __name__ == "__main__":
    main()
