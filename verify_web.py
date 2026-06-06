#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""逐格验证：HTML 输出 vs Excel 模板 (Issue #5)

对 web_roster 引擎生成的 HTML 进行逐单元格对比验证：
1. 打印区域 — 只检查 print_area 内的单元格
2. 隐藏行列 — 验证隐藏行列在 HTML 中不出现
3. 值 — 检查每个单元格的值是否相等（公式显示计算值）
4. 背景颜色 — 检查背景色是否一致
5. 宽高 — 检查单元格宽度高度是否一致
6. 边框 — 检查四周边框是否一致
7. 合并单元格 — rowspan/colspan 是否正确
8. 打印区域长宽-隐藏行列 与网页长宽一致

使用 东汇-保洁轮休表_template.xlsx 测试导出。
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter

sys.path.insert(0, r"D:\claude\claude_hk\backend")

from app.engine.html.web_roster import (
    run,
    _color_to_hex,
    _get_conditional_bg_color,
    _get_print_area_bounds,
    _find_date_start,
    _find_weekday_row,
    col_num_to_letter,
)
from app.engine.common.shift_info import WEEKDAY_CH

import calendar
from datetime import date


# ── HTML parsing ──

def parse_html_table(html: str) -> dict:
    """Parse single-sheet HTML table into structured dict.

    Returns:
        {
            "cells": {(row, col): {"value": str, "style": str, "rowspan": int, "colspan": int}},
            "max_row": int,
            "max_col": int,
            "col_widths": {col_index: px},
            "row_heights": {row_index: px},
        }
    """
    cells = {}
    col_widths = {}
    row_heights = {}

    # Parse <col> widths
    col_idx = 1
    for m in re.finditer(r'<col([^>]*)>', html):
        attrs = m.group(1)
        style_m = re.search(r'style="([^"]*)"', attrs)
        if style_m:
            w_m = re.search(r'width:(\d+)px', style_m.group(1))
            if w_m:
                col_widths[col_idx] = int(w_m.group(1))
        col_idx += 1

    # Parse rows — track rowspan so cells line up correctly
    row = 0
    rowspan_tracker: dict[int, int] = {}  # col -> remaining rows

    for tr_m in re.finditer(r'<tr([^>]*)>(.*?)</tr>', html, re.DOTALL):
        row += 1
        tr_attrs = tr_m.group(1)
        tr_html = tr_m.group(2)

        h_m = re.search(r'height:(\d+)px', tr_attrs)
        if h_m:
            row_heights[row] = int(h_m.group(1))

        # Decrement rowspan tracker
        rowspan_tracker = {c: r - 1 for c, r in rowspan_tracker.items() if r > 1}

        col = 0
        for td_m in re.finditer(r'<td([^>]*)>(.*?)</td>', tr_html, re.DOTALL):
            col += 1
            # Skip columns occupied by rowspan from previous rows
            while col in rowspan_tracker:
                col += 1

            attrs = td_m.group(1)
            content = td_m.group(2)

            rs = int(re.search(r'rowspan="(\d+)"', attrs).group(1)) if 'rowspan' in attrs else 1
            cs = int(re.search(r'colspan="(\d+)"', attrs).group(1)) if 'colspan' in attrs else 1
            style_m = re.search(r'style="([^"]*)"', attrs)
            style = style_m.group(1) if style_m else ""

            val = re.sub(r'<[^>]+>', '', content).strip()

            cells[(row, col)] = {
                "value": val,
                "style": style,
                "rowspan": rs,
                "colspan": cs,
            }

            # Track rowspan for subsequent rows
            if rs > 1:
                for c in range(col, col + cs):
                    rowspan_tracker[c] = rs

            col += cs - 1

    max_col = max(
        (c + info.get("colspan", 1) - 1 for (r, c), info in cells.items()),
        default=0,
    )

    return {
        "cells": cells,
        "max_row": row,
        "max_col": max_col,
        "col_widths": col_widths,
        "row_heights": row_heights,
    }


def parse_css_style(style_str: str) -> dict[str, str]:
    """Parse CSS style string into dict."""
    result = {}
    for part in style_str.split(';'):
        if ':' in part:
            k, v = part.split(':', 1)
            result[k.strip()] = v.strip()
    return result


# ── Excel cell info extraction ──

def _is_col_hidden(ws, col: int) -> bool:
    """Check if a column is hidden."""
    letter = get_column_letter(col)
    dim = ws.column_dimensions.get(letter)
    return bool(dim and dim.hidden)


def get_excel_cell_info(ws, row: int, col: int) -> dict:
    """Extract comprehensive cell info from openpyxl."""
    cell = ws.cell(row, col)
    info = {
        "value": cell.value,
        "width": None,
        "height": None,
    }

    # Column width
    letter = get_column_letter(col)
    dim = ws.column_dimensions.get(letter)
    if dim:
        info["width"] = dim.width

    # Row height
    rd = ws.row_dimensions.get(row)
    if rd:
        info["height"] = rd.height

    # Styles
    style = {}
    if cell.font:
        style["font_name"] = cell.font.name
        style["font_size"] = cell.font.size
        style["font_bold"] = cell.font.bold
        style["font_color"] = _color_to_hex(cell.font.color)

    if cell.fill:
        style["bg_color"] = _color_to_hex(cell.fill.fgColor)
        if not style.get("bg_color") and hasattr(cell.fill, "start_color") and cell.fill.start_color:
            style["bg_color"] = _color_to_hex(cell.fill.start_color)

    # Conditional formatting background colour
    style["cf_bg_color"] = _get_conditional_bg_color(ws, row, col)

    if cell.border:
        for side in ["top", "bottom", "left", "right"]:
            side_obj = getattr(cell.border, side)
            if side_obj:
                style[f"border_{side}_style"] = side_obj.style
                style[f"border_{side}_color"] = _color_to_hex(side_obj.color)

    if cell.alignment:
        style["align_h"] = cell.alignment.horizontal
        style["align_v"] = cell.alignment.vertical

    info["style"] = style
    return info


# ── Format helpers ──

def format_excel_value(value) -> str:
    """Format an Excel cell value for string comparison."""
    if value is None:
        return ""
    if hasattr(value, "year"):
        return str(value.day)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, int) or value == int(value):
            return str(int(value))
        return f"{value:.2f}"
    # Strip strings for comparison (Excel often has trailing spaces)
    if isinstance(value, str):
        return value.strip()
    return str(value)


def parse_html_border(border_css: str) -> tuple[str | None, str | None, str | None]:
    """Parse CSS border value into (width, style, color)."""
    if not border_css:
        return None, None, None
    parts = border_css.split()
    if len(parts) >= 3:
        return parts[0], parts[1], parts[2]
    if len(parts) == 2:
        return parts[0], parts[1], None
    if len(parts) == 1:
        return parts[0], None, None
    return None, None, None


# ── Core verification ──

def verify_sheet(html_data: dict, ws, value_ws=None) -> list[dict]:
    """Compare HTML sheet against Excel worksheet cell by cell.

    Returns list of error dicts.
    """
    errors = []
    html_cells = html_data["cells"]

    # ── 1. Parse print area ──
    bounds = _get_print_area_bounds(ws)
    if bounds:
        min_col, min_row, max_col, max_row = bounds
    else:
        min_col, min_row, max_col, max_row = 1, 1, ws.max_column, ws.max_row

    # ── 2. Build visible row/col lists ──
    visible_rows = [
        r for r in range(min_row, max_row + 1)
        if not ws.row_dimensions[r].hidden
    ]
    visible_cols = [
        c for c in range(min_col, max_col + 1)
        if not _is_col_hidden(ws, c)
    ]

    row_map = {r: i + 1 for i, r in enumerate(visible_rows)}
    col_map = {c: i + 1 for i, c in enumerate(visible_cols)}

    # ── 3. Build merged cell map from Excel ──
    excel_merged: dict[tuple[int, int], tuple[int, int]] = {}
    excel_merged_covered: set[tuple[int, int]] = set()

    for mr in ws.merged_cells.ranges:
        # Skip if outside print_area
        if mr.max_row < min_row or mr.min_row > max_row:
            continue
        if mr.max_col < min_col or mr.min_col > max_col:
            continue

        visible_mr_rows = [r for r in range(mr.min_row, mr.max_row + 1) if r in row_map]
        visible_mr_cols = [c for c in range(mr.min_col, mr.max_col + 1) if c in col_map]
        if not visible_mr_rows or not visible_mr_cols:
            continue

        leader_r = visible_mr_rows[0]
        leader_c = visible_mr_cols[0]
        rs = len(visible_mr_rows)
        cs = len(visible_mr_cols)

        excel_merged[(leader_r, leader_c)] = (rs, cs)
        for r in visible_mr_rows:
            for c in visible_mr_cols:
                if (r, c) != (leader_r, leader_c):
                    excel_merged_covered.add((r, c))

    # ── 4. Dimension check ──
    expected_rows = len(visible_rows)
    expected_cols = len(visible_cols)

    if html_data["max_row"] != expected_rows:
        errors.append({
            "type": "dimension_mismatch",
            "msg": f"行数不匹配: HTML={html_data['max_row']}, Excel可见行={expected_rows}",
        })

    if html_data["max_col"] != expected_cols:
        errors.append({
            "type": "dimension_mismatch",
            "msg": f"列数不匹配: HTML={html_data['max_col']}, Excel可见列={expected_cols}",
        })

    # Detect date/ weekday rows for special handling
    date_row, day_start_col = _find_date_start(ws)
    weekday_row = _find_weekday_row(ws, date_row) if date_row else None
    year, month_num = None, None
    days_in_month = None
    if date_row:
        # Infer month from first date cell
        first_date = ws.cell(date_row, day_start_col).value
        if first_date and hasattr(first_date, "year"):
            year, month_num = first_date.year, first_date.month
            days_in_month = calendar.monthrange(year, month_num)[1]

    # ── 5. Verify each visible cell ──
    check_limit = None  # Set to a number for faster debugging, None for full check
    checked = 0

    for excel_r in visible_rows:
        for excel_c in visible_cols:
            if check_limit and checked >= check_limit:
                break
            checked += 1

            html_r = row_map[excel_r]
            html_c = col_map[excel_c]

            # Skip Excel merged followers (they have no independent value)
            excel_cell = ws.cell(excel_r, excel_c)
            is_merged_follower = (
                isinstance(excel_cell, MergedCell)
                and (excel_r, excel_c) not in excel_merged
            )

            # Check if cell exists in HTML
            if (html_r, html_c) not in html_cells:
                if is_merged_follower:
                    continue
                errors.append({
                    "pos": (excel_r, excel_c),
                    "type": "missing_html_cell",
                    "msg": f"HTML 缺少单元格 ({excel_r},{excel_c}) -> HTML({html_r},{html_c})",
                })
                continue

            html_cell = html_cells[(html_r, html_c)]

            if is_merged_follower:
                # Merged followers should NOT appear as independent cells in HTML
                errors.append({
                    "pos": (excel_r, excel_c),
                    "type": "unexpected_html_cell",
                    "msg": f"HTML 不应包含合并单元格的跟随单元格 ({excel_r},{excel_c})",
                })
                continue

            # ── Value check ──
            html_val = html_cell["value"]
            excel_val = excel_cell.value

            # Special handling for date row (holidays appended in HTML)
            if excel_r == date_row and day_start_col and excel_c >= day_start_col:
                day = excel_c - day_start_col + 1
                if year and month_num and 1 <= day <= days_in_month:
                    expected = str(day)
                    # HTML may append holiday name — only check the leading number
                    if not html_val.startswith(expected):
                        errors.append({
                            "pos": (excel_r, excel_c),
                            "type": "value_mismatch",
                            "msg": f"日期值不匹配: HTML='{html_val}' vs expected='{expected}'",
                        })
                continue

            # Special handling for weekday row (compute expected value ourselves)
            if excel_r == weekday_row and day_start_col and excel_c >= day_start_col:
                day = excel_c - day_start_col + 1
                if year and month_num and 1 <= day <= days_in_month:
                    expected = WEEKDAY_CH[date(year, month_num, day).weekday()]
                    if html_val != expected:
                        errors.append({
                            "pos": (excel_r, excel_c),
                            "type": "value_mismatch",
                            "msg": f"星期值不匹配: HTML='{html_val}' vs expected='{expected}'",
                        })
                continue

            # Get computed value from value workbook
            computed_val = None
            if value_ws:
                try:
                    computed_val = value_ws.cell(excel_r, excel_c).value
                except Exception:
                    pass

            # For comparison, prefer computed value (what user sees in Excel)
            check_val = computed_val if computed_val is not None else excel_val
            excel_str = format_excel_value(check_val)

            # Special: formula cells must not display formula text
            if isinstance(excel_val, str) and excel_val.startswith("="):
                if html_val.startswith("="):
                    errors.append({
                        "pos": (excel_r, excel_c),
                        "type": "formula_not_computed",
                        "msg": f"公式显示为文本: {html_val[:60]}",
                    })
                    continue

            if html_val != excel_str:
                errors.append({
                    "pos": (excel_r, excel_c),
                    "type": "value_mismatch",
                    "msg": f"值不匹配: HTML='{html_val}' vs Excel='{excel_str}'",
                })

            # ── Background color check ──
            excel_info = get_excel_cell_info(ws, excel_r, excel_c)
            # Conditional format bg takes precedence over normal fill
            excel_bg = excel_info["style"].get("cf_bg_color") or excel_info["style"].get("bg_color")
            html_style = parse_css_style(html_cell["style"])
            html_bg = html_style.get("background-color")

            if excel_bg != html_bg:
                errors.append({
                    "pos": (excel_r, excel_c),
                    "type": "bgcolor_mismatch",
                    "msg": f"背景色: HTML={html_bg} vs Excel={excel_bg}",
                })

            # ── Border check (4 sides) ──
            # Excel border style -> CSS mapping (same as web_roster.py)
            _EXCEL_TO_CSS_STYLE = {
                "double": "double",
                "dashed": "dashed",
                "dotted": "dotted",
                "mediumDashDot": "dashed",
                "mediumDashDotDot": "dotted",
                "slantDashDot": "dashed",
            }
            _EXCEL_TO_CSS_WIDTH = {
                "hair": "0.5px",
                "thin": "1px",
                "medium": "2px",
                "thick": "3px",
                "double": "3px",
                "dashed": "1px",
                "dotted": "1px",
                "mediumDashDot": "2px",
                "mediumDashDotDot": "2px",
                "slantDashDot": "1px",
            }

            for side in ["top", "bottom", "left", "right"]:
                excel_bstyle = excel_info["style"].get(f"border_{side}_style")
                excel_bcolor = excel_info["style"].get(f"border_{side}_color")
                html_border = html_style.get(f"border-{side}", "")
                html_bwidth, html_bstyle, html_bcolor = parse_html_border(html_border)

                # Map Excel style to expected CSS style
                expected_style = _EXCEL_TO_CSS_STYLE.get(excel_bstyle, "solid") if excel_bstyle else None
                expected_width = _EXCEL_TO_CSS_WIDTH.get(excel_bstyle, "1px") if excel_bstyle else None

                # Compare style
                if expected_style != (html_bstyle or None):
                    errors.append({
                        "pos": (excel_r, excel_c),
                        "type": "border_style_mismatch",
                        "msg": f"边框 {side} style: HTML={html_bstyle} vs Excel={excel_bstyle} (expected {expected_style})",
                    })

                # Compare width
                if expected_width and html_bwidth and expected_width != html_bwidth:
                    errors.append({
                        "pos": (excel_r, excel_c),
                        "type": "border_width_mismatch",
                        "msg": f"边框 {side} width: HTML={html_bwidth} vs Excel={expected_width}",
                    })

                # Compare color (only if both have borders)
                if excel_bcolor and html_bcolor and excel_bcolor != html_bcolor:
                    errors.append({
                        "pos": (excel_r, excel_c),
                        "type": "border_color_mismatch",
                        "msg": f"边框 {side} color: HTML={html_bcolor} vs Excel={excel_bcolor}",
                    })

            # ── Merge check ──
            if (excel_r, excel_c) in excel_merged:
                expected_rs, expected_cs = excel_merged[(excel_r, excel_c)]
                if html_cell["rowspan"] != expected_rs or html_cell["colspan"] != expected_cs:
                    errors.append({
                        "pos": (excel_r, excel_c),
                        "type": "merge_mismatch",
                        "msg": (
                            f"合并单元格: HTML rs={html_cell['rowspan']},cs={html_cell['colspan']} "
                            f"vs Excel rs={expected_rs},cs={expected_cs}"
                        ),
                    })

            # ── Width check (only once per column, at first row) ──
            if excel_r == visible_rows[0]:
                html_width_px = html_data["col_widths"].get(html_c)
                excel_width = excel_info["width"]
                if html_width_px is not None and excel_width is not None:
                    expected_px = int(excel_width * 7)
                    if abs(html_width_px - expected_px) > 2:
                        errors.append({
                            "pos": (excel_r, excel_c),
                            "type": "width_mismatch",
                            "msg": f"列宽: HTML={html_width_px}px vs Excel~{expected_px}px",
                        })

            # ── Height check (only once per row, at first col) ──
            if excel_c == visible_cols[0]:
                html_height_px = html_data["row_heights"].get(html_r)
                excel_height = excel_info["height"]
                if html_height_px is not None and excel_height is not None:
                    expected_px = int(excel_height * 1.33)
                    if abs(html_height_px - expected_px) > 2:
                        errors.append({
                            "pos": (excel_r, excel_c),
                            "type": "height_mismatch",
                            "msg": f"行高: HTML={html_height_px}px vs Excel~{expected_px}px",
                        })

    # ── 6. Check hidden cells are NOT in HTML ──
    # (Already ensured by construction: we only iterate visible cells)
    # But double-check: no HTML cell maps to a hidden Excel cell
    for (html_r, html_c), html_cell in html_cells.items():
        # Reverse lookup
        excel_r_candidates = [r for r, hr in row_map.items() if hr == html_r]
        excel_c_candidates = [c for c, hc in col_map.items() if hc == html_c]
        if not excel_r_candidates or not excel_c_candidates:
            errors.append({
                "pos": (html_r, html_c),
                "type": "orphan_html_cell",
                "msg": f"HTML 单元格 ({html_r},{html_c}) 无法映射到 Excel 单元格",
            })

    return errors


# ── Main ──

def main():
    print("=" * 80)
    print("Web Roster 逐格验证 (Issue #5)")
    print("=" * 80)

    # Use 东汇-保洁轮休表_template.xlsx for testing
    template_path = "templates/东汇-保洁轮休表_template.xlsx"

    all_ok = True
    for month in ["2026-02", "2026-04"]:
        print(f"\n【验证】东汇-保洁 — {month}")

        # Generate HTML
        result = run("A", project_id=2, category_id=2, month=month)
        sheets_html = result["sheets"]
        wb = result.get("_wb")
        wb_values = result.get("_wb_values")

        if wb is None:
            print(f"  ⚠ 无法获取处理后的 workbook")
            continue

        # Use processed workbook for verification (matches HTML output)
        total_errors = 0
        for sheet_name, html in sheets_html.items():
            print(f"\n  Sheet: {sheet_name}")

            if sheet_name not in wb.sheetnames:
                print(f"    ⚠ Sheet '{sheet_name}' 不在处理后的 workbook 中")
                continue

            ws = wb[sheet_name]
            value_ws = wb_values[sheet_name] if wb_values and sheet_name in wb_values.sheetnames else None

            # Parse HTML
            html_data = parse_html_table(html)
            print(f"    HTML: {html_data['max_row']} 行 × {html_data['max_col']} 列")

            # Print area info
            bounds = _get_print_area_bounds(ws)
            if bounds:
                print(f"    Print area: {get_column_letter(bounds[0])}{bounds[1]}:{get_column_letter(bounds[2])}{bounds[3]}")

            errors = verify_sheet(html_data, ws, value_ws=value_ws)
            if errors:
                total_errors += len(errors)
                print(f"    ❌ {len(errors)} 个差异")
                for e in errors[:10]:
                    pos = e.get("pos", "")
                    print(f"      ({pos[0]},{pos[1]}) [{e['type']}] {e['msg'][:120]}")
                if len(errors) > 10:
                    print(f"      ... 还有 {len(errors) - 10} 个")
            else:
                print(f"    ✅ 无差异")

        if total_errors == 0:
            print(f"\n  ✅ 东汇-保洁 {month} — 全部通过")
        else:
            print(f"\n  ❌ 东汇-保洁 {month} — 共 {total_errors} 个差异")
            all_ok = False

    print("\n" + "=" * 80)
    if all_ok:
        print("✅ 所有测试全部通过逐格验证!")
    else:
        print("❌ 存在差异，需要进一步修复")
    print("=" * 80)


if __name__ == "__main__":
    main()
