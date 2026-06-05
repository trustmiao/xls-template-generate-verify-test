#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Common row-deletion utilities for Excel templates.

Replaces openpyxl's built-in ``ws.delete_rows()`` which suffers from a
phantom-empty-cell bug: ``iter_rows`` creates empty Cell objects for every
column in the traversed range; these empty cells are then moved upward and
overwrite valid data below the deleted row.

This module provides a **safe** replacement that directly mutates the
internal ``ws._cells`` dictionary, avoiding the phantom-cell problem
entirely.

Public API
----------

.. function:: delete_row_and_fixup(ws, row_to_delete)

   High-level helper — delete one row and fix all dependent structures
   (formulas, merged cells, print area, conditional formatting, row height).

.. function:: safe_delete_rows(ws, idx, amount=1)

   Low-level safe row deletion.  Mimics Excel behaviour: cells shift up,
   row heights/styles shift up, and no phantom empty cells are created.

All other functions are helpers called by the two functions above; they are
kept public so callers can compose their own pipelines if needed.
"""

import re
import copy

__all__ = [
    "adjust_formula",
    "adjust_all_formulas",
    "fix_merged_cells",
    "adjust_print_area",
    "adjust_conditional_formatting",
    "safe_delete_rows",
    "delete_row_and_fixup",
]

# ── Formula Adjustment ──
cell_ref = re.compile(r"(?<![A-Z$])(\$?[A-Z]{1,3})(\$?)(\d+)")
range_pattern = re.compile(
    r"(?<![A-Z$])(\$?[A-Z]{1,3})(\$?)(\d+)(:(\$?[A-Z]{1,3})(\$?)(\d+))"
)


def adjust_formula(formula, deleted_row):
    """
    Adjust cell references inside *formula* after deleting *deleted_row*.

    Rules (match Excel semantics):

    * Range refs (e.g. ``A1:A5``) — both bounds use ``>=`` rule.  The range
      shrinks by one row when the deleted row falls inside it.
    * Single refs (e.g. ``A1``) — use ``>`` rule.  A ref pointing *exactly*
      to the deleted row becomes ``#REF!``.
    * Absolute row refs (``$A$1``, ``A$1``) are never shifted.
    """
    range_positions = set()
    for m in range_pattern.finditer(formula):
        range_positions.add((m.start(), m.end()))

    result = []
    last_end = 0

    for match in cell_ref.finditer(formula):
        in_range = any(
            start <= match.start() < end for start, end in range_positions
        )

        result.append(formula[last_end:match.start()])
        col_part = match.group(1)
        abs_row = match.group(2) == "$"
        row_num = int(match.group(3))

        if not abs_row:
            if in_range:
                if row_num >= deleted_row:
                    row_num -= 1
            else:
                if row_num > deleted_row:
                    row_num -= 1
                elif row_num == deleted_row:
                    result.append("#REF!")
                    last_end = match.end()
                    continue

        result.append(f"{col_part}{'$' if abs_row else ''}{row_num}")
        last_end = match.end()

    result.append(formula[last_end:])
    return "".join(result)


def adjust_all_formulas(ws, deleted_row):
    """Scan every cell in *ws* and patch formulas referencing *deleted_row*."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = cell.value
            if val and isinstance(val, str) and val.startswith("="):
                cell.value = adjust_formula(val, deleted_row)


# ── Merged Cells ──
def adjust_merged_range(range_str, deleted_row):
    """Adjust a single merged-range string after a row deletion."""
    ref_pattern = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")

    def parse_ref(ref):
        m = ref_pattern.match(ref)
        if m:
            col_abs, col, row_abs, row = m.groups()
            row = int(row)
            if row > deleted_row:
                row -= 1
            return f"{col_abs}{col}{row_abs}{row}"
        return ref

    if ":" in range_str:
        start, end = range_str.split(":")
        return f"{parse_ref(start)}:{parse_ref(end)}"
    return parse_ref(range_str)


def fix_merged_cells(ws, deleted_row):
    """
    Adjust merged-cell ranges after a row deletion **without** unmerge+merge.

    openpyxl's ``unmerge_cells`` + ``merge_cells`` has a side-effect of
copying the top-left cell's style to **all** cells in the merged range,
    corrupting styles.  We directly mutate :class:`MergedCellRange` objects
    in place.

    Important: :class:`MergedCellRange` is hashable; mutating its bounds
    changes its hash and breaks the internal ``set``.  We **remove** each
    range before mutation and **re-add** it only if it survives.
    """
    to_keep = []
    to_remove = []
    for mr in list(ws.merged_cells.ranges):
        # Remove before mutation to avoid hash-inconsistency in the set
        ws.merged_cells.ranges.remove(mr)
        if mr.min_row > deleted_row:
            # Only adjust row numbers (``shift`` would also move columns,
            # causing ``Invalid shift value`` when ``min_col == 1``).
            mr.min_row -= 1
            mr.max_row -= 1
            to_keep.append(mr)
        elif mr.max_row < deleted_row:
            to_keep.append(mr)
        else:
            if mr.min_row == mr.max_row == deleted_row:
                to_remove.append(mr)
            else:
                mr.max_row -= 1
                if mr.min_row > mr.max_row:
                    to_remove.append(mr)
                else:
                    to_keep.append(mr)

    for mr in to_keep:
        ws.merged_cells.ranges.add(mr)


# ── Print Area ──
def adjust_print_area(ws, deleted_row):
    """Shrink or shift the worksheet print area after a row deletion."""
    if not ws.print_area or not ws._print_area:
        return
    for cr in ws._print_area.ranges:
        if cr.min_row > deleted_row:
            cr.min_row -= 1
        if cr.max_row > deleted_row:
            cr.max_row -= 1


# ── Conditional Formatting ──
def adjust_conditional_formatting(ws, deleted_row):
    """
    Adjust conditional-formatting ranges after a row deletion.

    Rules:
      * Ranges entirely below the deleted row → shift up by 1.
      * Ranges entirely above → unchanged.
      * Ranges that span the deleted row → shrink ``max_row`` by 1.
      * Single-row ranges on the deleted row → removed.

    Rebuilds the worksheet's ``ConditionalFormattingList`` from scratch to
    avoid internal key/hash inconsistencies when mutating existing objects.
    """
    from collections import OrderedDict
    from openpyxl.formatting.formatting import ConditionalFormatting
    from openpyxl.worksheet.cell_range import MultiCellRange

    new_entries = []          # list of (ConditionalFormatting, list_of_rules)

    for cf in ws.conditional_formatting:
        rules = list(cf.rules)
        adjusted_ranges = []

        for cr in cf.cells:
            if cr.min_row > deleted_row:
                # Entirely below — shift up by 1
                new_cr = copy.copy(cr)
                new_cr.min_row -= 1
                new_cr.max_row -= 1
                adjusted_ranges.append(str(new_cr))
            elif cr.max_row < deleted_row:
                # Entirely above — unchanged
                adjusted_ranges.append(str(cr))
            else:
                # deleted_row is inside or at the boundary
                if cr.min_row == cr.max_row == deleted_row:
                    # Single-row range completely deleted — drop it
                    continue
                else:
                    new_cr = copy.copy(cr)
                    new_cr.max_row -= 1
                    if new_cr.min_row > new_cr.max_row:
                        continue
                    adjusted_ranges.append(str(new_cr))

        if adjusted_ranges:
            new_cf = ConditionalFormatting()
            new_cf.cells = MultiCellRange(" ".join(adjusted_ranges))
            new_entries.append((new_cf, rules))

    # Rebuild _cf_rules from scratch
    ws.conditional_formatting._cf_rules = OrderedDict()
    for new_cf, rules in new_entries:
        ws.conditional_formatting._cf_rules[new_cf] = rules


# ── Safe Row Deletion (replaces ws.delete_rows) ──
def safe_delete_rows(ws, idx, amount=1):
    """
    Delete rows, correctly mimicking Excel behaviour.

    Replaces openpyxl's ``ws.delete_rows()`` which has an ``iter_rows`` bug
    that creates phantom empty cells which then overwrite valid data during
    the shift.

    This implementation directly mutates ``ws._cells``, moving each real
    :class:`Cell` object upward.  Because the entire Cell object (including
    its ``_style`` — font, fill, border, alignment, number_format,
    protection) is moved, all cell-level formatting travels with the data.

    Row-level attributes (height, hidden, outline, collapsed, and row style)
    are also copied from the source row to the target row.
    """
    max_row = max((row for row, col in ws._cells.keys()), default=0)

    # 1. Move cells upward by directly mutating ws._cells
    for r in range(idx, max_row + 1):
        source_r = r + amount
        if source_r > max_row:
            # Source row does not exist — clear the target row
            for key in list(ws._cells.keys()):
                if key[0] == r:
                    del ws._cells[key]
            continue

        target_cols = {col for row, col in ws._cells.keys() if row == r}
        source_cols = {col for row, col in ws._cells.keys() if row == source_r}

        # Delete target cells that have no counterpart in the source row
        for col in target_cols - source_cols:
            del ws._cells[(r, col)]

        # Move source cells to the target row
        for col in source_cols:
            if (source_r, col) in ws._cells:
                cell = ws._cells[(source_r, col)]
                ws._cells[(r, col)] = cell
                cell.row = r
                del ws._cells[(source_r, col)]

    # 2. Copy row dimensions (height + style + other attrs) upward
    for r in range(idx, max_row + 1):
        source_r = r + amount
        dst_rd = ws.row_dimensions[r]
        if source_r in ws.row_dimensions:
            src_rd = ws.row_dimensions[source_r]
            dst_rd.height = src_rd.height
            if src_rd._style is not None:
                dst_rd._style = copy.copy(src_rd._style)
            else:
                dst_rd._style = None
            dst_rd.hidden = src_rd.hidden
            dst_rd.outline_level = src_rd.outline_level
            dst_rd.collapsed = src_rd.collapsed
        else:
            dst_rd.height = None
            dst_rd._style = None
            dst_rd.hidden = False
            dst_rd.outline_level = 0
            dst_rd.collapsed = False

    # 3. Update internal cursor so ws.max_row reflects the new reality
    ws._current_row = max(
        row for row, col in ws._cells.keys()
    ) if ws._cells else 0


# ── High-level: delete one row and fix everything ──
def delete_row_and_fixup(ws, row_to_delete):
    """
    Delete a single row and fix all dependent structures.

    Applies, in order:

      1. Formula adjustment
      2. Safe physical row deletion (cell styles, borders, row heights)
      3. Merged-cell range adjustment
      4. Print-area adjustment
      5. Conditional-formatting adjustment
    """
    adjust_all_formulas(ws, row_to_delete)
    safe_delete_rows(ws, row_to_delete, 1)
    fix_merged_cells(ws, row_to_delete)
    adjust_print_area(ws, row_to_delete)
    adjust_conditional_formatting(ws, row_to_delete)
