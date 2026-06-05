#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Common column-deletion utilities for Excel templates.

Replaces openpyxl's built-in ``ws.delete_cols()`` which suffers from the
same phantom-empty-cell bug as ``delete_rows``: ``iter_cols`` creates empty
Cell objects for every row in the traversed range; these empty cells are
then moved leftward and overwrite valid data to the right of the deleted
column.

This module provides a **safe** replacement that directly mutates the
internal ``ws._cells`` dictionary, avoiding the phantom-cell problem
entirely.

Public API
----------

.. function:: delete_col_and_fixup(ws, col_to_delete)

   High-level helper — delete one column and fix all dependent structures
   (formulas, merged cells, print area, conditional formatting, column width).

.. function:: safe_delete_cols(ws, idx, amount=1)

   Low-level safe column deletion.  Mimics Excel behaviour: cells shift left,
   column widths/styles shift left, and no phantom empty cells are created.

All other functions are helpers called by the two functions above; they are
kept public so callers can compose their own pipelines if needed.
"""

import re
import copy

from openpyxl.utils import get_column_letter

__all__ = [
    "adjust_formula",
    "adjust_all_formulas",
    "fix_merged_cells",
    "adjust_print_area",
    "adjust_conditional_formatting",
    "safe_delete_cols",
    "delete_col_and_fixup",
]

# ── Formula Adjustment ──
cell_ref = re.compile(r"(?<![A-Z$])(\$?[A-Z]{1,3})(\$?)(\d+)")
range_pattern = re.compile(
    r"(?<![A-Z$])(\$?[A-Z]{1,3})(\$?)(\d+)(:(\$?[A-Z]{1,3})(\$?)(\d+))"
)


def _col_letter_to_num(letter: str) -> int:
    """Convert Excel column letter(s) to a 1-based column number."""
    result = 0
    for c in letter.upper():
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result


def _col_num_to_letter(n: int) -> str:
    """Convert a 1-based column number to Excel column letter(s)."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def adjust_formula(formula, deleted_col):
    """
    Adjust cell references inside *formula* after deleting *deleted_col*.

    Rules (match Excel semantics):

    * Range refs (e.g. ``A1:C1``) — both bounds use ``>=`` rule.  The range
      shrinks by one column when the deleted column falls inside it.
    * Single refs (e.g. ``A1``) — use ``>`` rule.  A ref pointing *exactly*
      to the deleted column becomes ``#REF!``.
    * Absolute column refs (``$A$1``, ``$A1``) are never shifted.
    """
    range_positions = set()
    for m in range_pattern.finditer(formula):
        range_positions.add((m.start(), m.end()))

    result = []
    last_end = 0

    for match in cell_ref.finditer(formula):
        in_range = any(
            start <= match.start() < end for start, end in range_positions
        )

        result.append(formula[last_end:match.start()])
        col_part = match.group(1)
        abs_row = match.group(2) == "$"
        row_num = match.group(3)

        # col_part may include leading '$' (e.g. '$A' or 'A')
        abs_col = col_part.startswith("$")
        col_letter = col_part.lstrip("$")
        col_num = _col_letter_to_num(col_letter)

        if not abs_col:
            if in_range:
                if col_num >= deleted_col:
                    col_num -= 1
            else:
                if col_num > deleted_col:
                    col_num -= 1
                elif col_num == deleted_col:
                    result.append("#REF!")
                    last_end = match.end()
                    continue

        if col_num <= 0:
            result.append("#REF!")
            last_end = match.end()
            continue

        new_letter = _col_num_to_letter(col_num)
        result.append(
            f"{'$' if abs_col else ''}{new_letter}{'$' if abs_row else ''}{row_num}"
        )
        last_end = match.end()

    result.append(formula[last_end:])
    return "".join(result)


def adjust_all_formulas(ws, deleted_col):
    """Scan every cell in *ws* and patch formulas referencing *deleted_col*."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = cell.value
            if val and isinstance(val, str) and val.startswith("="):
                cell.value = adjust_formula(val, deleted_col)


# ── Merged Cells ──
def fix_merged_cells(ws, deleted_col):
    """
    Adjust merged-cell ranges after a column deletion **without** unmerge+merge.

    openpyxl's ``unmerge_cells`` + ``merge_cells`` has a side-effect of
copying the top-left cell's style to **all** cells in the merged range,
    corrupting styles.  We directly mutate :class:`MergedCellRange` objects
    in place.
    """
    to_remove = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_col > deleted_col:
            # Only adjust column numbers (``shift`` would also move rows,
            # causing ``Invalid shift value`` when ``min_row == 1``).
            mr.min_col -= 1
            mr.max_col -= 1
        elif mr.max_col < deleted_col:
            pass
        else:
            if mr.min_col == mr.max_col == deleted_col:
                to_remove.append(mr)
            else:
                mr.max_col -= 1
                if mr.min_col > mr.max_col:
                    to_remove.append(mr)

    for mr in to_remove:
        ws.merged_cells.ranges.remove(mr)


# ── Print Area ──
def adjust_print_area(ws, deleted_col):
    """Shrink or shift the worksheet print area after a column deletion."""
    if not ws.print_area or not ws._print_area:
        return
    for cr in ws._print_area.ranges:
        if cr.min_col > deleted_col:
            cr.min_col -= 1
        if cr.max_col > deleted_col:
            cr.max_col -= 1


# ── Conditional Formatting ──
def adjust_conditional_formatting(ws, deleted_col):
    """
    Adjust conditional-formatting ranges after a column deletion.

    Rules:
      * Ranges entirely to the right of the deleted column → shift left by 1.
      * Ranges entirely to the left → unchanged.
      * Ranges that span the deleted column → shrink ``max_col`` by 1.
      * Single-column ranges on the deleted column → removed.

    Rebuilds the worksheet's ``ConditionalFormattingList`` from scratch to
    avoid internal key/hash inconsistencies when mutating existing objects.
    """
    from collections import OrderedDict
    from openpyxl.formatting.formatting import ConditionalFormatting
    from openpyxl.worksheet.cell_range import MultiCellRange

    new_entries = []          # list of (ConditionalFormatting, list_of_rules)

    for cf in ws.conditional_formatting:
        rules = list(cf.rules)
        adjusted_ranges = []

        for cr in cf.cells:
            if cr.min_col > deleted_col:
                # Entirely to the right — shift left by 1
                new_cr = copy.copy(cr)
                new_cr.min_col -= 1
                new_cr.max_col -= 1
                adjusted_ranges.append(str(new_cr))
            elif cr.max_col < deleted_col:
                # Entirely to the left — unchanged
                adjusted_ranges.append(str(cr))
            else:
                # deleted_col is inside or at the boundary
                if cr.min_col == cr.max_col == deleted_col:
                    # Single-column range completely deleted — drop it
                    continue
                else:
                    new_cr = copy.copy(cr)
                    new_cr.max_col -= 1
                    if new_cr.min_col > new_cr.max_col:
                        continue
                    adjusted_ranges.append(str(new_cr))

        if adjusted_ranges:
            new_cf = ConditionalFormatting()
            new_cf.cells = MultiCellRange(" ".join(adjusted_ranges))
            new_entries.append((new_cf, rules))

    # Rebuild _cf_rules from scratch
    ws.conditional_formatting._cf_rules = OrderedDict()
    for new_cf, rules in new_entries:
        ws.conditional_formatting._cf_rules[new_cf] = rules


# ── Safe Column Deletion (replaces ws.delete_cols) ──
def safe_delete_cols(ws, idx, amount=1):
    """
    Delete columns, correctly mimicking Excel behaviour.

    Replaces openpyxl's ``ws.delete_cols()`` which has an ``iter_cols`` bug
    that creates phantom empty cells which then overwrite valid data during
    the shift.

    This implementation directly mutates ``ws._cells``, moving each real
    :class:`Cell` object leftward.  Because the entire Cell object
    (including its ``_style`` — font, fill, border, alignment, number_format,
    protection) is moved, all cell-level formatting travels with the data.

    Column-level attributes (width, hidden, outline, collapsed, and column
    style) are also copied from the source column to the target column.
    """
    max_col = max((col for row, col in ws._cells.keys()), default=0)
    if ws.column_dimensions:
        max_col = max(
            max_col,
            max(
                _col_letter_to_num(k)
                for k in ws.column_dimensions.keys()
            ),
        )

    # 1. Move cells leftward by directly mutating ws._cells
    for c in range(idx, max_col + 1):
        source_c = c + amount
        if source_c > max_col:
            # Source column does not exist — clear the target column
            for key in list(ws._cells.keys()):
                if key[1] == c:
                    del ws._cells[key]
            continue

        target_rows = {row for row, col in ws._cells.keys() if col == c}
        source_rows = {row for row, col in ws._cells.keys() if col == source_c}

        # Delete target cells that have no counterpart in the source column
        for row in target_rows - source_rows:
            del ws._cells[(row, c)]

        # Move source cells to the target column
        for row in source_rows:
            if (row, source_c) in ws._cells:
                cell = ws._cells[(row, source_c)]
                ws._cells[(row, c)] = cell
                cell.column = c
                del ws._cells[(row, source_c)]

    # 2. Copy column dimensions (width + style + other attrs) leftward
    for c in range(idx, max_col + 1):
        source_c = c + amount
        col_letter = get_column_letter(c)
        dst_cd = ws.column_dimensions[col_letter]
        if source_c <= max_col:
            src_letter = get_column_letter(source_c)
            if src_letter in ws.column_dimensions:
                src_cd = ws.column_dimensions[src_letter]
                dst_cd.width = src_cd.width
                if src_cd._style is not None:
                    dst_cd._style = copy.copy(src_cd._style)
                else:
                    dst_cd._style = None
                dst_cd.hidden = src_cd.hidden
                dst_cd.outline_level = src_cd.outline_level
                dst_cd.collapsed = src_cd.collapsed
            else:
                # Source column has no explicit dimension — reset target to defaults.
                # ColumnDimension.width is a Float (cannot be None); use the
                # openpyxl default of 13 characters.
                dst_cd.width = 13
                dst_cd._style = None
                dst_cd.hidden = False
                dst_cd.outline_level = 0
                dst_cd.collapsed = False
        else:
            dst_cd.width = 13
            dst_cd._style = None
            dst_cd.hidden = False
            dst_cd.outline_level = 0
            dst_cd.collapsed = False


# ── High-level: delete one column and fix everything ──
def delete_col_and_fixup(ws, col_to_delete):
    """
    Delete a single column and fix all dependent structures.

    Applies, in order:

      1. Formula adjustment
      2. Safe physical column deletion (cell styles, borders, column widths)
      3. Merged-cell range adjustment
      4. Print-area adjustment
      5. Conditional-formatting adjustment
    """
    adjust_all_formulas(ws, col_to_delete)
    safe_delete_cols(ws, col_to_delete, 1)
    fix_merged_cells(ws, col_to_delete)
    adjust_print_area(ws, col_to_delete)
    adjust_conditional_formatting(ws, col_to_delete)
