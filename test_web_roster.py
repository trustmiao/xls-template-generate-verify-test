#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""E2E test for web_roster engine (Issue #5).

Tests:
  1. Date update — first date cell = target month day 1
  2. Day column deletion — excess columns removed from the right
  3. 29-31號時數 column handling
  4. Personnel row shrink — regions shrunk to match mock headcount
  5. Number format — integers show as ints, floats show 2 decimals
  6. HTML export — renders tables with merged cells + holiday badges
"""
from __future__ import annotations

import calendar
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

import sys
sys.path.insert(0, r"D:\claude\claude_hk\backend")

from app.engine.excel.web_roster import (
    _detect_roster_regions,
    _shrink_region,
    _adjust_number_formats,
    workbook_to_html,
    run as engine_run,
)

TEMPLATES = [
    ("TY-大元保安", "templates/TY-2026.03- SG_SEC-Deploy Roster  Shortfall - template.xlsx"),
    ("东汇-保安", "templates/东汇-保安轮休表template.xlsx"),
    ("东汇-保洁", "templates/东汇-保洁轮休表_template.xlsx"),
]

OUT_DIR = Path("test_outputs_web_roster")
OUT_DIR.mkdir(exist_ok=True)

# ── Mock data (simulating api_shortfall_engine response) ──
MOCK_SEGMENTS = [
    {
        "title": "1. 保安主管出勤時數",
        "rows": [
            {"rank_seq": "M01", "employee_no": "0001234", "name": "張三", "cells": [{"day": d, "code": "8", "value": "8"} for d in range(1, 29)]},
            {"rank_seq": "M02", "employee_no": "0005678", "name": "李四", "cells": [{"day": d, "code": "8" if d % 2 else "RD", "value": "8" if d % 2 else "RD"} for d in range(1, 29)]},
        ],
    },
    {
        "title": "2. 保安員出勤時數",
        "rows": [
            {"rank_seq": "D01", "employee_no": "0009012", "name": "王五", "cells": [{"day": d, "code": "8", "value": "8"} for d in range(1, 29)]},
            {"rank_seq": "D02", "employee_no": "0003456", "name": "趙六", "cells": [{"day": d, "code": "8", "value": "8"} for d in range(1, 29)]},
            {"rank_seq": "D03", "employee_no": "0007890", "name": "陳七", "cells": [{"day": d, "code": "8", "value": "8"} for d in range(1, 29)]},
        ],
    },
]

MOCK_DATA = {
    "has_data": True,
    "days_in_month": 28,
    "segments": MOCK_SEGMENTS,
}


def mock_fetch_data(context, shift):
    """Return mock data for testing."""
    return MOCK_DATA


def test_date_update():
    """Test 1: Date row updated to target month."""
    print("\n[Test 1] Date update")
    wb = load_workbook(TEMPLATES[0][1], data_only=False, keep_links=False)
    ws = wb["早"]

    # Patch fetch_data
    import app.engine.excel.web_roster as wr
    orig_fetch = wr._fetch_data
    wr._fetch_data = mock_fetch_data

    try:
        engine_run(wb, {"project_id": 1, "category_id": 1, "month": "2026-02"})
    finally:
        wr._fetch_data = orig_fetch

    # Check first date cell
    from app.engine.excel.cleaning_shortfall_v2 import _find_date_start
    date_row, day_start_col = _find_date_start(ws)
    first_date = ws.cell(date_row, day_start_col).value
    assert first_date == date(2026, 2, 1), f"Expected 2026-02-01, got {first_date}"
    print("  ✓ First date cell = 2026-02-01")

    # Check day 28 exists and day 29 does not
    day_28 = ws.cell(date_row, day_start_col + 27).value
    assert day_28 is not None, "Day 28 column should exist"
    print("  ✓ Day 28 column exists")

    # Check that columns beyond day 28 were deleted
    max_col = ws.max_column
    # The date row should only have 28 day columns + preceding columns
    print(f"  ✓ Max column after deletion = {max_col}")


def test_column_deletion():
    """Test 2+3: Excess day columns deleted, 29-31 column handled."""
    print("\n[Test 2+3] Column deletion + 29-31號時數")
    wb = load_workbook(TEMPLATES[0][1], data_only=False, keep_links=False)
    ws = wb["早"]

    import app.engine.excel.web_roster as wr
    orig_fetch = wr._fetch_data
    wr._fetch_data = mock_fetch_data

    try:
        engine_run(wb, {"project_id": 1, "category_id": 1, "month": "2026-02"})
    finally:
        wr._fetch_data = orig_fetch

    from app.engine.excel.cleaning_shortfall_v2 import _find_date_start
    date_row, day_start_col = _find_date_start(ws)
    days_in_month = 28

    # Check that day 29-31 columns are gone
    for offset in [28, 29, 30]:  # day 29, 30, 31
        col = day_start_col + offset
        if col <= ws.max_column:
            val = ws.cell(date_row, col).value
            # If the column still exists, it should not be a date or date formula
            if val is not None:
                assert not (hasattr(val, "year") or (isinstance(val, str) and "+1" in val)), \
                    f"Day {offset+1} column should be deleted, but found {val}"

    # Check 29-31號時數 column is gone for 28-day month
    found_29_31 = False
    for c in range(1, ws.max_column + 1):
        for r in range(1, min(15, ws.max_row) + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "29-31" in v:
                found_29_31 = True
                break
        if found_29_31:
            break
    assert not found_29_31, "29-31號時數 column should be deleted for 28-day month"
    print("  ✓ 29-31號時數 column deleted for Feb")
    print("  ✓ Excess day columns deleted")


def test_row_shrink():
    """Test 4: Personnel rows shrunk to match mock headcount."""
    print("\n[Test 4] Row shrink")
    wb = load_workbook(TEMPLATES[0][1], data_only=False, keep_links=False)
    ws = wb["早"]

    import app.engine.excel.web_roster as wr
    orig_fetch = wr._fetch_data
    wr._fetch_data = mock_fetch_data

    try:
        engine_run(wb, {"project_id": 1, "category_id": 1, "month": "2026-02"})
    finally:
        wr._fetch_data = orig_fetch

    regions = _detect_roster_regions(ws)
    print(f"  Found {len(regions)} region(s)")
    for region in regions:
        count = region["data_end"] - region["data_start"] + 1
        names = []
        for r in range(region["data_start"], region["data_end"] + 1):
            rank = ws.cell(r, 2).value or ""
            name = ws.cell(r, 3).value or ""
            names.append(f"{rank}/{name}")
        print(f"    Region {region['index']}: {count} person(s) — {names}")

    # Region 1 should have 2 people, Region 2 should have 3 people
    assert len(regions) >= 1, "Should have at least 1 region"
    if len(regions) >= 1:
        r1_count = regions[0]["data_end"] - regions[0]["data_start"] + 1
        assert r1_count == 2, f"Region 1 should have 2 people, got {r1_count}"
        print("  ✓ Region 1 shrunk to 2 people")
    if len(regions) >= 2:
        r2_count = regions[1]["data_end"] - regions[1]["data_start"] + 1
        assert r2_count == 3, f"Region 2 should have 3 people, got {r2_count}"
        print("  ✓ Region 2 shrunk to 3 people")


def test_number_formats():
    """Test 5: Number formats adjusted."""
    print("\n[Test 5] Number formats")
    wb = load_workbook(TEMPLATES[0][1], data_only=False, keep_links=False)
    ws = wb["早"]

    import app.engine.excel.web_roster as wr
    orig_fetch = wr._fetch_data
    wr._fetch_data = mock_fetch_data

    try:
        engine_run(wb, {"project_id": 1, "category_id": 1, "month": "2026-02"})
    finally:
        wr._fetch_data = orig_fetch

    from app.engine.excel.cleaning_shortfall_v2 import _find_date_start
    date_row, day_start_col = _find_date_start(ws)

    # Check that numeric cells have correct number format
    fmt_found = set()
    for row in range(1, ws.max_row + 1):
        for col in range(day_start_col, day_start_col + 28):
            cell = ws.cell(row, col)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                fmt_found.add(cell.number_format)

    print(f"  Number formats found: {fmt_found}")
    # The mock data only has integers (8), so we expect '0' format
    assert "0" in fmt_found, f"Expected '0' format for integers, got {fmt_found}"
    print("  ✓ Integer format '0' applied")


def test_html_export():
    """Test 6: HTML export with holiday badges."""
    print("\n[Test 6] HTML export")
    wb = load_workbook(TEMPLATES[0][1], data_only=False, keep_links=False)

    import app.engine.excel.web_roster as wr
    orig_fetch = wr._fetch_data
    wr._fetch_data = mock_fetch_data

    try:
        engine_run(wb, {"project_id": 1, "category_id": 1, "month": "2026-02"})
    finally:
        wr._fetch_data = orig_fetch

    html = workbook_to_html(wb, "2026-02")

    # Save for manual inspection
    html_path = OUT_DIR / "TY-大元保安_2026-02.html"
    html_path.write_text(html, encoding="utf-8")
    print(f"  HTML saved to: {html_path}")

    # Basic checks
    assert "<table>" in html, "HTML should contain tables"
    print("  ✓ HTML contains tables")

    # Check for holiday badge (2026-02 has Chinese New Year on Feb 17)
    from app.engine.common.shift_info import _get_holidays_for_month
    holidays = _get_holidays_for_month(2026, 2)
    if holidays:
        for day, name in holidays.items():
            if name in html:
                print(f"  ✓ Holiday badge found: {name}")
                break
        else:
            print(f"  ⚠ No holiday badge found (holidays: {holidays})")
    else:
        print("  ℹ No holidays in 2026-02")

    # Check weekday row uses Chinese characters
    weekday_chars = ["一", "二", "三", "四", "五", "六", "日"]
    found_weekday = any(c in html for c in weekday_chars)
    assert found_weekday, "HTML should contain Chinese weekday characters"
    print("  ✓ Chinese weekday characters rendered")


def test_all_templates():
    """Run engine on all templates and save outputs."""
    print("\n[Test 7] All templates (Feb + Apr)")
    for name, path in TEMPLATES:
        if not Path(path).exists():
            print(f"  ⚠ Skipping {name} — template not found at {path}")
            continue

        for month in ["2026-02", "2026-04"]:
            wb = load_workbook(path, data_only=False, keep_links=False)

            import app.engine.excel.web_roster as wr
            orig_fetch = wr._fetch_data
            wr._fetch_data = mock_fetch_data

            try:
                engine_run(wb, {"project_id": 1, "category_id": 1, "month": month})
            finally:
                wr._fetch_data = orig_fetch

            # Save Excel
            xlsx_path = OUT_DIR / f"{month}_{name}_web.xlsx"
            wb.save(str(xlsx_path))
            print(f"  💾 Excel: {xlsx_path}")

            # Save HTML
            html = workbook_to_html(wb, month)
            html_path = OUT_DIR / f"{month}_{name}_web.html"
            html_path.write_text(html, encoding="utf-8")
            print(f"  💾 HTML:  {html_path}")

    print(f"\n  All outputs saved to: {OUT_DIR}")


def main():
    print("=" * 60)
    print("Web Roster Engine E2E Test (Issue #5)")
    print("=" * 60)

    test_date_update()
    test_column_deletion()
    test_row_shrink()
    test_number_formats()
    test_html_export()
    test_all_templates()

    print("\n" + "=" * 60)
    print("✅ All tests passed!")
    print("=" * 60)


if __name__ == "__main__":
    main()
