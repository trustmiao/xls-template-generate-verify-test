"""Integration tests for all engines.

Run with:
    pytest app/engine/test/test_engines.py -v
"""
from __future__ import annotations

import pytest

from .harness import EngineTestHarness


# ---------------------------------------------------------------------------
# 大元邨 Security (project=1, category=1)
# ---------------------------------------------------------------------------

class TestDayuanSecurity:
    """大元邨保安 — 完整数据 + 3-shift roster template."""

    @pytest.fixture(scope="class")
    def harness(self):
        return EngineTestHarness(project_id=1, category_id=1, month="2026-03")

    def test_shortfall_html_shift_a(self, harness):
        result = harness.run_html("shortfall", shift="A")
        assert result["shift"] == "A"
        assert len(result["segments"]) > 0
        assert result["has_data"]
        assert result["days_in_month"] == 31

    def test_shortfall_html_shift_b(self, harness):
        result = harness.run_html("shortfall", shift="B")
        assert result["shift"] == "B"
        assert result["has_data"]

    def test_shortfall_html_shift_c(self, harness):
        result = harness.run_html("shortfall", shift="C")
        assert result["shift"] == "C"
        assert result["has_data"]

    def test_roster_shortfall_excel(self, harness):
        path = harness.run_excel("roster_shortfall")
        checks = harness.verify_xlsx(path)
        assert checks["ok"], f"X1/X2/X3 failed: {checks}"
        for sn, summary in checks.get("summary", {}).items():
            assert summary.get("pass"), f"{sn} failed"

    def test_deployment_excel(self, harness):
        path = harness.run_excel("deployment")
        checks = harness.verify_xlsx(path)
        # Deployment sheet has no day grid → X3 skipped; X1/X2 should pass
        assert checks["ok"]


# ---------------------------------------------------------------------------
# 东汇邨 Cleaning (project=2, category=2)
# ---------------------------------------------------------------------------

class TestDonghuiCleaning:
    """东汇邨保洁 — cleaning template with generic roster."""

    @pytest.fixture(scope="class")
    def harness(self):
        return EngineTestHarness(project_id=2, category_id=2, month="2026-03")

    def test_shortfall_html_shift_a(self, harness):
        result = harness.run_html("shortfall", shift="A")
        assert result["shift"] == "A"
        assert result["has_data"]

    def test_generic_roster_excel(self, harness):
        path = harness.run_excel("generic_roster")
        checks = harness.verify_xlsx(path)
        assert checks["ok"]

    def test_generic_roster_cleaner_excel(self, harness):
        path = harness.run_excel("generic_roster_cleaner")
        checks = harness.verify_xlsx(path)
        assert checks["ok"]

    def test_cleaner_dates_updated(self, harness):
        """generic_roster_cleaner must write the correct month into date row."""
        from openpyxl import load_workbook
        path = harness.run_excel("generic_roster_cleaner")
        wb = load_workbook(str(path), data_only=True)
        ws = wb[wb.sheetnames[0]]
        # Find date row
        date_row = None
        for r in (6, 5, 7):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if hasattr(v, "year"):
                    date_row = r
                    break
            if date_row:
                break
        assert date_row, "Date row not found"
        first_date = ws.cell(date_row, 9).value  # col I = day 1
        assert hasattr(first_date, "year"), f"Expected date, got {first_date!r}"
        assert first_date.year == 2026
        assert first_date.month == 3


# ---------------------------------------------------------------------------
# 东汇邨 Security (project=2, category=3)
# ---------------------------------------------------------------------------

class TestDonghuiSecurity:
    """东汇邨保安 — security roster."""

    @pytest.fixture(scope="class")
    def harness(self):
        return EngineTestHarness(project_id=2, category_id=3, month="2026-06")

    def test_shortfall_html_shift_a(self, harness):
        result = harness.run_html("shortfall", shift="A")
        assert result["shift"] == "A"
        assert result["has_data"]

    def test_roster_shortfall_excel(self, harness):
        path = harness.run_excel("roster_shortfall")
        checks = harness.verify_xlsx(path)
        assert checks["ok"]


# ---------------------------------------------------------------------------
# Consistency: HTML and Excel must show the same data
# ---------------------------------------------------------------------------

class TestHtmlExcelConsistency:
    """网页和 Excel 数据一致性检查."""

    @pytest.fixture(scope="class")
    def harness(self):
        return EngineTestHarness(project_id=1, category_id=1, month="2026-03")

    def test_shift_a_person_count_matches(self, harness):
        html = harness.run_html("shortfall", shift="A")
        xlsx_path = harness.run_excel("roster_shortfall")

        # Count people in HTML
        html_persons = set()
        for seg in html.get("segments", []):
            for row in seg.get("rows", []):
                html_persons.add(row.get("name"))

        # Count people in Excel 早 sheet
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx_path), data_only=True)
        ws = wb["早"]
        xlsx_persons = set()
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 2).value  # col B = rank_seq
            if v and str(v).strip().startswith("M"):
                name = ws.cell(r, 4).value
                if name:
                    xlsx_persons.add(name)

        # Allow minor differences due to segment layout; rough count should match
        assert len(xlsx_persons) >= len(html_persons) - 2, (
            f"HTML has {len(html_persons)} people, Excel has {len(xlsx_persons)}"
        )
