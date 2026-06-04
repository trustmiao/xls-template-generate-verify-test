"""Test harness for running engines standalone.

Usage:
    h = EngineTestHarness(project_id=1, category_id=1, month="2026-03")
    html = h.run_html("shortfall", shift="A")
    xlsx_path = h.run_excel("roster_shortfall")
    checks = h.verify_xlsx(xlsx_path)
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from ...config import DOC_DIR, OUTPUT_DIR
from ..common.data_source import (
    build_context,
    load_pages,
    resolve_template,
    resolve_template_for_engine,
)
from ..registry import EXCEL_ENGINES, HTML_ENGINES


class EngineTestHarness:
    """Load data + template and run any HTML or Excel engine for testing."""

    def __init__(self, project_id: int, category_id: int, month: str):
        self.project_id = project_id
        self.category_id = category_id
        self.month = month

    # ------------------------------------------------------------------
    # Data / Template loaders
    # ------------------------------------------------------------------

    def load_template(self, template_type: str = "roster_shortfall") -> Path:
        """Resolve template path for this project/category/month."""
        path, _ = resolve_template(self.project_id, self.category_id, self.month, template_type)
        if not path:
            raise FileNotFoundError(
                f"No template for p={self.project_id} c={self.category_id} m={self.month}"
            )
        return path

    def load_data(self, shift: str) -> List[Dict[str, Any]]:
        """Load pages with entries for a shift."""
        return load_pages(
            shift=shift,
            project_id=self.project_id,
            category_id=self.category_id,
            month=self.month,
        )

    # ------------------------------------------------------------------
    # HTML engines
    # ------------------------------------------------------------------

    def run_html(self, engine_name: str, shift: str) -> Dict[str, Any]:
        """Run an HTML engine and return JSON."""
        engine = HTML_ENGINES.get(engine_name)
        if not engine:
            raise ValueError(f"Unknown HTML engine: {engine_name}")
        return engine(
            shift=shift,
            project_id=self.project_id,
            category_id=self.category_id,
            month=self.month,
        )

    # ------------------------------------------------------------------
    # Excel engines
    # ------------------------------------------------------------------

    def run_excel(
        self,
        engine_name: str,
        template_path: Optional[Path] = None,
        out_name: Optional[str] = None,
    ) -> Path:
        """Run an Excel engine and return the saved xlsx path.

        Args:
            engine_name: builtin key (e.g. "roster_shortfall", "generic_roster")
            template_path: optional explicit template; defaults to DB lookup
            out_name: optional output filename; defaults to test_<engine>_<month>.xlsx
        """
        engine = EXCEL_ENGINES.get(engine_name)
        if not engine:
            raise ValueError(f"Unknown Excel engine: {engine_name}")

        if template_path is None:
            template_path = self.load_template()

        wb = load_workbook(str(template_path), data_only=False)
        context = build_context(self.project_id, self.category_id, self.month)
        context["template_path"] = str(template_path)
        engine(wb, context)

        safe_name = out_name or f"test_{engine_name}_{self.month}.xlsx"
        out_path = OUTPUT_DIR / safe_name
        wb.save(str(out_path))
        return out_path

    def run_excel_by_engine_id(self, engine_id: int) -> Path:
        """Run an Excel engine by its DB engine_id (uses template from DB config)."""
        from ...models import get_output_engine
        engine_config = get_output_engine(engine_id)
        if not engine_config:
            raise ValueError(f"Engine {engine_id} not found")
        builtin_key = engine_config.get("builtin_key")
        template_path, _ = resolve_template_for_engine(engine_config, self.month)
        if not template_path:
            raise FileNotFoundError(f"No template for engine {engine_id}")
        return self.run_excel(builtin_key, template_path=template_path)

    # ------------------------------------------------------------------
    # Verification helpers
    # ------------------------------------------------------------------

    def verify_xlsx(self, path: Path) -> Dict[str, Any]:
        """Run X1/X2/X3 checks on a generated xlsx."""
        from ...utils.excel_verify import verify_generated_xlsx
        return verify_generated_xlsx(path)

    def read_xlsx_day_cols(self, path: Path, sheet_name: str, row: int) -> List[Any]:
        """Read all day-column values for a given row in a sheet."""
        wb = load_workbook(str(path), data_only=True)
        ws = wb[sheet_name]
        # Find date row to know day_start/day_end
        day_start = day_end = None
        for r in (6, 5, 7):
            if r > ws.max_row:
                continue
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if hasattr(v, "year"):
                    if day_start is None:
                        day_start = c
                    day_end = c
            if day_start:
                break
        if not day_start:
            return []
        return [ws.cell(row, c).value for c in range(day_start, day_end + 1)]
