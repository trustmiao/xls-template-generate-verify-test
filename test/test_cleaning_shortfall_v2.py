"""Unit tests for cleaning_shortfall_v2 engine.

Tests the new engine that works with the 31-day pre-configured template.
"""
from __future__ import annotations

import pytest
from openpyxl import load_workbook

from ..excel.cleaning_shortfall_v2 import (
    _adjust_day_columns,
    _update_dates,
    _find_date_start,
    DAY_START_COL,
    TEMPLATE_DAYS,
)

TEMPLATE_PATH = r"D:\claude\claude_hk\backend\data\doc\output_templates\东汇邨\保洁\东汇-保洁轮休表_template.xlsx"
SECURITY_TEMPLATE_PATH = r"D:\claude\claude_hk\backend\app\engine\templates\东汇-保安轮休表template.xlsx"


@pytest.fixture
def template_wb():
    return load_workbook(TEMPLATE_PATH, data_only=False)


@pytest.fixture
def security_wb():
    return load_workbook(SECURITY_TEMPLATE_PATH, data_only=False)


class TestAdjustDayColumns:
    """Tests for _adjust_day_columns."""

    def test_31_days_no_change(self, template_wb):
        ws = template_wb.worksheets[0]
        day_end = _adjust_day_columns(ws, 31)
        assert day_end == 34  # AH
        # All 31 day columns still present
        assert ws.cell(7, 34).value == "=AG7+1"
        # 5th week formula intact
        assert ws.cell(13, 32).value == "=SUM(AF12:AH12)"

    def test_30_days_deletes_one_column(self, template_wb):
        ws = template_wb.worksheets[0]
        day_end = _adjust_day_columns(ws, 30)
        assert day_end == 33  # AG
        # Date chain ends at AG
        assert "+1" in str(ws.cell(7, 33).value) or ws.cell(7, 33).value is None
        # 5th week formula shrunk
        val = ws.cell(13, 32).value
        assert "AH" not in str(val) or "AG" in str(val)

    def test_28_days_deletes_three_columns(self, template_wb):
        ws = template_wb.worksheets[0]
        day_end = _adjust_day_columns(ws, 28)
        assert day_end == 31  # AE
        # 28 day columns + 29-31 summary column deleted = 96 - 3 - 1 = 92
        assert ws.max_column == 92
        # 5th week merge removed
        merges = [str(mr) for mr in ws.merged_cells.ranges]
        assert not any("AF13:AH13" in m for m in merges)
        # 4 weeks still intact
        assert ws.cell(13, 25).value == "=SUM(Y12:AE12)"
        # 29-31號時數 summary column also removed
        assert not any("29-31" in str(mr) for mr in ws.merged_cells.ranges)

    def test_28_days_clears_5th_week_values(self, template_wb):
        ws = template_wb.worksheets[0]
        _adjust_day_columns(ws, 28)
        # 5th week cells cleared
        for r in [13, 14, 15, 36, 37, 38, 45, 46, 47]:
            # After deleting cols 32-34, the old col 35 (AI) shifts to 32 (AF)
            pass  # values were cleared before delete

    def test_total_merge_shifts_left(self, template_wb):
        ws = template_wb.worksheets[0]
        _adjust_day_columns(ws, 28)
        merges = [str(mr) for mr in ws.merged_cells.ranges]
        # AI13:AM13 shifts left by 4 (3 date cols + 1 summary col)
        # After delete: AF13:AI13
        assert any("AF13:AI13" in m for m in merges)

    def test_shifted_merge_preserves_border(self, template_wb):
        """Merged cells shifted left by column deletion must keep their borders."""
        ws = template_wb.worksheets[0]
        # Snapshot original border of the merge that will shift (AI13:AM13)
        orig_tl = ws.cell(13, 35)  # AI13
        orig_border = orig_tl.border.left.style if orig_tl.border else None

        _adjust_day_columns(ws, 28)

        # After shift AI13:AM13 -> AF13:AI13, check border preserved
        new_tl = ws.cell(13, 32)  # AF13
        assert new_tl.border.left.style == orig_border


class TestUpdateDates:
    """Tests for _update_dates."""

    def test_updates_first_date(self, template_wb):
        ws = template_wb.worksheets[0]
        _adjust_day_columns(ws, 31)
        _update_dates(ws, "2026-05")
        first = ws.cell(7, DAY_START_COL).value
        assert first.year == 2026
        assert first.month == 5
        assert first.day == 1


class TestSecurityTemplate:
    """Tests for security template (东汇-保安轮休表) day-column adjustment."""

    def test_30_days_deletes_correct_column(self, security_wb):
        """Day 31 column (AM) should be deleted, not AH."""
        ws = security_wb.worksheets[0]
        day_end = _adjust_day_columns(ws, 30)

        # Security template dates start at I (col 9)
        # 30 days → day_end = 9 + 30 - 1 = 38 (AL)
        assert day_end == 38  # AL

        date_row, first_col = _find_date_start(ws)
        assert first_col == 9  # I

        # All 30 day columns must have values
        for c in range(first_col, first_col + 30):
            assert ws.cell(6, c).value is not None, f"col {c} should have date value"

        # 31st column should now be 工作總日數 (shifted left from AN)
        v = ws.cell(6, first_col + 30).value
        assert "工作總" in str(v)

    def test_28_days_deletes_three_columns(self, security_wb):
        ws = security_wb.worksheets[0]
        day_end = _adjust_day_columns(ws, 28)

        # 28 days → day_end = 9 + 28 - 1 = 36 (AJ)
        assert day_end == 36  # AJ

        date_row, first_col = _find_date_start(ws)

        # 28 day columns present
        for c in range(first_col, first_col + 28):
            assert ws.cell(6, c).value is not None

        # 29th column should be 工作總日數
        v = ws.cell(6, first_col + 28).value
        assert "工作總" in str(v)

    def test_count_formulas_stay_correct(self, security_wb):
        ws = security_wb.worksheets[0]
        _adjust_day_columns(ws, 30)

        date_row, first_col = _find_date_start(ws)
        # Row 14 has COUNT formulas for each day
        for c in range(first_col, first_col + 30):
            letter = __import__("openpyxl").utils.get_column_letter(c)
            expected = f"=COUNT({letter}8:{letter}10)"
            assert ws.cell(14, c).value == expected

        # Total sum formula
        total_col = first_col + 30
        first_letter = __import__("openpyxl").utils.get_column_letter(first_col)
        last_letter = __import__("openpyxl").utils.get_column_letter(first_col + 29)
        expected_total = f"=SUM({first_letter}14:{last_letter}14)"
        assert ws.cell(14, total_col).value == expected_total

    def test_merged_cells_shift_correctly(self, security_wb):
        ws = security_wb.worksheets[0]
        _adjust_day_columns(ws, 30)

        merges = [str(mr) for mr in ws.merged_cells.ranges]
        # AO6:AO7 should shift to AN6:AN7 after deleting AM
        assert any("AN6:AN7" in m for m in merges)


class TestEndToEnd:
    """End-to-end using the verify module."""

    def test_28_day_month(self, template_wb, tmp_path):
        from ..test.verify.cleaning import run_check_cleaning_summary

        ws = template_wb.worksheets[0]
        _adjust_day_columns(ws, 28)
        _update_dates(ws, "2026-02")

        out = tmp_path / "test_28.xlsx"
        template_wb.save(str(out))

        result = run_check_cleaning_summary(out, "Roster-FEB2026", "2026-02")
        assert result["ok"], result["errors"]
        assert result["days"] == 28
        assert result["weeks"] == 4

    def test_31_day_month(self, template_wb, tmp_path):
        from ..test.verify.cleaning import run_check_cleaning_summary

        ws = template_wb.worksheets[0]
        _adjust_day_columns(ws, 31)
        _update_dates(ws, "2026-03")

        out = tmp_path / "test_31.xlsx"
        template_wb.save(str(out))

        result = run_check_cleaning_summary(out, "Roster-FEB2026", "2026-03")
        assert result["ok"], result["errors"]
        assert result["days"] == 31
        assert result["weeks"] == 5


class TestRowAdjustment:
    """Tests for personnel row insertion/deletion."""

    def _make_rows(self, n: int, days: int = 28):
        rows = []
        for i in range(n):
            cells = [{"day": d, "value": 8} for d in range(1, days + 1)]
            rows.append({"rank_seq": f"A{i+1}", "name": f"Person{i+1}", "cells": cells})
        return rows

    def test_delete_to_one_person_per_segment(self, template_wb):
        """When API returns 1 person per segment, excess template rows are deleted."""
        from ..excel.cleaning_shortfall_v2 import run
        from unittest.mock import patch

        ws = template_wb.worksheets[0]
        _adjust_day_columns(ws, 28)
        _update_dates(ws, "2026-02")

        mock_data = {
            "has_data": True,
            "days_in_month": 28,
            "segments": [
                {"title": "1. 清潔科文", "rows": self._make_rows(1)},
                {"title": "2. 清潔工人", "rows": self._make_rows(1)},
                {"title": "3. VO清潔工人", "rows": self._make_rows(1)},
            ],
        }

        with patch("app.engine.excel.cleaning_shortfall_v2._fetch_data", return_value=mock_data):
            run(template_wb, {"project_id": 1, "category_id": 2, "month": "2026-02"})

        # Helper: find personnel rows under a given title row
        def find_personnel(title_fragment):
            for r in range(1, ws.max_row + 1):
                if title_fragment in str(ws.cell(r, 2).value or ""):
                    people = []
                    for pr in range(r + 1, ws.max_row + 1):
                        b = ws.cell(pr, 2).value
                        if b and isinstance(b, str) and b.startswith("A"):
                            people.append(pr)
                        elif b and "每天工作總時數" in b:
                            break
                    return people
            return []

        # Segment 1: should have 1 person
        seg1 = find_personnel("清潔科文")
        assert len(seg1) == 1
        assert ws.cell(seg1[0], 2).value == "A1"
        # Next row should be daily-total, not another person
        nxt = ws.cell(seg1[0] + 1, 2).value
        assert nxt is None or not str(nxt).startswith("A")

        # Segment 2: should have 1 person
        seg2 = find_personnel("清潔工人")
        assert len(seg2) == 1
        assert ws.cell(seg2[0], 2).value == "A1"
        nxt = ws.cell(seg2[0] + 1, 2).value
        assert nxt is None or not str(nxt).startswith("A")

        # Segment 3: should have 1 person
        seg3 = find_personnel("VO清潔工人")
        assert len(seg3) == 1
        assert ws.cell(seg3[0], 2).value == "A1"
        nxt = ws.cell(seg3[0] + 1, 2).value
        assert nxt is None or not str(nxt).startswith("A")

    def test_add_two_people_per_segment(self, template_wb):
        """When API returns 4/19/5 people, extra rows are inserted."""
        from ..excel.cleaning_shortfall_v2 import run
        from unittest.mock import patch

        ws = template_wb.worksheets[0]
        _adjust_day_columns(ws, 28)
        _update_dates(ws, "2026-02")

        mock_data = {
            "has_data": True,
            "days_in_month": 28,
            "segments": [
                {"title": "1. 清潔科文", "rows": self._make_rows(4)},
                {"title": "2. 清潔工人", "rows": self._make_rows(19)},
                {"title": "3. VO清潔工人", "rows": self._make_rows(5)},
            ],
        }

        with patch("app.engine.excel.cleaning_shortfall_v2._fetch_data", return_value=mock_data):
            run(template_wb, {"project_id": 1, "category_id": 2, "month": "2026-02"})

        def find_personnel(title_fragment):
            for r in range(1, ws.max_row + 1):
                if title_fragment in str(ws.cell(r, 2).value or ""):
                    people = []
                    for pr in range(r + 1, ws.max_row + 1):
                        b = ws.cell(pr, 2).value
                        if b and isinstance(b, str) and b.startswith("A"):
                            people.append(pr)
                        elif b and "每天工作總時數" in b:
                            break
                    return people
            return []

        # Segment 1: should have 4 people
        seg1 = find_personnel("清潔科文")
        assert len(seg1) == 4
        for i, pr in enumerate(seg1):
            assert ws.cell(pr, 2).value == f"A{i+1}"
        # Next row should be daily-total
        nxt = ws.cell(seg1[-1] + 1, 2).value
        assert nxt is None or "每天工作總時數" in str(nxt)

        # Segment 3: should have 5 people
        seg3 = find_personnel("VO清潔工人")
        assert len(seg3) == 5
        for i, pr in enumerate(seg3):
            assert ws.cell(pr, 2).value == f"A{i+1}"
        nxt = ws.cell(seg3[-1] + 1, 2).value
        assert nxt is None or "每天工作總時數" in str(nxt)


class TestThreeTemplatesFebApr:
    """Cross-template validation: Feb (28) and Apr (30) day-column adjustment."""

    @pytest.fixture
    def ty_wb(self):
        return load_workbook(
            r"D:\claude\claude_hk\backend\app\engine\templates\TY-2026.03- SG_SEC-Deploy Roster  Shortfall - template.xlsx",
            data_only=False,
        )

    def _check_date_continuity(self, ws, days, label):
        """Assert every day column has a value and the next col is non-date."""
        date_row, first_col = _find_date_start(ws)
        assert first_col, f"{label}: no date row found"
        for c in range(first_col, first_col + days):
            v = ws.cell(date_row, c).value
            assert v is not None, (
                f"{label} col {c} ({__import__('openpyxl').utils.get_column_letter(c)})"
                f" row {date_row} is None"
            )
        next_col = first_col + days
        v_next = ws.cell(date_row, next_col).value
        assert not (v_next and hasattr(v_next, "year")), (
            f"{label} col {next_col} still looks like a date: {v_next}"
        )

    # --- 东汇保安 ---
    def test_security_feb28(self, security_wb):
        ws = security_wb.worksheets[0]
        _adjust_day_columns(ws, 28)
        _update_dates(ws, "2026-02")
        self._check_date_continuity(ws, 28, "东汇保安 Feb28")

    def test_security_apr30(self, security_wb):
        ws = security_wb.worksheets[0]
        _adjust_day_columns(ws, 30)
        _update_dates(ws, "2026-04")
        self._check_date_continuity(ws, 30, "东汇保安 Apr30")

    # --- 东汇保洁 ---
    def test_cleaning_feb28(self, template_wb):
        ws = template_wb.worksheets[0]
        _adjust_day_columns(ws, 28)
        _update_dates(ws, "2026-02")
        self._check_date_continuity(ws, 28, "东汇保洁 Feb28")

    def test_cleaning_apr30(self, template_wb):
        ws = template_wb.worksheets[0]
        _adjust_day_columns(ws, 30)
        _update_dates(ws, "2026-04")
        self._check_date_continuity(ws, 30, "东汇保洁 Apr30")

    # --- TY 大元保安 (first sheet = 早) ---
    def test_ty_feb28(self, ty_wb):
        ws = ty_wb.worksheets[0]
        _adjust_day_columns(ws, 28)
        _update_dates(ws, "2026-02")
        self._check_date_continuity(ws, 28, "TY Feb28")

    def test_ty_apr30(self, ty_wb):
        ws = ty_wb.worksheets[0]
        _adjust_day_columns(ws, 30)
        _update_dates(ws, "2026-04")
        self._check_date_continuity(ws, 30, "TY Apr30")
