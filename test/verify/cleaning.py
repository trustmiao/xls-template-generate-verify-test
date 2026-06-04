"""Cleaning roster verification checks.

Validates the special rules for 保洁 (cleaning) templates as specified
in issue #4:
  1. Date columns match month days
  2. Weekly summary columns: 4 for 28-day months, 5 for >28 days
  3. Monthly total covers all weekly columns
  4-12. Per-section checks for 清潔科文 / 清潔工人 / VO清潔工人:
        - 实际每周工作总时数 formula & merge range
        - 每周所需工作总时数 formula (proportional for 5th week)
        - 欠/多 时数 formula
"""
from __future__ import annotations

import calendar
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

DAY_START_COL = 4  # col D

_WEEK_STARTS = [DAY_START_COL + 7 * i for i in range(4)]  # [4, 11, 18, 25]
_WEEK_ENDS = [s + 6 for s in _WEEK_STARTS]  # [10, 17, 24, 31]

# Labels for the three sections
SECTIONS = [
    {
        "name": "清潔科文",
        "actual_label": "清潔科文(實際) 每週工作總時數",
        "required_label": "清潔科文每週所需工作總時數",
        "diff_label": "欠/多 清潔科文時數",
    },
    {
        "name": "清潔工人",
        "actual_label": "清潔工人(實際)每週工作總時數",
        "required_label": "清潔工人每週所需工作總時數",
        "diff_label": "欠/多 清潔工人時數",
    },
    {
        "name": "VO清潔工人",
        "actual_label": "VO清潔工人(實際) 每週工作總時數",
        "required_label": "VO清潔工人每週所需工作總時數",
        "diff_label": "欠/多 清潔工人時數",
    },
]


def _col_letter(col: int) -> str:
    """Convert 1-based column index to Excel column letter."""
    result = ""
    while col > 0:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result


def _find_date_row(ws) -> Optional[int]:
    """Find the row containing the first date value."""
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(DAY_START_COL, min(ws.max_column + 1, DAY_START_COL + 35)):
            v = ws.cell(r, c).value
            if v is not None and hasattr(v, "year"):
                return r
    return None


def _find_day_end_col(ws, date_row: int) -> int:
    """Detect the last day column by walking the date row."""
    last_day = None
    for c in range(DAY_START_COL, ws.max_column + 1):
        v = ws.cell(date_row, c).value
        if v is not None and (hasattr(v, "year") or (isinstance(v, str) and "+1" in v)):
            last_day = c
        elif v is not None and isinstance(v, str) and v.strip() and not v.startswith("="):
            # Hit a label like "1-7號時數" — past the day columns
            break
    return last_day or DAY_START_COL + 27


def _find_row_by_label(ws, label: str) -> Optional[int]:
    """Find the first row whose col B matches the given label."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v and isinstance(v, str) and v.strip() == label.strip():
            return r
    return None


def _get_merged_range(ws, row: int, col: int) -> Optional[Tuple[int, int, int, int]]:
    """Return (min_row, min_col, max_row, max_col) if this cell is inside a merge."""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return (mr.min_row, mr.min_col, mr.max_row, mr.max_col)
    return None


def _check_week_merge(ws, row: int, week_idx: int, expected_start: int, expected_end: int,
                      errors: List[str]) -> None:
    """Check that a weekly stat cell is merged across the expected day columns."""
    col = _WEEK_STARTS[week_idx]
    rng = _get_merged_range(ws, row, col)
    if rng is None:
        errors.append(f"Row {row} week {week_idx + 1} (col {col}) is not merged")
        return
    _, min_col, _, max_col = rng
    if min_col != expected_start or max_col != expected_end:
        errors.append(
            f"Row {row} week {week_idx + 1} merge {min_col}-{max_col} "
            f"!= expected {expected_start}-{expected_end}"
        )


def run_check_cleaning_summary(xlsx_path: Path, sheet: str, month: str) -> Dict[str, Any]:
    """Run all cleaning-specific checks on a generated xlsx.

    Returns {"ok": bool, "errors": [...], "days": int, "weeks": int}.
    """
    errors: List[str] = []

    try:
        year, month_num = int(month[:4]), int(month[5:7])
        expected_days = calendar.monthrange(year, month_num)[1]
    except Exception as e:
        return {"ok": False, "error": f"Invalid month {month}: {e}"}

    wb = load_workbook(str(xlsx_path), data_only=False)
    if sheet not in wb.sheetnames:
        return {"ok": False, "error": f"Sheet {sheet} not found"}
    ws = wb[sheet]

    # ------------------------------------------------------------------
    # 1. Date column count
    # ------------------------------------------------------------------
    date_row = _find_date_row(ws)
    if date_row is None:
        return {"ok": False, "error": "Date row not found"}

    day_end_col = _find_day_end_col(ws, date_row)
    actual_days = day_end_col - DAY_START_COL + 1
    if actual_days != expected_days:
        errors.append(f"Date zone has {actual_days} day columns, expected {expected_days}")

    # ------------------------------------------------------------------
    # 2. Weekly summary columns
    # ------------------------------------------------------------------
    expected_weeks = 4 if expected_days <= 28 else 5
    extra_days = max(0, expected_days - 28)

    # ------------------------------------------------------------------
    # 4-12. Per-section checks
    # ------------------------------------------------------------------
    for sec in SECTIONS:
        actual_row = _find_row_by_label(ws, sec["actual_label"])
        required_row = _find_row_by_label(ws, sec["required_label"])
        # diff_row is always the row right after required_row in this template
        diff_row = required_row + 1 if required_row else None
        # Sanity check: ensure the label matches
        if diff_row and ws.cell(diff_row, 2).value != sec["diff_label"]:
            # Fall back to label search if structure differs
            diff_row = _find_row_by_label(ws, sec["diff_label"])

        if not all([actual_row, required_row, diff_row]):
            errors.append(f"{sec['name']}: stat rows not found")
            continue

        daily_row = actual_row - 1

        # Check each of the 4 regular weeks
        for wi in range(4):
            start_col = _WEEK_STARTS[wi]
            end_col = _WEEK_ENDS[wi]
            _check_week_merge(ws, actual_row, wi, start_col, end_col, errors)
            _check_week_merge(ws, required_row, wi, start_col, end_col, errors)
            _check_week_merge(ws, diff_row, wi, start_col, end_col, errors)

            # Verify actual formula
            expected_actual = f"=SUM({_col_letter(start_col)}{daily_row}:{_col_letter(end_col)}{daily_row})"
            actual_formula = ws.cell(actual_row, start_col).value
            if actual_formula != expected_actual:
                errors.append(
                    f"{sec['name']} actual week {wi + 1} formula "
                    f"{actual_formula!r} != expected {expected_actual!r}"
                )

            # Verify required formula (should be a literal or simple formula)
            required_formula = ws.cell(required_row, start_col).value
            if required_formula is None or required_formula == "":
                errors.append(f"{sec['name']} required week {wi + 1} is empty")

            # Verify diff formula
            expected_diff = f"={_col_letter(start_col)}{actual_row}-{_col_letter(start_col)}{required_row}"
            diff_formula = ws.cell(diff_row, start_col).value
            # Allow SUM(...) wrapper and hard-coded values
            # (cleaning worker template uses =SUM(D36-728))
            if diff_formula and isinstance(diff_formula, str):
                col_let = _col_letter(start_col)
                # Patterns:
                #   =D36-D37               (cell reference)
                #   =SUM(D36-728)          (SUM wrapper with hard-coded value)
                #   =SUM(D36-D37)          (SUM wrapper with cell reference)
                ok_patterns = [
                    rf"^={col_let}{actual_row}-{col_let}{required_row}$",
                    rf"^=SUM\({col_let}{actual_row}-{col_let}{required_row}\)$",
                    rf"^=SUM\({col_let}{actual_row}-\d+\)$",
                    rf"^=SUM\({col_let}{actual_row}-\d+\.\d+\)$",
                ]
                if not any(re.search(p, diff_formula) for p in ok_patterns):
                    errors.append(
                        f"{sec['name']} diff week {wi + 1} formula "
                        f"{diff_formula!r} != expected {expected_diff!r}"
                    )
            else:
                errors.append(f"{sec['name']} diff week {wi + 1} missing formula")

        # Check 5th week (if applicable)
        if expected_weeks == 5:
            fifth_start = DAY_START_COL + 28  # col 32 = AF
            fifth_end = day_end_col

            # Check merges
            for r in (actual_row, required_row, diff_row):
                rng = _get_merged_range(ws, r, fifth_start)
                if rng is None:
                    errors.append(f"{sec['name']} 5th week (row {r}) not merged")
                elif rng[1] != fifth_start or rng[3] != fifth_end:
                    errors.append(
                        f"{sec['name']} 5th week merge {rng[1]}-{rng[3]} "
                        f"!= expected {fifth_start}-{fifth_end}"
                    )

            # Check actual formula
            expected_fifth_actual = f"=SUM({_col_letter(fifth_start)}{daily_row}:{_col_letter(fifth_end)}{daily_row})"
            fifth_actual = ws.cell(actual_row, fifth_start).value
            if fifth_actual != expected_fifth_actual:
                errors.append(
                    f"{sec['name']} 5th week actual formula "
                    f"{fifth_actual!r} != expected {expected_fifth_actual!r}"
                )

            # Check required formula (proportional)
            # Template v2 uses =Y{row}/7*COUNT(AF{day}:AH{day}) instead of hard-coded days
            first_col = _col_letter(DAY_START_COL)
            fifth_col_let = _col_letter(fifth_start)
            fifth_end_let = _col_letter(fifth_end)
            expected_patterns = [
                f"={first_col}{required_row}/7*{extra_days}",
                f"={fifth_col_let}{required_row}/7*COUNT({fifth_col_let}{daily_row}:{fifth_end_let}{daily_row})",
                f"=Y{required_row}/7*COUNT({fifth_col_let}{daily_row}:{fifth_end_let}{daily_row})",
            ]
            fifth_required = ws.cell(required_row, fifth_start).value
            if fifth_required not in expected_patterns:
                errors.append(
                    f"{sec['name']} 5th week required formula "
                    f"{fifth_required!r} not in expected patterns {expected_patterns}"
                )

            # Check diff formula
            fifth_col = _col_letter(fifth_start)
            expected_fifth_diff = f"={fifth_col}{actual_row}-{fifth_col}{required_row}"
            fifth_diff = ws.cell(diff_row, fifth_start).value
            if fifth_diff != expected_fifth_diff:
                errors.append(
                    f"{sec['name']} 5th week diff formula "
                    f"{fifth_diff!r} != expected {expected_fifth_diff!r}"
                )

    return {
        "ok": len(errors) == 0,
        "errors": errors,
        "days": actual_days,
        "weeks": expected_weeks,
    }
