"""Mark compare_template diffs in Excel with red background.

Reads the latest test report JSON to find xlsx/sheet mappings,
then for each compare CSV: copy the sheet, mark non-match cells red.
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

OUTPUT_DIR = Path(__file__).parent / "outputs"
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def _addr_to_cell(ws, addr: str):
    """Parse A1 notation like 'AN59' into openpyxl cell."""
    m = re.fullmatch(r"([A-Z]+)(\d+)", addr or "")
    if not m:
        return None
    return ws[f"{m.group(1)}{m.group(2)}"]


def main() -> int:
    reports = sorted(OUTPUT_DIR.glob("report_*.json"))
    if not reports:
        print("No report JSON found")
        return 1
    report_path = reports[-1]
    print(f"Using report: {report_path.name}")
    data = json.loads(report_path.read_text(encoding="utf-8"))

    marked = 0
    for res in data.get("results", []):
        label = res["label"]
        safe_label = label.replace("/", "_").replace("\\", "_")
        xlsx_name = res.get("output", {}).get("path", "")
        if not xlsx_name:
            continue
        xlsx_path = OUTPUT_DIR / xlsx_name
        if not xlsx_path.exists():
            print(f"  SKIP: xlsx not found: {xlsx_path}")
            continue

        wb = openpyxl.load_workbook(xlsx_path)
        any_change = False

        for shift, v in res.get("verify_excel", {}).items():
            if shift.startswith("_") or shift == "skipped":
                continue
            dr = v.get("detect_regions", {})
            sheet = dr.get("sheet")
            if not sheet or sheet not in wb.sheetnames:
                continue

            csv_path = OUTPUT_DIR / f"{safe_label}_{shift}_compare.csv"
            if not csv_path.exists():
                continue

            # Count diffs
            diff_addrs = []
            with open(csv_path, encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row.get("status") == "match":
                        continue
                    addr = row.get("addr_gen", "").strip()
                    if addr:
                        diff_addrs.append(addr)

            if not diff_addrs:
                print(f"  {label} {shift}({sheet}): no diffs")
                continue

            # Copy sheet
            src_ws = wb[sheet]
            diff_sheet_name = f"{sheet}_diff"
            # Remove existing diff sheet if present
            if diff_sheet_name in wb.sheetnames:
                del wb[diff_sheet_name]
            dst_ws = wb.copy_worksheet(src_ws)
            dst_ws.title = diff_sheet_name

            for addr in diff_addrs:
                cell = _addr_to_cell(dst_ws, addr)
                if cell:
                    cell.fill = RED_FILL

            print(f"  {label} {shift}({sheet}): {len(diff_addrs)} cells marked red -> sheet '{diff_sheet_name}'")
            marked += len(diff_addrs)
            any_change = True

        if any_change:
            wb.save(xlsx_path)
            print(f"  Saved: {xlsx_path.name}")

    print(f"\nTotal diff cells marked: {marked}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
