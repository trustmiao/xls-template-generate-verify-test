"""Unit tests for cleaning_shortfall engine fixes (issue #4).

These tests verify:
  1. Date columns extend correctly for months > 28 days
  2. 5th-week formulas are added with proper merge ranges
  3. Required-hours 5th week uses proportional formula
  4. Diff formulas reference the correct cells
"""
from __future__ import annotations

import pytest
from openpyxl import load_workbook

from ..excel.cleaning_shortfall import (
    _adjust_cleaning_day_columns,
    _add_fifth_week_formulas,
    DAY_START_COL,
)


TEMPLATE_PATH = r"D:\claude\claude_hk\backend\data\doc\output_templates\东汇邨\保洁\东汇-保洁轮休表.xlsx"


@pytest.fixture
def template_wb():
    return load_workbook(TEMPLATE_PATH, data_only=False)


class TestAdjustDayColumns:
    """Tests for _adjust_cleaning_day_columns."""

    def test_28_days_no_change(self, template_wb):
        ws = template_wb.worksheets[0]
        day_end, date_row, weekday_row = _adjust_cleaning_day_columns(ws, 28)
        assert day_end == 31  # AE
        assert date_row == 7
        assert weekday_row == 9
        # No extra columns inserted
        assert ws.cell(7, 32).value == "1-7號\n時數"  # col AF

    def test_31_days_extends_date_chain(self, template_wb):
        ws = template_wb.worksheets[0]
        day_end, date_row, weekday_row = _adjust_cleaning_day_columns(ws, 31)
        assert day_end == 34  # AH
        # Date chain continues through new columns
        assert ws.cell(7, 32).value == "=AE7+1"   # AF
        assert ws.cell(7, 33).value == "=AF7+1"   # AG
        assert ws.cell(7, 34).value == "=AG7+1"   # AH
        # Weekday chain continues
        assert "AF7" in str(ws.cell(9, 32).value)  # AF
        assert "AG7" in str(ws.cell(9, 33).value)  # AG
        assert "AH7" in str(ws.cell(9, 34).value)  # AH

    def test_31_days_keeps_4th_week_intact(self, template_wb):
        ws = template_wb.worksheets[0]
        _adjust_cleaning_day_columns(ws, 31)
        # 4th week (days 22-28) should still end at AE (col 31)
        y13 = ws.cell(13, 25).value
        assert "AE12" in str(y13) or "AE13" in str(y13)

    def test_31_days_shifts_merged_totals(self, template_wb):
        ws = template_wb.worksheets[0]
        _adjust_cleaning_day_columns(ws, 31)
        # Total merge should shift from AF:AI to AI:AL
        merges = [str(mr) for mr in ws.merged_cells.ranges]
        assert any("AI13:AL13" in m for m in merges)


class TestFifthWeekFormulas:
    """Tests for _add_fifth_week_formulas."""

    def test_28_days_skips(self, template_wb):
        ws = template_wb.worksheets[0]
        _adjust_cleaning_day_columns(ws, 28)
        _add_fifth_week_formulas(ws, 31, 28)
        # No 5th week formulas added
        assert ws.cell(13, 32).value != "=SUM(AF12:AH12)"

    def test_31_days_adds_5th_week_actual(self, template_wb):
        ws = template_wb.worksheets[0]
        day_end, _, _ = _adjust_cleaning_day_columns(ws, 31)
        _add_fifth_week_formulas(ws, day_end, 31)
        assert ws.cell(13, 32).value == "=SUM(AF12:AH12)"

    def test_31_days_adds_5th_week_required(self, template_wb):
        ws = template_wb.worksheets[0]
        day_end, _, _ = _adjust_cleaning_day_columns(ws, 31)
        _add_fifth_week_formulas(ws, day_end, 31)
        # 清潔科文: 8*7 / 7 * 3 = 8*3 = 24
        assert ws.cell(14, 32).value == "=D14/7*3"
        # 清潔工人: 728 / 7 * 3
        assert ws.cell(37, 32).value == "=D37/7*3"
        # VO: 5.25*7 / 7 * 3 = 5.25*3
        assert ws.cell(46, 32).value == "=D46/7*3"

    def test_31_days_adds_5th_week_diff(self, template_wb):
        ws = template_wb.worksheets[0]
        day_end, _, _ = _adjust_cleaning_day_columns(ws, 31)
        _add_fifth_week_formulas(ws, day_end, 31)
        assert ws.cell(15, 32).value == "=AF13-AF14"
        assert ws.cell(38, 32).value == "=AF36-AF37"
        assert ws.cell(47, 32).value == "=AF45-AF46"

    def test_31_days_merges_5th_week_cells(self, template_wb):
        ws = template_wb.worksheets[0]
        day_end, _, _ = _adjust_cleaning_day_columns(ws, 31)
        _add_fifth_week_formulas(ws, day_end, 31)
        merges = [(mr.min_row, mr.min_col, mr.max_row, mr.max_col)
                  for mr in ws.merged_cells.ranges]
        # 清潔科文 5th week merges
        assert any(r == 13 and c1 == 32 and c2 == 34 for _, c1, _, c2 in merges
                   for r in [13])
        assert any(r == 14 and c1 == 32 and c2 == 34 for r, c1, _, c2 in merges)
        assert any(r == 15 and c1 == 32 and c2 == 34 for r, c1, _, c2 in merges)


class TestCleaningVerify:
    """End-to-end test using the verify module."""

    def test_31_day_passes_verification(self, template_wb, tmp_path):
        from ..test.verify.cleaning import run_check_cleaning_summary

        ws = template_wb.worksheets[0]
        day_end, _, _ = _adjust_cleaning_day_columns(ws, 31)
        _add_fifth_week_formulas(ws, day_end, 31)

        out = tmp_path / "test.xlsx"
        template_wb.save(str(out))

        result = run_check_cleaning_summary(out, "Roster-FEB2026", "2026-03")
        assert result["ok"], result["errors"]
        assert result["days"] == 31
        assert result["weeks"] == 5

    def test_28_day_passes_verification(self, template_wb, tmp_path):
        from ..test.verify.cleaning import run_check_cleaning_summary

        ws = template_wb.worksheets[0]
        day_end, _, _ = _adjust_cleaning_day_columns(ws, 28)
        _add_fifth_week_formulas(ws, day_end, 28)

        out = tmp_path / "test.xlsx"
        template_wb.save(str(out))

        result = run_check_cleaning_summary(out, "Roster-FEB2026", "2026-02")
        assert result["ok"], result["errors"]
        assert result["days"] == 28
        assert result["weeks"] == 4
