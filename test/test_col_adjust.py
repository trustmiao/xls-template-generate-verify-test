"""Unit tests for column adjustment logic."""
from __future__ import annotations

import calendar
from datetime import date

import pytest
from openpyxl import Workbook

from ..common.col_adjust import (
    col_letter_to_num,
    col_num_to_letter,
    shift_formula_cols,
)
from ..excel.roster_shortfall import _scan_sheet, _adjust_day_columns
from ..excel.cleaning_shortfall import _adjust_cleaning_day_columns


class TestColHelpers:
    def test_col_letter_to_num(self):
        assert col_letter_to_num("A") == 1
        assert col_letter_to_num("Z") == 26
        assert col_letter_to_num("AA") == 27
        assert col_letter_to_num("AM") == 39

    def test_col_num_to_letter(self):
        assert col_num_to_letter(1) == "A"
        assert col_num_to_letter(26) == "Z"
        assert col_num_to_letter(27) == "AA"
        assert col_num_to_letter(39) == "AM"


class TestShiftFormulaCols:
    def test_insert(self):
        # B1 (col 2) < col_idx 3, so it stays B1
        assert shift_formula_cols("=A1+B1", 3, 1) == "=A1+B1"
        # C1 (col 3) >= col_idx 3, so it shifts to D1
        assert shift_formula_cols("=SUM(A1:C1)", 3, 1) == "=SUM(A1:D1)"
        assert shift_formula_cols("=D1+E1", 4, 1) == "=E1+F1"

    def test_delete(self):
        assert shift_formula_cols("=A1+D1", 3, -1, is_delete=True) == "=A1+C1"
        # Range shrink
        assert shift_formula_cols("=SUM(A1:D1)", 3, -1, is_delete=True) == "=SUM(A1:C1)"


class TestRosterShortfallColAdjust:
    def test_31day_template_to_28days(self):
        from openpyxl import load_workbook
        from pathlib import Path

        tpl = Path(r"D:\claude\claude_hk\backend\data\doc\output_templates\大元邨\保安\roster_shortfall.xlsx")
        wb = load_workbook(str(tpl), data_only=False)
        ws = wb["早"]
        struct = _scan_sheet(ws, "A")
        diff = _adjust_day_columns(ws, struct, 28)

        assert diff == -3
        assert struct.day_end_col == 36
        # Date row should have 28 formulas
        date_count = 0
        for c in range(struct.day_start_col, struct.day_end_col + 1):
            v = ws.cell(struct.date_row, c).value
            if v is not None:
                date_count += 1
        assert date_count == 28
        # Summary formula should reference new end
        assert "AJ" in str(ws.cell(16, struct.day_end_col + 1).value)

    def test_31day_template_to_31days_no_change(self):
        from openpyxl import load_workbook
        from pathlib import Path

        tpl = Path(r"D:\claude\claude_hk\backend\data\doc\output_templates\大元邨\保安\roster_shortfall.xlsx")
        wb = load_workbook(str(tpl), data_only=False)
        ws = wb["早"]
        struct = _scan_sheet(ws, "A")
        diff = _adjust_day_columns(ws, struct, 31)

        assert diff == 0


class TestCleaningShortfallColAdjust:
    def test_28day_template_to_31days(self):
        from openpyxl import load_workbook
        from pathlib import Path

        tpl = Path(r"D:\claude\claude_hk\backend\data\doc\output_templates\东汇邨\保洁\东汇-保洁轮休表.xlsx")
        wb = load_workbook(str(tpl), data_only=False)
        ws = wb[wb.sheetnames[0]]
        day_end, date_row, weekday_row = _adjust_cleaning_day_columns(ws, 31)

        assert day_end == 34
        # Weekly summary should include new days
        assert "AH" in str(ws.cell(13, 25).value)
        # Total should include new days
        assert "AH" in str(ws.cell(13, 35).value)
        # Daily summary should NOT be changed
        assert str(ws.cell(12, 31).value) == "=SUM(AE10:AE11)"

    def test_28day_template_to_28days_no_change(self):
        from openpyxl import load_workbook
        from pathlib import Path

        tpl = Path(r"D:\claude\claude_hk\backend\data\doc\output_templates\东汇邨\保洁\东汇-保洁轮休表.xlsx")
        wb = load_workbook(str(tpl), data_only=False)
        ws = wb[wb.sheetnames[0]]
        day_end, date_row, weekday_row = _adjust_cleaning_day_columns(ws, 28)

        assert day_end == 31
