#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Delete rows from roster regions in Excel templates and adjust SUM/COUNT formulas.

Usage:
    python delete_roster_rows.py <input.xlsx> <output.xlsx> <region_deletions>

Example:
    python delete_roster_rows.py template.xlsx output.xlsx "1:1,2:3"
    # Deletes 1 row from region 1, 3 rows from region 2
"""

import re
import sys
import copy
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────────────────────────────────────
# Region Detection
# ──────────────────────────────────────────────────────────────────────────────

def detect_roster_regions(ws):
    """
    Detect all roster (personnel list) regions in a worksheet.

    A region is defined as:
      - Title row: B-column starts with "N. " (e.g. "1. 保安主管出勤時數")
      - Data rows: B-column contains a rank code like M01, A1, D01, etc.
      - Ends at the first non-data row below the data block.

    Returns a list of dicts:
        [{
            'index': 1-based region index,
            'title_row': row number of the title,
            'data_start': first data row,
            'data_end':   last data row,
            'title':      title text,
        }, ...]
    """
    regions = []
    rank_pattern = re.compile(r'^[A-Z]+\d+$')
    title_pattern = re.compile(r'^\d+\.\s+')

    for row in range(1, ws.max_row + 1):
        b_val = ws.cell(row=row, column=2).value
        if b_val and isinstance(b_val, str) and title_pattern.match(b_val):
            title_row = row

            # Find first data row (rank code in column B)
            data_start = None
            for r in range(title_row + 1, ws.max_row + 1):
                b = ws.cell(row=r, column=2).value
                if b and isinstance(b, str) and rank_pattern.match(b.strip()):
                    data_start = r
                    break

            if data_start is None:
                continue

            # Find data end: last consecutive row with a rank code in column B
            data_end = data_start
            for r in range(data_start + 1, ws.max_row + 1):
                b = ws.cell(row=r, column=2).value
                is_data = b and isinstance(b, str) and rank_pattern.match(b.strip())
                if is_data:
                    data_end = r
                else:
                    break

            regions.append({
                'index': len(regions) + 1,
                'title_row': title_row,
                'data_start': data_start,
                'data_end': data_end,
                'title': b_val,
            })

    return regions


# ──────────────────────────────────────────────────────────────────────────────
# Formula Adjustment
# ──────────────────────────────────────────────────────────────────────────────

def adjust_formula(formula, deleted_row):
    """
    Adjust all cell references in an Excel formula after a row is deleted.

    Rules:
      - Any row reference > deleted_row is decremented by 1.
      - Row references <= deleted_row are left unchanged.
    """
    # Pattern matches: AN20 or AN8:AN20
    cell_ref_pattern = re.compile(r'([A-Z]+)(\d+)')

    def replace_range(match):
        full = match.group(0)
        if ':' in full:
            parts = full.split(':')
            new_parts = []
            for part in parts:
                m = cell_ref_pattern.match(part)
                if m:
                    col, row_num = m.group(1), int(m.group(2))
                    if row_num > deleted_row:
                        row_num -= 1
                    new_parts.append(f"{col}{row_num}")
                else:
                    new_parts.append(part)
            return ':'.join(new_parts)
        else:
            m = cell_ref_pattern.match(full)
            if m:
                col, row_num = m.group(1), int(m.group(2))
                if row_num > deleted_row:
                    row_num -= 1
                return f"{col}{row_num}"
            return full

    # Match either a range like AN8:AN20 or a single cell like AN20
    return re.sub(r'[A-Z]+\d+(?::[A-Z]+\d+)?', replace_range, formula)


def adjust_all_formulas(ws, deleted_row):
    """Scan entire worksheet and adjust every formula after a row deletion."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = cell.value
            if val and isinstance(val, str) and val.startswith('='):
                cell.value = adjust_formula(val, deleted_row)


# ──────────────────────────────────────────────────────────────────────────────
# Row Deletion
# ──────────────────────────────────────────────────────────────────────────────

def safe_delete_rows(ws, idx, amount=1):
    """
    Delete rows, correctly mimicking Excel behavior.
    Replaces openpyxl's ws.delete_rows() which has an iter_rows bug that
    creates phantom empty cells which then overwrite valid data during shift.
    """
    max_row = max((row for row, col in ws._cells.keys()), default=0)

    for r in range(idx, max_row + 1):
        source_r = r + amount
        if source_r > max_row:
            for key in list(ws._cells.keys()):
                if key[0] == r:
                    del ws._cells[key]
            continue

        target_cols = {col for row, col in ws._cells.keys() if row == r}
        source_cols = {col for row, col in ws._cells.keys() if row == source_r}

        for col in target_cols - source_cols:
            del ws._cells[(r, col)]

        for col in source_cols:
            if (source_r, col) in ws._cells:
                cell = ws._cells[(source_r, col)]
                ws._cells[(r, col)] = cell
                cell.row = r
                del ws._cells[(source_r, col)]

    for r in range(idx, max_row + 1):
        source_r = r + amount
        dst_rd = ws.row_dimensions[r]
        if source_r in ws.row_dimensions:
            src_rd = ws.row_dimensions[source_r]
            dst_rd.height = src_rd.height
            if src_rd._style is not None:
                from copy import copy
                dst_rd._style = copy(src_rd._style)
            else:
                dst_rd._style = None
            dst_rd.hidden = src_rd.hidden
            dst_rd.outline_level = src_rd.outline_level
            dst_rd.collapsed = src_rd.collapsed
        else:
            dst_rd.height = None
            dst_rd._style = None
            dst_rd.hidden = False
            dst_rd.outline_level = 0
            dst_rd.collapsed = False

    ws._current_row = max(row for row, col in ws._cells.keys()) if ws._cells else 0


def delete_rows_in_region(ws, region, n):
    """
    Delete n rows from a roster region, starting at the 2nd data row.

    The region's first data row is kept; rows 2 .. (n+1) are removed.
    After deletion, all formulas in the worksheet are adjusted.

    Returns the list of original row numbers that were deleted.
    """
    data_start = region['data_start']
    # Delete rows data_start+1 through data_start+n
    rows_to_delete = list(range(data_start + 1, data_start + n + 1))

    if not rows_to_delete:
        return []

    # Delete in descending order so row numbers of earlier deletions
    # don't shift the targets of later deletions.
    for row in sorted(rows_to_delete, reverse=True):
        # 1. Adjust formulas first (while row number is still valid in references)
        adjust_all_formulas(ws, row)
        # 2. Physically remove the row (safe replacement for ws.delete_rows)
        safe_delete_rows(ws, row, 1)

    return rows_to_delete


# ──────────────────────────────────────────────────────────────────────────────
# Main Processing
# ──────────────────────────────────────────────────────────────────────────────

def process_workbook(input_path, output_path, deletions):
    """
    Process an Excel workbook, deleting rows from specified roster regions.

    Parameters
    ----------
    input_path : str or Path
        Source Excel file.
    output_path : str or Path
        Destination Excel file.
    deletions : dict
        Mapping of {region_index: number_of_rows_to_delete}.
        Example: {1: 1, 2: 3}
    """
    wb = openpyxl.load_workbook(input_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        regions = detect_roster_regions(ws)

        print(f"\n📋 Sheet '{sheet_name}' — detected {len(regions)} roster region(s):")
        for r in regions:
            print(f"   Region {r['index']}: rows {r['data_start']}-{r['data_end']} "
                  f"({r['data_end'] - r['data_start'] + 1} rows) — {r['title']}")

        # Collect all rows to delete (using original row numbers)
        all_rows_to_delete = []
        deletion_log = {}
        for region_idx, n in deletions.items():
            region = next((r for r in regions if r['index'] == region_idx), None)
            if region is None:
                print(f"   ⚠️  Region {region_idx} not found, skipping.")
                continue

            if n <= 0:
                continue

            available = region['data_end'] - region['data_start'] + 1
            if n >= available:
                print(f"   ⚠️  Cannot delete {n} rows from region {region_idx} "
                      f"(only {available} rows available). Skipping.")
                continue

            data_start = region['data_start']
            rows = list(range(data_start + 1, data_start + n + 1))
            all_rows_to_delete.extend(rows)
            deletion_log[region_idx] = rows

        # Delete all rows in descending order so earlier deletions
        # don't shift the row numbers of later deletions.
        if all_rows_to_delete:
            for row in sorted(all_rows_to_delete, reverse=True):
                adjust_all_formulas(ws, row)
                safe_delete_rows(ws, row, 1)

            for region_idx, rows in deletion_log.items():
                print(f"   ✅ Deleted {len(rows)} row(s) from region {region_idx}: "
                      f"original rows {rows}")

    wb.save(output_path)
    print(f"\n💾 Saved to: {output_path}")
    return wb


def parse_deletions(arg):
    """Parse a deletion spec like '1:1,2:3' into a dict."""
    result = {}
    for part in arg.split(','):
        part = part.strip()
        if ':' in part:
            m, n = part.split(':', 1)
            result[int(m)] = int(n)
    return result


def main():
    if len(sys.argv) >= 4:
        input_path = sys.argv[1]
        output_path = sys.argv[2]
        deletions = parse_deletions(sys.argv[3])
    else:
        # Default test: 东汇保安 template, delete 1 row from region 1, 3 rows from region 2
        input_path = "templates/东汇-保安轮休表template.xlsx"
        output_path = "test_outputs_delete_rows/东汇-保安轮休表_deleted.xlsx"
        deletions = {1: 1, 2: 3}
        print("No arguments provided; running default test case.")
        print(f"  Input:  {input_path}")
        print(f"  Output: {output_path}")
        print(f"  Deletions: {deletions}")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    process_workbook(input_path, output_path, deletions)


if __name__ == '__main__':
    main()
