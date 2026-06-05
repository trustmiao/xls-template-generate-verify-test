"""Built-in engine: cleaning roster fill v2 for the new 31-day template.

The new template ships with 31 days pre-configured, including 5th-week
formulas using COUNT() for proportional required hours.  For months with
<31 days we delete excess day columns and let the template's formulas
adapt automatically.
"""
from __future__ import annotations

import calendar
import re
from copy import copy
from datetime import date
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook

from ..common.col_adjust import (
    update_all_formulas_for_col_change,
    col_num_to_letter,
)

DAY_START_COL = 4   # col D
RANK_COL = 2        # col B
NAME_COL = 3        # col C
TEMPLATE_DAYS = 31  # new template has 31 days pre-configured

_TITLE_RE = re.compile(r"^(\d+)\.\s")
_SUM_SINGLE_RE = re.compile(r"^=SUM\(([A-Z]+)(\d+):\1(\d+)\)$")
_CELL_REF_RE = re.compile(r"(\$?[A-Z]{1,3}\$?)(\d+)")
_RANGE_RE = re.compile(r"([A-Z$]+)(\d+):\1(\d+)")


# ---------------------------------------------------------------------------
# Merged-cell helpers
# ---------------------------------------------------------------------------

def _shift_merged_cols_for_delete(ws, delete_from: int, delete_count: int) -> List[Tuple[str, Optional[str], Any, Any, bool]]:
    """Adjust merged cell ranges after column deletion.

    Returns a list of (old_range, new_range, tl_value, tl_style, tl_has_style)
    tuples so the caller can restore top-left cell values AFTER delete_cols
    (otherwise openpyxl wipes the restored cell during column deletion).
    """
    from openpyxl.worksheet.cell_range import CellRange
    affected: List[Tuple[str, Optional[str], Any, Any, bool]] = []
    delete_end = delete_from + delete_count - 1

    for mr in list(ws.merged_cells.ranges):
        if mr.max_col < delete_from:
            continue
        elif mr.min_col > delete_end:
            new = CellRange(
                min_col=mr.min_col - delete_count, min_row=mr.min_row,
                max_col=mr.max_col - delete_count, max_row=mr.max_row,
            )
            # Snapshot top-left value + style before unmerge wipes it
            tl = ws.cell(mr.min_row, mr.min_col)
            affected.append((str(mr), str(new), tl.value, tl._style, tl.has_style))
        elif mr.min_col >= delete_from and mr.max_col <= delete_end:
            affected.append((str(mr), None, None, None, False))
        else:
            if mr.min_col < delete_from:
                new_max = mr.max_col - delete_count
                if new_max < mr.min_col:
                    new_max = mr.min_col
                new = CellRange(
                    min_col=mr.min_col, min_row=mr.min_row,
                    max_col=new_max, max_row=mr.max_row,
                )
            else:
                new_min = delete_from
                new_max = mr.max_col - delete_count
                if new_max < new_min:
                    new_max = new_min
                new = CellRange(
                    min_col=new_min, min_row=mr.min_row,
                    max_col=new_max, max_row=mr.max_row,
                )
            # For partially-overlapping ranges the top-left stays the same,
            # so no need to copy value/style.
            affected.append((str(mr), str(new), None, None, False))

    # Phase 1: unmerge all old ranges first.
    for old, new, tl_value, tl_style, tl_has_style in affected:
        ws.unmerge_cells(old)

    # Phase 2: merge new ranges.
    for old, new, tl_value, tl_style, tl_has_style in affected:
        if new:
            ws.merge_cells(new)

    return affected


# ---------------------------------------------------------------------------
# Day-column adjustment (delete excess from 31-day template)
# ---------------------------------------------------------------------------

def _adjust_day_columns(ws, days_in_month: int) -> int:
    """Adjust day columns from template's 31 days to target month days.

    For months < 31 days, deletes excess columns from the right edge
    (AF onwards) and adjusts merged cells + formulas.
    """
    if days_in_month >= TEMPLATE_DAYS:
        return DAY_START_COL + TEMPLATE_DAYS - 1

    diff = TEMPLATE_DAYS - days_in_month
    delete_from = DAY_START_COL + days_in_month

    # Clear all cell values in the columns to be deleted
    from openpyxl.cell.cell import MergedCell
    for row in range(1, ws.max_row + 1):
        for col in range(delete_from, delete_from + diff):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # Adjust merged cells before deleting (returns info needed for post-delete restore)
    affected = _shift_merged_cols_for_delete(ws, delete_from, diff)

    # Update formulas to shrink ranges ending in deleted columns
    update_all_formulas_for_col_change(ws, delete_from, -diff, is_delete=True)

    # Delete the excess columns
    ws.delete_cols(delete_from, diff)

    # Restore top-left values for merged ranges that were shifted left.
    # Must happen AFTER delete_cols because openpyxl removes the restored
    # cell when the deleted column is removed.
    from openpyxl.worksheet.cell_range import CellRange
    for old, new, tl_value, tl_style, tl_has_style in affected:
        if new and (tl_has_style or tl_value is not None):
            new_cr = CellRange(new)
            key = (new_cr.min_row, new_cr.min_col)
            if key in ws._cells and type(ws._cells[key]).__name__ == "MergedCell":
                del ws._cells[key]
            new_tl = ws.cell(new_cr.min_row, new_cr.min_col)
            if tl_value is not None:
                # Formulas captured before delete_cols still reference the old
                # column letters; shift them to match the post-delete layout.
                if isinstance(tl_value, str) and tl_value.startswith("="):
                    from ..common.col_adjust import shift_formula_cols
                    tl_value = shift_formula_cols(tl_value, delete_from, -diff, is_delete=True)
                new_tl.value = tl_value
            if tl_has_style and tl_style is not None:
                new_tl._style = tl_style

    return DAY_START_COL + days_in_month - 1


# ---------------------------------------------------------------------------
# Date update
# ---------------------------------------------------------------------------

def _update_dates(ws, month: str) -> None:
    """Update date row to target month (first cell = date, rest = +1 chain)."""
    date_row = None
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(DAY_START_COL, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and hasattr(v, "year"):
                date_row = r
                break
        if date_row:
            break

    if not date_row:
        return

    year, month_num = int(month[:4]), int(month[5:7])
    days_in_month = calendar.monthrange(year, month_num)[1]

    ws.cell(date_row, DAY_START_COL, value=date(year, month_num, 1))

    weekday_row = date_row + 2
    if weekday_row <= ws.max_row:
        for c in range(DAY_START_COL, DAY_START_COL + days_in_month):
            v = ws.cell(weekday_row, c).value
            if isinstance(v, str) and "TEXT" in v:
                col_letter = col_num_to_letter(c)
                ws.cell(weekday_row, c,
                        value=f'=SUBSTITUTE(TEXT({col_letter}{date_row},"aaa"),"週","")')


# ---------------------------------------------------------------------------
# Row insertion helpers (reused from v1)
# ---------------------------------------------------------------------------

def _find_segment_title_rows(ws) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, RANK_COL).value
        if v and isinstance(v, str):
            m = _TITLE_RE.match(v)
            if m:
                out[m.group(1) + "."] = r
    return out


def _find_data_row(ws, start_row: int, end_row: int) -> int | None:
    for r in range(start_row + 1, end_row + 1):
        v = ws.cell(r, DAY_START_COL).value
        if v and isinstance(v, str):
            m = _SUM_SINGLE_RE.match(v)
            if m and m.group(2) == m.group(3):
                return int(m.group(2))
    return None


def _insert_and_style(ws, insert_at: int, amount: int, source_row: int) -> None:
    from openpyxl.worksheet.cell_range import CellRange
    affected: List[Tuple[str, str]] = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= insert_at:
            new = CellRange(min_col=mr.min_col, min_row=mr.min_row + amount,
                            max_col=mr.max_col, max_row=mr.max_row + amount)
        elif mr.max_row >= insert_at:
            new = CellRange(min_col=mr.min_col, min_row=mr.min_row,
                            max_col=mr.max_col, max_row=mr.max_row + amount)
        else:
            continue
        affected.append((str(mr), str(new)))
    for old, _ in affected:
        ws.unmerge_cells(old)
    ws.insert_rows(insert_at, amount=amount)
    for _, new in affected:
        ws.merge_cells(new)
    _shift_print_area(ws, insert_at, amount)
    for off in range(amount):
        target_row = insert_at + off
        for col in range(1, ws.max_column + 1):
            src = ws.cell(source_row, col)
            dst = ws.cell(target_row, col)
            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = copy(src.number_format)
                dst.protection = copy(src.protection)
                dst.alignment = copy(src.alignment)


def _shift_print_area(ws, insert_at: int, amount: int) -> None:
    _PRINT_AREA_RE = re.compile(r"\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)")
    pa = ws.print_area
    if pa:
        def repl(m: re.Match) -> str:
            c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
            if r1 >= insert_at: r1 += amount
            if r2 >= insert_at: r2 += amount
            return f"${c1}${r1}:${c2}${r2}"
        new_pa = _PRINT_AREA_RE.sub(repl, str(pa))
        if new_pa != pa:
            ws.print_area = new_pa
    titles = ws.print_title_rows
    if titles:
        m = re.match(r"\$(\d+):\$(\d+)", titles)
        if m:
            r1, r2 = int(m.group(1)), int(m.group(2))
            if r1 >= insert_at: r1 += amount
            if r2 >= insert_at: r2 += amount
            new_titles = f"${r1}:${r2}"
            if new_titles != titles:
                ws.print_title_rows = new_titles


def _shift_formulas(ws, insert_at: int, amount: int) -> None:
    def repl(m: re.Match) -> str:
        col, r = m.group(1), int(m.group(2))
        if r >= insert_at:
            return f"{col}{r + amount}"
        return m.group(0)
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row, col).value
            if isinstance(v, str) and v.startswith("="):
                new = _CELL_REF_RE.sub(repl, v)
                if new != v:
                    ws.cell(row, col).value = new


def _expand_range(ws, segment_data_row: int, segment_end_after: int) -> None:
    def repl(m: re.Match) -> str:
        col, r1, r2 = m.group(1), int(m.group(2)), int(m.group(3))
        if r1 == r2 == segment_data_row:
            return f"{col}{r1}:{col}{segment_end_after}"
        return m.group(0)
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row, col).value
            if isinstance(v, str) and v.startswith("="):
                new = _RANGE_RE.sub(repl, v)
                if new != v:
                    ws.cell(row, col).value = new


def _copy_row_formulas(ws, source_row: int, insert_at: int, amount: int) -> None:
    for off in range(amount):
        target_row = insert_at + off
        for col in range(1, ws.max_column + 1):
            sv = ws.cell(source_row, col).value
            if isinstance(sv, str) and sv.startswith("="):
                def _rewrite(m: re.Match) -> str:
                    col_part, ref_row = m.group(1), int(m.group(2))
                    if ref_row == source_row:
                        return f"{col_part}{target_row}"
                    return m.group(0)
                ws.cell(target_row, col).value = _CELL_REF_RE.sub(_rewrite, sv)


def _fill_data_row(ws, r: int, row_data: Dict[str, Any], days_in_month: int) -> None:
    ws.cell(r, RANK_COL, value=row_data.get("rank_seq") or row_data.get("employee_no") or "")
    ws.cell(r, NAME_COL, value=row_data.get("name") or "")
    cells = {c["day"]: c for c in row_data.get("cells", [])}
    for d in range(1, days_in_month + 1):
        col = DAY_START_COL + d - 1
        c = cells.get(d)
        if not c:
            ws.cell(r, col, value="")
            continue
        v = c.get("value")
        if v is None:
            v = c.get("code")
        try:
            num = float(v)
            ws.cell(r, col, value=int(num) if num == int(num) else num)
        except (TypeError, ValueError):
            ws.cell(r, col, value=str(v) if v else "")


# ---------------------------------------------------------------------------
# Main entry
# ---------------------------------------------------------------------------

def run(wb: Workbook, context: Dict[str, Any]) -> None:
    """Fill the cleaning roster template (v2) for the 31-day pre-configured template."""
    project_id = context["project_id"]
    category_id = context["category_id"]
    month = context["month"]

    from ...api.shortfall_engine import api_shortfall_engine
    data = api_shortfall_engine(
        shift="A", project_id=project_id, category_id=category_id, month=month
    )
    if not data.get("has_data"):
        return

    segments = data.get("segments", [])
    days_in_month = data.get("days_in_month", 31)

    ws = wb.worksheets[0]

    # Phase 0: Adjust day columns (delete excess for months < 31 days)
    day_end_col = _adjust_day_columns(ws, days_in_month)

    # Phase 1: Update dates to target month
    _update_dates(ws, month)

    # Discover segment positions
    title_rows = _find_segment_title_rows(ws)
    if not title_rows:
        return
    sorted_titles = sorted(title_rows.items(), key=lambda x: x[1])
    section_bounds: Dict[str, Dict[str, int]] = {}
    for i, (prefix, tr) in enumerate(sorted_titles):
        end = sorted_titles[i + 1][1] - 1 if i + 1 < len(sorted_titles) else ws.max_row
        dr = _find_data_row(ws, tr, end)
        if dr is not None:
            section_bounds[prefix] = {"title_row": tr, "data_row": dr, "end_row": end}

    # Match API segments
    api_by_prefix: Dict[str, Dict[str, Any]] = {}
    for s in segments:
        m = _TITLE_RE.match(s.get("title", ""))
        if m:
            api_by_prefix[m.group(1) + "."] = s

    # Bottom-up: insert + fill per section
    ordered = sorted(section_bounds.keys(), key=lambda p: section_bounds[p]["data_row"], reverse=True)
    for prefix in ordered:
        sb = section_bounds[prefix]
        seg = api_by_prefix.get(prefix)
        if not seg or not seg.get("rows"):
            continue
        rows = seg["rows"]
        n = len(rows)
        template_data_row = sb["data_row"]
        end_row = template_data_row + n - 1
        if n > 1:
            insert_at = template_data_row + 1
            _insert_and_style(ws, insert_at, n - 1, source_row=template_data_row)
            _shift_formulas(ws, insert_at, n - 1)
            _expand_range(ws, template_data_row, end_row)
            _copy_row_formulas(ws, template_data_row, insert_at, n - 1)
        for idx, row_data in enumerate(rows):
            _fill_data_row(ws, template_data_row + idx, row_data, days_in_month)
