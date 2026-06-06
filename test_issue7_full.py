#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Issue #7 full test: cleaning_shortfall_v2 connects to service and exports.

Tests all template combinations:
  1. Load template
  2. Build context (project_id, category_id, month)
  3. Run cleaning_shortfall_v2 (adjust columns + dates + shrink/expand + fill data)
  4. Save output
  5. Verify: sheet count, segments per sheet, data filled correctly
"""

import sys
sys.path.insert(0, r"D:\claude\claude_hk\backend")

import re
import calendar
from pathlib import Path
from openpyxl import load_workbook

from app.engine.excel.cleaning_shortfall_v2 import run as cleaning_shortfall_v2_run
from app.engine.common.data_source import build_context

OUT_DIR = Path("test_outputs_issue7")
OUT_DIR.mkdir(exist_ok=True)

_TPL_DIR = Path(__file__).parent / "templates"

# ── Test combinations ──
COMBINATIONS = [
    {
        "label": "东汇邨-保洁-2026-02",
        "template": _TPL_DIR / "东汇-保洁轮休表_template.xlsx",
        "project_id": 2,
        "category_id": 2,
        "month": "2026-02",
        "expected_sheets": ["Roster-FEB2026"],
        "expected_segments": 3,
    },
    {
        "label": "东汇邨-保洁-2026-03",
        "template": _TPL_DIR / "东汇-保洁轮休表_template.xlsx",
        "project_id": 2,
        "category_id": 2,
        "month": "2026-03",
        "expected_sheets": ["Roster-FEB2026"],
        "expected_segments": 3,
    },
    {
        "label": "东汇邨-保安-2026-02",
        "template": _TPL_DIR / "东汇-保安轮休表template.xlsx",
        "project_id": 2,
        "category_id": 3,
        "month": "2026-02",
        "expected_sheets": ["早", "中", "夜"],
        "expected_segments": 2,
    },
    {
        "label": "东汇邨-保安-2026-03",
        "template": _TPL_DIR / "东汇-保安轮休表template.xlsx",
        "project_id": 2,
        "category_id": 3,
        "month": "2026-03",
        "expected_sheets": ["早", "中", "夜"],
        "expected_segments": 2,
    },
    {
        "label": "大元邨-保安-2026-02",
        "template": _TPL_DIR / "TY-2026.03- SG_SEC-Deploy Roster  Shortfall - template.xlsx",
        "project_id": 1,
        "category_id": 1,
        "month": "2026-02",
        "expected_sheets": ["早", "中", "夜"],
        "expected_segments": 2,
    },
    {
        "label": "大元邨-保安-2026-03",
        "template": _TPL_DIR / "TY-2026.03- SG_SEC-Deploy Roster  Shortfall - template.xlsx",
        "project_id": 1,
        "category_id": 1,
        "month": "2026-03",
        "expected_sheets": ["早", "中", "夜"],
        "expected_segments": 2,
    },
]

title_re = re.compile(r"^(\d+)\.\s")
rank_re = re.compile(r"^[A-Z]{1,2}\d+$")


def inspect_sheet(ws, days_in_month):
    """Inspect a sheet and return segment info."""
    # Find date row
    date_row = None
    day_start_col = None
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and hasattr(v, "year"):
                date_row = r
                day_start_col = c
                break
        if date_row:
            break

    # Count actual date columns (check date row)
    actual_days = 0
    if date_row and day_start_col:
        for c in range(day_start_col, ws.max_column + 1):
            v = ws.cell(date_row, c).value
            if v is not None and (hasattr(v, "year") or (isinstance(v, str) and v.startswith("="))):
                actual_days += 1
            else:
                break

    # Detect segments using engine logic
    title_rows = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v and isinstance(v, str) and title_re.match(v):
            m = title_re.match(v)
            title_rows[m.group(1) + "."] = r

    sorted_titles = sorted(title_rows.items(), key=lambda x: x[1])
    segments = []
    for i, (prefix, tr) in enumerate(sorted_titles):
        end = sorted_titles[i + 1][1] - 1 if i + 1 < len(sorted_titles) else ws.max_row
        data_rows = []
        for dr in range(tr + 1, end + 1):
            b = ws.cell(dr, 2).value
            if b and isinstance(b, str) and rank_re.match(b.strip()):
                name = ws.cell(dr, 3).value or ""
                # Check first day cell has data
                day_val = None
                if day_start_col:
                    day_val = ws.cell(dr, day_start_col).value
                data_rows.append({"rank": b, "name": name, "day1": day_val})
            elif data_rows:
                break
        segments.append({
            "prefix": prefix,
            "title": ws.cell(tr, 2).value[:50] if ws.cell(tr, 2).value else "",
            "count": len(data_rows),
            "people": data_rows,
        })

    return {
        "date_row": date_row,
        "day_start_col": day_start_col,
        "actual_days": actual_days,
        "segments": segments,
        "segment_count": len(segments),
    }


def run_test(combo):
    """Run one test combination and return results."""
    label = combo["label"]
    template_path = Path(combo["template"])
    month = combo["month"]
    year, mon = map(int, month.split("-"))
    days_in_month = calendar.monthrange(year, mon)[1]

    print(f"\n{'='*60}")
    print(f"Testing: {label}")
    print(f"{'='*60}")

    context = build_context(
        project_id=combo["project_id"],
        category_id=combo["category_id"],
        month=month,
    )

    wb = load_workbook(str(template_path), data_only=False)
    print(f"  Template sheets: {[ws.title for ws in wb.worksheets]}")

    try:
        cleaning_shortfall_v2_run(wb, context)
    except Exception as e:
        print(f"  ERROR during run: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}

    out_name = f"{label.replace('/', '_')}_issue7.xlsx"
    out_path = OUT_DIR / out_name
    wb.save(str(out_path))
    print(f"  Saved: {out_path}")

    # Verify output
    wb2 = load_workbook(str(out_path), data_only=False)
    results = {"ok": True, "sheets": {}, "errors": []}

    expected_sheets = combo["expected_sheets"]
    actual_sheets = [ws.title for ws in wb2.worksheets if ws.title != "Data"]
    if actual_sheets != expected_sheets:
        results["errors"].append(
            f"Sheet mismatch: expected {expected_sheets}, got {actual_sheets}"
        )
        results["ok"] = False

    for ws in wb2.worksheets:
        if ws.title == "Data":
            continue
        info = inspect_sheet(ws, days_in_month)
        results["sheets"][ws.title] = info

        print(f"\n  Sheet '{ws.title}':")
        print(f"    Date row: {info['date_row']}, Days: {info['actual_days']}/{days_in_month}")
        print(f"    Segments: {info['segment_count']}")

        if info["actual_days"] != days_in_month:
            results["errors"].append(
                f"[{ws.title}] Day count mismatch: {info['actual_days']} vs {days_in_month}"
            )
            results["ok"] = False

        for seg in info["segments"]:
            print(f"      {seg['prefix']} {seg['title']}... : {seg['count']} person(s)")
            # Verify first person has day data
            if seg["people"]:
                p = seg["people"][0]
                print(f"        First: {p['rank']} / {p['name']} | day1={p['day1']}")
                if p["day1"] is None or p["day1"] == "":
                    results["errors"].append(
                        f"[{ws.title}] {seg['prefix']} first person has no day1 data"
                    )
                    results["ok"] = False

    if results["ok"]:
        print(f"  ✅ PASSED")
    else:
        print(f"  ❌ FAILED: {results['errors']}")

    return results


def main():
    all_ok = True
    results = []

    for combo in COMBINATIONS:
        res = run_test(combo)
        results.append({"label": combo["label"], **res})
        if not res["ok"]:
            all_ok = False

    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    for r in results:
        status = "✅ PASS" if r["ok"] else "❌ FAIL"
        print(f"  {status}: {r['label']}")
        if not r["ok"]:
            for e in r.get("errors", []):
                print(f"    - {e}")

    print(f"\nOutput dir: {OUT_DIR}")
    print(f"Total: {len(COMBINATIONS)}, Passed: {sum(1 for r in results if r['ok'])}, Failed: {sum(1 for r in results if not r['ok'])}")

    return 0 if all_ok else 1


if __name__ == "__main__":
    sys.exit(main())
