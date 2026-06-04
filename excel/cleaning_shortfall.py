"""Built-in engine: 保洁 single-sheet Roster Shortfall fill.

Designed for 东汇邨 保洁 (cleaning) template which has a single sheet with
3 sections (清潔科文 / 清潔工人 / VO清潔工人). Each section's data goes
between its title row and its summary rows. The template ships with 1 blank
data row per section; this engine inserts extra rows to fit N people and
rewrites SUM formulas to cover the new range.

Source of truth for row data is /api/shortfall-engine, so anything that
endpoint already knows about (rank classification, hour codes, VO handling)
applies here without duplication.
"""
from __future__ import annotations

import re
from copy import copy
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook

from ..common.col_adjust import (
    shift_formula_cols,
    delete_cols_with_formulas,
    insert_cols_with_formulas,
    col_letter_to_num,
    col_num_to_letter,
)

DAY_START_COL = 4  # col D
RANK_COL = 2       # col B
NAME_COL = 3       # col C

_TITLE_RE = re.compile(r"^(\d+)\.\s")
_SUM_SINGLE_RE = re.compile(r"^=SUM\(([A-Z]+)(\d+):\1(\d+)\)$")
_CELL_REF_RE = re.compile(r"(\$?[A-Z]{1,3}\$?)(\d+)")
_RANGE_RE = re.compile(r"([A-Z$]+)(\d+):\1(\d+)")


def _find_segment_title_rows(ws) -> Dict[str, int]:
    """Find rows whose col B starts with 'N. '. Returns {'N.': row}."""
    out: Dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, RANK_COL).value
        if v and isinstance(v, str):
            m = _TITLE_RE.match(v)
            if m:
                out[m.group(1) + "."] = r
    return out


def _find_data_row(ws, start_row: int, end_row: int) -> int | None:
    """First row in (start_row..end_row] referenced by =SUM(Dn:Dn) at col D below."""
    for r in range(start_row + 1, end_row + 1):
        v = ws.cell(r, DAY_START_COL).value
        if v and isinstance(v, str):
            m = _SUM_SINGLE_RE.match(v)
            if m and m.group(2) == m.group(3):
                return int(m.group(2))
    return None


# ---------------------------------------------------------------------------
# Day-column detection / adjustment for cleaning templates
# ---------------------------------------------------------------------------

def _detect_cleaning_date_row(ws) -> int | None:
    """Find the row containing the first date value (e.g. 2026-02-01)."""
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(DAY_START_COL, min(ws.max_column + 1, DAY_START_COL + 35)):
            v = ws.cell(r, c).value
            if v is not None and hasattr(v, "year"):
                return r
    return None


def _detect_cleaning_weekday_row(ws, date_row: int) -> int | None:
    """Find the weekday row right below the date row."""
    for r in range(date_row + 1, min(ws.max_row, date_row + 5) + 1):
        v = ws.cell(r, DAY_START_COL).value
        if v is not None:
            return r
    return None


def _detect_day_end_col(ws, date_row: int) -> int:
    """Detect the last day column by walking the date row."""
    last_day = None
    for c in range(DAY_START_COL, ws.max_column + 1):
        v = ws.cell(date_row, c).value
        if v is not None and hasattr(v, "year"):
            last_day = c
        elif isinstance(v, str) and v.startswith("=") and "+1" in v:
            last_day = c
        elif v is not None and isinstance(v, str) and v.strip() and not v.startswith("="):
            # Hit a label like "1-7號時數" — this is past the day columns
            break
    return last_day or DAY_START_COL + 27


def _adjust_cleaning_day_columns(ws, days_in_month: int) -> tuple[int, int, int]:
    """Adjust day columns in cleaning template to match month days.

    Returns (day_end_col, date_row, weekday_row).
    For months > 28 days, extra columns are inserted after day 28.
    Date-row and weekday-row formulas are extended so the chain continues.
    Summary formulas are updated to include the extra days while keeping
    the 4th week (days 22-28) range intact.
    """
    date_row = _detect_cleaning_date_row(ws)
    weekday_row = _detect_cleaning_weekday_row(ws, date_row) if date_row else None
    day_end_col = _detect_day_end_col(ws, date_row) if date_row else DAY_START_COL + 27
    tpl_days = day_end_col - DAY_START_COL + 1
    diff = days_in_month - tpl_days

    if diff == 0:
        return day_end_col, date_row or 7, weekday_row or 9

    if diff < 0:
        # Delete excess columns
        delete_from = DAY_START_COL + days_in_month
        delete_cols_with_formulas(ws, delete_from, -diff)
        day_end_col = DAY_START_COL + days_in_month - 1
        return day_end_col, date_row or 7, weekday_row or 9

    # Insert extra columns after the last day column
    insert_at = day_end_col + 1
    insert_cols_with_formulas(ws, insert_at, diff, date_row=date_row, weekday_row=weekday_row)
    _shift_merged_cols(ws, insert_at, diff)
    new_day_end = day_end_col + diff

    # Fix date-row formulas for inserted columns so the date chain continues.
    # insert_cols_with_formulas copies the formula from the column before the
    # insertion point, but that produces an incorrect reference (e.g. copying
    # AE7=AD7+1 to AF7 still yields AD7+1 instead of AE7+1).  Overwrite with
    # the correct chained formula.
    if date_row:
        for i in range(diff):
            col = insert_at + i
            prev_letter = col_num_to_letter(col - 1)
            ws.cell(date_row, col, value=f"={prev_letter}{date_row}+1")

    # Fix weekday-row formulas similarly.
    if weekday_row:
        for i in range(diff):
            col = insert_at + i
            col_letter = col_num_to_letter(col)
            ws.cell(weekday_row, col,
                    value=f'=SUBSTITUTE(TEXT({col_letter}{date_row},"aaa"),"週","")')

    # Update summary formulas that reference the old day-end (AE) to new end,
    # but keep the 4th-week (days 22-28) range unchanged.
    _update_cleaning_summary_formulas(ws, day_end_col, new_day_end)

    return new_day_end, date_row or 7, weekday_row or 9


def _update_cleaning_summary_formulas(ws, old_end_col: int, new_end_col: int) -> None:
    """Update weekly / total summary formulas to include extra day columns.

    Rules:
      - 4th-week ranges (start_col = Y, days 22-28) are kept unchanged.
      - Single-col daily summaries (start_col == old_end_col) are untouched.
      - All other multi-column ranges ending at old_end_col are expanded to
        new_end_col (e.g. monthly totals D:AE -> D:AH).
    """
    old_end_letter = col_num_to_letter(old_end_col)
    new_end_letter = col_num_to_letter(new_end_col)
    # For a template whose day columns start at D (col 4), the 4th week
    # starts at col 4 + 7*3 = 25 -> letter Y.
    fourth_week_start = col_num_to_letter(DAY_START_COL + 7 * 3)

    def _update_formula(v: str) -> str:
        if not isinstance(v, str) or not v.startswith("="):
            return v
        # Pattern: start_col + start_row : old_end_col + end_row
        pat = re.compile(rf'([A-Z]+)(\d+):{old_end_letter}(\d+)')

        def repl(m: re.Match) -> str:
            start_col = m.group(1)
            start_row = m.group(2)
            end_row = m.group(3)
            if start_col == old_end_letter:
                # Single-col range — leave untouched
                return m.group(0)
            if start_col == fourth_week_start:
                # 4th week covers days 22-28 — keep the original range
                return m.group(0)
            return f"{start_col}{start_row}:{new_end_letter}{end_row}"

        return pat.sub(repl, v)

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                new_v = _update_formula(cell.value)
                if new_v != cell.value:
                    cell.value = new_v


def _shift_merged_ranges_around_insert(ws, insert_at: int, amount: int) -> None:
    """Move merged-range metadata across a pending row insertion.

    Order matters:
      1. Compute target ranges (read mr.min_row / mr.max_row from current
         positions — these are pre-insert).
      2. unmerge_cells BEFORE ws.insert_rows: openpyxl's unmerge walks sub-
         cells inside the range and del self._cells[r,c] them; those sub-
         cells only exist at the original positions, so calling unmerge
         after insert_rows would KeyError.
      3. ws.insert_rows shifts cell contents.
      4. merge_cells at the new (shifted) positions — top-left of the new
         merge now holds the content openpyxl moved there.
    """
    from openpyxl.worksheet.cell_range import CellRange
    affected: List[Tuple[str, str]] = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= insert_at:
            new = CellRange(min_col=mr.min_col, min_row=mr.min_row + amount,
                            max_col=mr.max_col, max_row=mr.max_row + amount)
        elif mr.max_row >= insert_at:
            new = CellRange(min_col=mr.min_col, min_row=mr.min_row,
                            max_col=mr.max_col, max_row=mr.max_row + amount)
        else:
            continue
        affected.append((str(mr), str(new)))
    return affected


def _shift_merged_cols(ws, insert_at: int, amount: int) -> None:
    """Shift merged cell ranges that are at or after a column insertion point."""
    from openpyxl.worksheet.cell_range import CellRange
    affected: List[Tuple[str, str]] = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_col >= insert_at:
            new = CellRange(
                min_col=mr.min_col + amount, min_row=mr.min_row,
                max_col=mr.max_col + amount, max_row=mr.max_row,
            )
            affected.append((str(mr), str(new)))
        elif mr.max_col >= insert_at:
            new = CellRange(
                min_col=mr.min_col, min_row=mr.min_row,
                max_col=mr.max_col + amount, max_row=mr.max_row,
            )
            affected.append((str(mr), str(new)))
    for old, _ in affected:
        ws.unmerge_cells(old)
    for _, new in affected:
        ws.merge_cells(new)


def _shift_merged_ranges(ws, insert_at: int, amount: int) -> None:
    """Legacy wrapper: used by callers that have already inserted rows."""
    # When the caller has already done ws.insert_rows, the merge metadata is
    # still at old positions but the sub-cells have moved. This path is only
    # safe when called BEFORE insert_rows.
    from openpyxl.worksheet.cell_range import CellRange
    affected: List[Tuple[str, str]] = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= insert_at:
            new = CellRange(min_col=mr.min_col, min_row=mr.min_row + amount,
                            max_col=mr.max_col, max_row=mr.max_row + amount)
        elif mr.max_row >= insert_at:
            new = CellRange(min_col=mr.min_col, min_row=mr.min_row,
                            max_col=mr.max_col, max_row=mr.max_row + amount)
        else:
            continue
        affected.append((str(mr), str(new)))
    for old, _ in affected:
        ws.unmerge_cells(old)
    for _, new in affected:
        ws.merge_cells(new)


_PRINT_AREA_RE = re.compile(r"\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)")


def _shift_print_area(ws, insert_at: int, amount: int) -> None:
    """Shift the bottom row of print_area and any row reference in print_title_rows."""
    pa = ws.print_area
    if pa:
        def repl(m: re.Match) -> str:
            c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
            if r1 >= insert_at:
                r1 += amount
            if r2 >= insert_at:
                r2 += amount
            return f"${c1}${r1}:${c2}${r2}"
        new_pa = _PRINT_AREA_RE.sub(repl, str(pa))
        if new_pa != pa:
            ws.print_area = new_pa

    titles = ws.print_title_rows
    if titles:
        # Format: "$1:$9" — single row range
        m = re.match(r"\$(\d+):\$(\d+)", titles)
        if m:
            r1, r2 = int(m.group(1)), int(m.group(2))
            if r1 >= insert_at: r1 += amount
            if r2 >= insert_at: r2 += amount
            new_titles = f"${r1}:${r2}"
            if new_titles != titles:
                ws.print_title_rows = new_titles


def _insert_and_style(ws, insert_at: int, amount: int, source_row: int) -> None:
    # Three-step merge-aware insertion:
    #   1. Compute target merged ranges + UNMERGE old ones (sub-cells must
    #      still exist for openpyxl's unmerge — i.e. BEFORE insert_rows).
    #   2. insert_rows shifts cell contents.
    #   3. Re-merge at new positions; the value that openpyxl shifted INTO
    #      the new top-left now becomes the merged cell's visible value.
    # Previously this was pre-shift-then-insert, which left the new merge's
    # top-left blank because the value landed one cell over (the legend row
    # in 东汇邨保洁 lost its 備註/總時數 cells that way).
    affected = _shift_merged_ranges_around_insert(ws, insert_at, amount)
    for old, _ in affected:
        ws.unmerge_cells(old)
    ws.insert_rows(insert_at, amount=amount)
    for _, new in affected:
        ws.merge_cells(new)
    _shift_print_area(ws, insert_at, amount)
    for off in range(amount):
        target_row = insert_at + off
        for col in range(1, ws.max_column + 1):
            src = ws.cell(source_row, col)
            dst = ws.cell(target_row, col)
            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = copy(src.number_format)
                dst.protection = copy(src.protection)
                dst.alignment = copy(src.alignment)


def _copy_row_formulas(ws, source_row: int, insert_at: int, amount: int) -> None:
    """Copy formulas from source_row to inserted rows, rewriting source_row → target.

    Run this AFTER _shift_formulas — _shift_formulas only sees the existing
    formulas (which it shifts), so newly-written formulas here are safe.

    Per-day cells will be overwritten by _fill_data_row with literal values,
    but weekly/monthly aggregation columns (e.g. AF: =SUM(D{r}:J{r}),
    AJ: =SUM(AF{r}:AI{r})) only ever come from the template's data row, so
    every inserted row needs its own copy referencing its own row number.
    """
    for off in range(amount):
        target_row = insert_at + off
        for col in range(1, ws.max_column + 1):
            sv = ws.cell(source_row, col).value
            if isinstance(sv, str) and sv.startswith("="):
                def _rewrite(m: re.Match) -> str:
                    col_part, ref_row = m.group(1), int(m.group(2))
                    if ref_row == source_row:
                        return f"{col_part}{target_row}"
                    return m.group(0)
                ws.cell(target_row, col).value = _CELL_REF_RE.sub(_rewrite, sv)


def _shift_formulas(ws, insert_at: int, amount: int) -> None:
    """Shift every cell-ref row number by +amount if it's >= insert_at."""
    def repl(m: re.Match) -> str:
        col, r = m.group(1), int(m.group(2))
        if r >= insert_at:
            return f"{col}{r + amount}"
        return m.group(0)
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row, col).value
            if isinstance(v, str) and v.startswith("="):
                new = _CELL_REF_RE.sub(repl, v)
                if new != v:
                    ws.cell(row, col).value = new


def _expand_range(ws, segment_data_row: int, segment_end_after: int) -> None:
    """Where SUM(Cn:Cn) with n=segment_data_row exists, expand to SUM(Cn:C{end})."""
    def repl(m: re.Match) -> str:
        col, r1, r2 = m.group(1), int(m.group(2)), int(m.group(3))
        if r1 == r2 == segment_data_row:
            return f"{col}{r1}:{col}{segment_end_after}"
        return m.group(0)
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row, col).value
            if isinstance(v, str) and v.startswith("="):
                new = _RANGE_RE.sub(repl, v)
                if new != v:
                    ws.cell(row, col).value = new


def _fill_data_row(ws, r: int, row_data: Dict[str, Any], days_in_month: int) -> None:
    ws.cell(r, RANK_COL, value=row_data.get("rank_seq") or row_data.get("employee_no") or "")
    ws.cell(r, NAME_COL, value=row_data.get("name") or "")
    cells = {c["day"]: c for c in row_data.get("cells", [])}
    for d in range(1, days_in_month + 1):
        col = DAY_START_COL + d - 1
        c = cells.get(d)
        if not c:
            ws.cell(r, col, value="")
            continue
        v = c.get("value")
        if v is None:
            v = c.get("code")
        try:
            num = float(v)
            ws.cell(r, col, value=int(num) if num == int(num) else num)
        except (TypeError, ValueError):
            ws.cell(r, col, value=str(v) if v else "")


def _add_fifth_week_formulas(ws, day_end_col: int, days_in_month: int) -> None:
    """Add 5th-week formulas and merge cells for months with >28 days.

    For each section (清潔科文 / 清潔工人 / VO清潔工人) adds:
      - 5th-week actual hours: SUM(daily-total row for days 29+)
      - 5th-week required hours: first-week value / 7 * extra_days
      - 5th-week diff: actual - required
    All placed in col AF (32) merged across the extra day columns.
    """
    if days_in_month <= 28:
        return

    extra_days = days_in_month - 28
    first_extra_col = DAY_START_COL + 28          # col 32 = AF
    first_extra_letter = col_num_to_letter(first_extra_col)
    new_end_letter = col_num_to_letter(day_end_col)
    first_col_letter = col_num_to_letter(DAY_START_COL)

    # Map labels to rows
    stat_rows: Dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 2).value
        if label and isinstance(label, str):
            stat_rows[label.strip()] = r

    sections = [
        ("清潔科文(實際) 每週工作總時數", "清潔科文每週所需工作總時數", "欠/多 清潔科文時數"),
        ("清潔工人(實際)每週工作總時數", "清潔工人每週所需工作總時數", "欠/多 清潔工人時數"),
        ("VO清潔工人(實際) 每週工作總時數", "VO清潔工人每週所需工作總時數", "欠/多 VO清潔工人時數"),
    ]

    for actual_label, required_label, diff_label in sections:
        actual_row = stat_rows.get(actual_label)
        required_row = stat_rows.get(required_label)
        diff_row = stat_rows.get(diff_label)

        if not all([actual_row, required_row, diff_row]):
            continue

        daily_row = actual_row - 1

        # 5th week actual hours
        ws.cell(actual_row, first_extra_col,
                value=f"=SUM({first_extra_letter}{daily_row}:{new_end_letter}{daily_row})")
        ws.merge_cells(start_row=actual_row, start_column=first_extra_col,
                       end_row=actual_row, end_column=day_end_col)

        # 5th week required hours (proportional)
        ws.cell(required_row, first_extra_col,
                value=f"={first_col_letter}{required_row}/7*{extra_days}")
        ws.merge_cells(start_row=required_row, start_column=first_extra_col,
                       end_row=required_row, end_column=day_end_col)

        # 5th week diff
        ws.cell(diff_row, first_extra_col,
                value=f"={first_extra_letter}{actual_row}-{first_extra_letter}{required_row}")
        ws.merge_cells(start_row=diff_row, start_column=first_extra_col,
                       end_row=diff_row, end_column=day_end_col)


def run(wb: Workbook, context: Dict[str, Any]) -> None:
    """Fill the 保洁 Roster Shortfall single-sheet template."""
    project_id = context["project_id"]
    category_id = context["category_id"]
    month = context["month"]

    # Pull rendered data from /api/shortfall-engine (single source of truth).
    from ...api.shortfall_engine import api_shortfall_engine
    data = api_shortfall_engine(
        shift="A", project_id=project_id, category_id=category_id, month=month
    )
    if not data.get("has_data"):
        return

    segments = data.get("segments", [])
    days_in_month = data.get("days_in_month", 31)

    ws = wb.worksheets[0]

    # Phase 0: Adjust day columns to match month length
    day_end_col, date_row, weekday_row = _adjust_cleaning_day_columns(ws, days_in_month)

    # Discover segment positions in this template.
    title_rows = _find_segment_title_rows(ws)
    if not title_rows:
        return
    sorted_titles = sorted(title_rows.items(), key=lambda x: x[1])
    section_bounds: Dict[str, Dict[str, int]] = {}
    for i, (prefix, tr) in enumerate(sorted_titles):
        end = sorted_titles[i + 1][1] - 1 if i + 1 < len(sorted_titles) else ws.max_row
        dr = _find_data_row(ws, tr, end)
        if dr is not None:
            section_bounds[prefix] = {"title_row": tr, "data_row": dr, "end_row": end}

    # Match API segments by "N." prefix.
    api_by_prefix: Dict[str, Dict[str, Any]] = {}
    for s in segments:
        m = _TITLE_RE.match(s.get("title", ""))
        if m:
            api_by_prefix[m.group(1) + "."] = s

    # Bottom-up: insert + fill per section. Filling immediately after each
    # section's own insertion avoids stale row positions caused by later
    # inserts shifting rows above downward.
    ordered = sorted(section_bounds.keys(), key=lambda p: section_bounds[p]["data_row"], reverse=True)
    for prefix in ordered:
        sb = section_bounds[prefix]
        seg = api_by_prefix.get(prefix)
        if not seg or not seg.get("rows"):
            continue
        rows = seg["rows"]
        n = len(rows)
        template_data_row = sb["data_row"]
        end_row = template_data_row + n - 1
        if n > 1:
            insert_at = template_data_row + 1
            _insert_and_style(ws, insert_at, n - 1, source_row=template_data_row)
            _shift_formulas(ws, insert_at, n - 1)
            _expand_range(ws, template_data_row, end_row)
            # Must run AFTER _shift_formulas so the row-rewrite isn't double-shifted.
            _copy_row_formulas(ws, template_data_row, insert_at, n - 1)
        for idx, row_data in enumerate(rows):
            _fill_data_row(ws, template_data_row + idx, row_data, days_in_month)

    # Add 5th-week formulas for months with >28 days
    _add_fifth_week_formulas(ws, day_end_col, days_in_month)
