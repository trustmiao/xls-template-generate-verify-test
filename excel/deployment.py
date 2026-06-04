"""Built-in engine: Deployment sheet.

Replaces the external-file VLOOKUP formulas in the Deployment sheet (cols
E English Name, G Committed Monthly Wages, H Statutory Holiday Pay, I Extra
Typhoon Wages, J Gratuity, L Wage Received) with literal values from the
salary_master table.

Lookup chain for each data row R12..R124:
  Deployment!B(r)  →  Seq# like "M01" / "D01" / "N01"
                  ↓ first letter selects 早 / 中 / 夜 sheet
                  →  match against B column of that sheet, read C (员工编号)
                  →  match employee_no against salary_master for the project's
                     DEPT + the engine's run month

Preserved (NOT touched):
  D — Rank, F — Name 中文  : cross-sheet INDEX/MATCH within the workbook
  K — Total Hours          : SUMIF across 早/中/夜 (filled by roster_shortfall)
  M — Av Hourly Wage       : =L/K, recalculates automatically when L changes
  A — Site, B — Seq#, C/N — empty/Remarks

Also updates the "FOR THE MONTH: ..." label in R6 D-column to reflect the
engine's month parameter.
"""
from __future__ import annotations

import calendar
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook

from ...models import get_conn, normalize_emp_no


SHIFT_SHEET_BY_PREFIX = {"M": "早", "D": "中", "N": "夜"}

_MONTH_NAMES = [
    "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
    "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER",
]


def _parse_month(month: str) -> Tuple[int, int]:
    """Accept '2026-03' or '2026.03'. Default to (2026, 3) on failure."""
    if not month:
        return 2026, 3
    sep = "-" if "-" in month else "."
    try:
        y, m = month.split(sep)
        return int(y), int(m)
    except Exception:
        return 2026, 3


def _norm_name(s: Optional[str]) -> str:
    if not s:
        return ""
    return "".join(str(s).strip().upper().split())


def _load_salary_master(project_id: int, month: str) -> Dict[str, Dict[str, Any]]:
    """Load salary_master rows for the project's DEPT and the given month.
    Returns: {employee_no: row_dict}. Empty if no DEPT mapping or no data.
    """
    if not project_id:
        return {}
    sm_month = month.replace("-", ".") if month else ""
    with get_conn() as conn:
        proj = conn.execute(
            "SELECT enterprise_id FROM projects WHERE id=?", (project_id,)
        ).fetchone()
        if not proj:
            return {}
        ent_id = proj["enterprise_id"]
        dept_row = conn.execute(
            "SELECT dept_code FROM dept_project_mapping "
            "WHERE enterprise_id=? AND project_id=? LIMIT 1",
            (ent_id, project_id),
        ).fetchone()
        if not dept_row:
            return {}
        rows = conn.execute(
            """SELECT employee_no, chinese_name, english_name, position_code,
                      monthly_wage, holiday_pay, gratuity, typhoon_pay,
                      total_salary, mpf_employee
               FROM salary_master
               WHERE enterprise_id=? AND month=? AND dept_code=?""",
            (ent_id, sm_month, dept_row["dept_code"]),
        ).fetchall()
    return {r["employee_no"]: dict(r) for r in rows if r["employee_no"]}


def _build_seq_to_employee_map(wb: Workbook) -> Dict[str, Tuple[str, str]]:
    """Walk 早/中/夜 sheets and build {Seq#: (employee_no, chinese_name)}.

    Sheets have headers in R7 and data from R8 down. Col B=Seq#, C=工号, D=姓名.
    """
    seq_map: Dict[str, Tuple[str, str]] = {}
    for sheet_name in ("早", "中", "夜"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(8, ws.max_row + 1):
            seq = ws.cell(r, 2).value
            emp = ws.cell(r, 3).value
            name = ws.cell(r, 4).value
            if not seq:
                continue
            seq_str = str(seq).strip()
            if not seq_str or seq_str[0] not in "MDN":
                continue
            emp_str = normalize_emp_no(str(emp)) if emp else ""
            name_str = str(name).strip() if name else ""
            seq_map[seq_str] = (emp_str, name_str)
    return seq_map


def _update_month_label(ws, year: int, month_num: int) -> None:
    """Update the 'FOR THE MONTH: ...' line in R6. The cell is typically D6.
    Be lenient about which cell holds it — find by text prefix.
    """
    target_text = f"FOR THE MONTH: {_MONTH_NAMES[month_num - 1]} {year}"
    for r in (5, 6, 7):
        for c in range(1, 15):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.upper().startswith("FOR THE MONTH"):
                ws.cell(r, c).value = target_text
                return


def run(wb: Workbook, context: Dict[str, Any]) -> None:
    """Fill the Deployment sheet from salary_master.

    Required context keys: project_id, month (e.g. '2026-03' or '2026.03').
    """
    project_id = context.get("project_id")
    month = context.get("month") or ""
    year, month_num = _parse_month(month)

    if "Deployment" not in wb.sheetnames:
        raise ValueError("Deployment sheet not found in template")
    ws = wb["Deployment"]

    salary = _load_salary_master(int(project_id), month) if project_id else {}
    seq_map = _build_seq_to_employee_map(wb)

    # Header: update FOR THE MONTH label
    _update_month_label(ws, year, month_num)

    # Stats
    matched = 0
    unmatched: List[str] = []

    # Iterate data rows. Treat any row whose col B starts with M/D/N as data.
    for r in range(12, ws.max_row + 1):
        seq = ws.cell(r, 2).value
        if not seq:
            continue
        seq_str = str(seq).strip()
        if not seq_str or seq_str[0] not in "MDN":
            continue

        info = seq_map.get(seq_str)
        if not info:
            unmatched.append(seq_str)
            continue
        emp_no, _name = info
        master = salary.get(emp_no) if emp_no else None
        if not master:
            unmatched.append(f"{seq_str}({emp_no})")
            continue

        matched += 1

        # Column E (5) — English Name
        if master.get("english_name"):
            ws.cell(r, 5).value = master["english_name"]

        # Column G (7) — Committed Monthly Wages
        if master.get("monthly_wage") is not None:
            ws.cell(r, 7).value = float(master["monthly_wage"])

        # Column H (8) — Statutory Holiday Pay
        if master.get("holiday_pay") is not None:
            ws.cell(r, 8).value = float(master["holiday_pay"])

        # Column I (9) — Extra Typhoon Wages
        if master.get("typhoon_pay") is not None:
            ws.cell(r, 9).value = float(master["typhoon_pay"])

        # Column J (10) — Gratuity
        if master.get("gratuity") is not None:
            ws.cell(r, 10).value = float(master["gratuity"])

        # Column L (12) — Wage Received
        if master.get("total_salary") is not None:
            ws.cell(r, 12).value = float(master["total_salary"])

    # Stash a quick report so the caller can surface in metadata.
    # Output_engine __init__ writes this into the generated_files row.
    context["_deployment_report"] = {
        "matched": matched,
        "unmatched_seqs": unmatched,
        "salary_master_rows": len(salary),
    }
