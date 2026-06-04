"""Built-in engine: Generic roster fill via detect_regions + segment ops.

A simpler, more reusable alternative to `roster_shortfall.py`. Where
roster_shortfall hard-codes the 早/中/夜 sheet layout + the 大元邨 column
positions, this engine:

  1. Auto-detects template structure via excel-verify's `detect_regions`
     (any number of data segments, any date col range, header positions
     inferred from header text 排位/員工編號/姓名/Rank/職銜/工作時數/更份).
  2. Pulls rendered roster data via the existing `parse_template` +
     `render_sheet` flow used by the /api/shortfall-engine endpoint (no
     HTTP — direct function call).
  3. Resizes each data segment to fit the engine's row count using the
     spike helpers (now living in app.utils.excel_segment_ops).
  4. Clears template stock data and fills from engine rows.
  5. Trims oversized conditional-formatting sqrefs that extend past
     print_area (otherwise LO recalc on the cleaning template balloons
     from ~3s to ~60s evaluating empty XFD-range cells).

Designed to coexist with the existing roster_shortfall + cleaning_* builtins
— pick this one when the template doesn't fit a specific builtin's hard-coded
layout, or when you want the new xlsx_checks pipeline to find layout drift
in the generated file.
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook

from ...config import DOC_DIR
from ...models import list_pages_with_entries
from ..common import (
    RenderedRow,
    RenderedSheet,
    render_sheet,
    parse_template,
    clear_data_row,
    expand_segment_for_count,
    fill_data_row,
    shrink_segment_for_count,
)


SHIFT_TO_SHEET = {"A": "早", "B": "中", "C": "夜"}

# Header text → fill_data_row arg name. Used to auto-detect column positions.
HEADER_PATTERNS: Dict[str, str] = {
    "排位": "col_seq",
    "員工編號": "col_emp",
    "员工编号": "col_emp",
    "姓名": "col_name",
    "Rank": "col_rank",
    "職銜": "col_pos",
    "职衔": "col_pos",
    "工作時數": "col_hours",
    "工作时数": "col_hours",
    "更份": "col_shift",
}


def _detect_module():
    """Lazy import excel-verify's detect_regions (lives outside app/)."""
    skill_dir = (
        Path(__file__).resolve().parents[3]
        / "tools" / "skills" / "excel-verify" / "scripts"
    )
    if str(skill_dir) not in sys.path:
        sys.path.insert(0, str(skill_dir))
    import detect_regions  # noqa: E402
    return detect_regions


def _detect_header_columns(ws, header_rows=(7, 8)) -> Dict[str, int]:
    """Find column positions of well-known header labels. Returns dict of
    {fill_data_row_arg → col_index}. Missing labels → arg defaults to None."""
    found: Dict[str, int] = {}
    for r in header_rows:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None or not isinstance(v, str):
                continue
            text = v.strip()
            for kw, arg_name in HEADER_PATTERNS.items():
                if kw in text and arg_name not in found:
                    found[arg_name] = c
    return found


def _row_to_person_dict(row: RenderedRow) -> Dict[str, Any]:
    """Convert RenderedRow (dataclass) → dict suitable for fill_data_row."""
    return {
        "page_id": row.page_id,
        "rank_seq": row.rank_seq,
        "employee_no": row.employee_no,
        "name": row.name,
        "rank": row.rank_code,
        "rank_code": row.rank_code,
        "position": row.position_text,
        "position_text": row.position_text,
        "contracted_hours": row.contracted_hours,
        "shift_label": row.shift_label,
        "cells": [{"day": c.day, "code": c.code} for c in row.cells],
    }


def _resolve_template_path(context: Dict[str, Any]) -> Optional[Path]:
    """Find the template xlsx path. execute_engine() doesn't currently pass
    it through context, so we accept either:
      - context["template_path"]: explicit (set by caller)
      - context["_template_file_path"]: backward-compat fallback
    """
    tp = context.get("template_path") or context.get("_template_file_path")
    if tp:
        p = Path(tp)
        if p.is_absolute() and p.is_file():
            return p
        # Fall back to DOC_DIR-relative
        rel = DOC_DIR / str(tp)
        if rel.is_file():
            return rel
    return None


def _fill_segment(ws, *, seg_tpl: Dict[str, Any], rendered_rows: List[RenderedRow],
                  day_start: int, day_end: int, cols: Dict[str, int],
                  cum_row_delta: int, seq_prefix: str) -> int:
    """Resize a single segment to fit `rendered_rows` and fill its cells.

    Returns the row-count delta for this segment (positive = expanded,
    negative = shrunk, 0 = exact fit) so the caller can shift subsequent
    segments' base rows.
    """
    tpl_count = len(seg_tpl["data_rows"])
    base = seg_tpl["first_row"] + cum_row_delta
    needed = len(rendered_rows)

    if needed > tpl_count:
        delta = expand_segment_for_count(ws, base, tpl_count, needed)
    elif needed < tpl_count:
        delta = shrink_segment_for_count(ws, base, tpl_count, needed)
    else:
        delta = 0

    slot_count = tpl_count + delta
    data_rows = list(range(base, base + slot_count))

    # Clear template stock data
    for r in data_rows:
        clear_data_row(ws, r, last_col=day_end)

    # Fill from engine rows
    for i, rr in enumerate(rendered_rows[:slot_count]):
        fill_data_row(
            ws, data_rows[i], person=_row_to_person_dict(rr),
            seq_label=rr.rank_seq or f"{seq_prefix}{i+1:02d}",
            day_start=day_start, day_end=day_end,
            col_seq=cols.get("col_seq", 2),
            col_emp=cols.get("col_emp"),
            col_name=cols.get("col_name"),
            col_rank=cols.get("col_rank"),
            col_pos=cols.get("col_pos"),
            col_hours=cols.get("col_hours"),
            col_shift=cols.get("col_shift"),
        )
    return delta


def run(wb: Workbook, context: Dict[str, Any]) -> None:
    """Generic fill: detect_regions → render_sheet per shift → fill segments."""
    project_id = context.get("project_id")
    category_id = context.get("category_id")
    month = context.get("month", "")

    if not (project_id and category_id and month):
        raise ValueError("generic_roster requires project_id, category_id, month in context")

    template_path = _resolve_template_path(context)
    if template_path is None:
        raise FileNotFoundError(
            "generic_roster: template_path missing from context. "
            "Caller (execute_engine) must inject context['template_path']."
        )

    detect_regions = _detect_module()

    # Decide which sheets to fill. For security templates this is 早/中/夜
    # whichever are present. For cleaning templates it's a single sheet name
    # that doesn't match A/B/C; let the caller override via context.
    explicit = context.get("sheets")
    if explicit:
        sheet_shift_pairs = []
        for sn in explicit:
            shift = next((s for s, name in SHIFT_TO_SHEET.items() if name == sn), "A")
            sheet_shift_pairs.append((sn, shift))
    else:
        sheet_shift_pairs = [
            (name, shift) for shift, name in SHIFT_TO_SHEET.items() if name in wb.sheetnames
        ]
        if not sheet_shift_pairs:
            # Cleaning-style single-sheet template: pick the first non-Data sheet
            for sn in wb.sheetnames:
                if sn != "Data":
                    sheet_shift_pairs = [(sn, "A")]
                    break

    seq_prefixes = ["M", "N", "O", "P", "Q"]

    for sheet_name, shift in sheet_shift_pairs:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Detect template structure
        regions = detect_regions.detect_all(ws)
        segments_tpl = regions.get("data_segments") or []
        if not segments_tpl:
            continue
        dz = regions.get("date_zone") or {}
        day_start = dz.get("col_start", 9)
        day_end = dz.get("col_end", 36)
        cols = _detect_header_columns(ws)
        if "col_seq" not in cols:
            cols["col_seq"] = 2

        # Build rendered roster via the existing engine flow.
        # render_sheet expects each page to have its entries nested as
        # page["entries"], which is what list_pages_with_entries returns
        # (get_data_context's flat lists wouldn't work here).
        try:
            shift_pages = list_pages_with_entries(
                shift=shift,
                project_id=project_id,
                category_id=category_id,
                month=month,
            )
            tpl = parse_template(str(template_path), shift)
            rendered: RenderedSheet = render_sheet(
                tpl, shift_pages, shift, month, str(template_path)
            )
        except Exception as exc:
            # If shortfall-engine can't render this shift (e.g. cleaning
            # template, single-sheet, shift label mismatch), skip with a
            # log marker rather than failing the whole generate call.
            print(f"[generic_roster] {sheet_name} (shift={shift}): "
                  f"render_sheet failed: {type(exc).__name__}: {exc}")
            continue

        engine_segments = rendered.segments or []
        n_segments = min(len(engine_segments), len(segments_tpl))
        if n_segments == 0:
            continue

        # Resize + fill each segment in order, tracking row delta for later
        # segments since expanding segment 1 pushes segment 2 base down.
        cum_delta = 0
        for i in range(n_segments):
            seg_tpl = segments_tpl[i]
            seg_eng = engine_segments[i]
            prefix = seq_prefixes[i] if i < len(seq_prefixes) else "X"
            cum_delta += _fill_segment(
                ws,
                seg_tpl=seg_tpl,
                rendered_rows=seg_eng.rows,
                day_start=day_start,
                day_end=day_end,
                cols=cols,
                cum_row_delta=cum_delta,
                seq_prefix=prefix,
            )

        # Note: CF trim to print_area is done at the workbook level by
        # execute_engine after wb.save (see app/utils/excel_cf.py).
