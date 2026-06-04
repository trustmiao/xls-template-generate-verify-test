"""Built-in engine: Deploy Roster Shortfall fill.

Fills attendance data from pages+entries into the Deploy Roster Shortfall
Excel template, preserving formulas, formatting, and style.

Template structure (大元邨 format):
- Sheets: 早 / 中 / 夜 + Data + Deployment
- Each shift sheet has supervisor segment + security segment + summary rows
- Data rows identified by col 2 values starting with M/D/N prefix
"""
from __future__ import annotations

import calendar
import json
import re
import time
from copy import copy
from dataclasses import dataclass, field
from datetime import date
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.styles import PatternFill

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SHIFT_SHEETS = {
    "A": "早",
    "B": "中",
    "C": "夜",
}

SHIFT_PREFIX = {
    "A": "M",
    "B": "D",
    "C": "N",
}

WEEKDAY_CH = ["一", "二", "三", "四", "五", "六", "日"]

# Regex for cell references in Excel formulas (e.g., A1, $A$1, A$1, $A1)
_CELL_REF_RE = re.compile(r"(\$?[A-Z]{1,3}\$?)(\d+)")

# Regex for same-column vertical ranges (e.g., I8:I11, $I$8:$I$11)
_RANGE_RE = re.compile(r"([A-Z$]+)(\d+):\1(\d+)")

# Month name mapping (English -> for Data!B5)
_MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# Part-time GS background colours
_SUPERVISOR_PT_FILL = PatternFill(start_color="D2DAE4", end_color="D2DAE4", fill_type="solid")
_SECURITY_PT_FILL = PatternFill(start_color="D2DAE4", end_color="D2DAE4", fill_type="solid")

# Sunday / public holiday column header fill
_HOLIDAY_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class SheetSegment:
    """A contiguous block of data rows in a sheet."""
    start_row: int          # first data row (inclusive)
    end_row: int            # last data row (inclusive)
    original_end: int       # original end row before any insert/delete


@dataclass
class SheetStruct:
    """Scanned structure of a shift worksheet."""
    sheet_name: str
    data_rows: List[int] = field(default_factory=list)
    segments: List[SheetSegment] = field(default_factory=list)
    summary_rows: List[int] = field(default_factory=list)
    header_row: Optional[int] = None
    date_row: Optional[int] = None
    title_row: Optional[int] = None
    day_start_col: int = 9   # default for 早/夜; 中 uses 8
    day_end_col: Optional[int] = None  # last column holding a day cell (templates vary by month length)
    has_work_hours_col: bool = True


# ---------------------------------------------------------------------------
# Remark map (mirrors shortfall.py)
# ---------------------------------------------------------------------------

_REMARK_MAP_CACHE: Optional[Dict[str, str]] = None
_REMARK_MAP_CACHE_TS: float = 0.0
_REMARK_MAP_TTL: float = 5.0


def _build_remark_map() -> Dict[str, str]:
    global _REMARK_MAP_CACHE, _REMARK_MAP_CACHE_TS
    now = time.time()
    if _REMARK_MAP_CACHE is not None and (now - _REMARK_MAP_CACHE_TS) < _REMARK_MAP_TTL:
        return _REMARK_MAP_CACHE
    defaults: Dict[str, str] = {
        "AL": "AL",
        "SH": "FH",
        "FH": "FH",
        "H": "FH",
        "NSH": "NSH",
        "NP": "NP",
        "RD": "RD",
        "RP": "RD",
        "R": "RD",
        "SL": "SL",
        "NC": "NC",
        "ND": "ND",
        "EC": "EC",
        "F": "FS",
        "FS": "FS",
        "FW": "FS",
        "W": "WTT",
        "WTT": "WTT",
    }
    try:
        from ...models import get_settings
        settings = get_settings()
        user = json.loads(settings.get("remark_map_json") or "{}")
        _REMARK_MAP_CACHE = {**defaults, **user}
    except Exception:
        _REMARK_MAP_CACHE = defaults
    _REMARK_MAP_CACHE_TS = now
    return _REMARK_MAP_CACHE


def _cell_code(entry: Dict[str, Any]) -> str:
    """Map an OCR entry to a shortfall code."""
    remark_map = _build_remark_map()
    remark = (entry.get("remark") or "").strip()
    if remark in remark_map:
        return remark_map[remark]
    ci = (entry.get("clock_in") or "").strip()
    co = (entry.get("clock_out") or "").strip()
    for val in (ci, co):
        if val in remark_map:
            return remark_map[val]
    if entry.get("is_rest_day"):
        return "RD"
    if ci and co:
        return "8"
    if ci or co:
        return "?"
    return "X"


# ---------------------------------------------------------------------------
# Template scanning
# ---------------------------------------------------------------------------

def _scan_sheet(ws, shift: str) -> SheetStruct:
    """Scan a worksheet to identify its structure."""
    struct = SheetStruct(sheet_name=ws.title)

    data_rows: List[int] = []
    summary_keywords = ["總人數", "總時數", "合計", "Summary", "Actual", "Required", "Shortfall"]
    # Data rows are identified by a one-letter prefix + 1-3 digits in col 2
    # (e.g. M01, D02, N31). Accept any letter prefix so that cross-shift marker
    # rows like M13/M14 in the 中 sheet are also recognized — otherwise the
    # security segment won't include them and they'd get pushed past the
    # bottom of the segment by row insertion, duplicating people.
    rank_seq_re = re.compile(r"^[A-Z]\d{1,3}$")

    for r in range(1, ws.max_row + 1):
        v2 = str(ws.cell(r, 2).value or "").strip()

        if v2 and rank_seq_re.match(v2):
            data_rows.append(r)
            continue

        # Summary row
        if any(k in v2 for k in summary_keywords):
            struct.summary_rows.append(r)
            continue

        # Header row — labels are split across columns (排位 | 員工編號 | 姓名 | ...)
        if v2 == "排位":
            row_labels = " ".join(
                str(ws.cell(r, c).value or "") for c in range(2, min(ws.max_column + 1, 10))
            )
            if "員工編號" in row_labels:
                struct.header_row = r
                continue

        # Date row: contains actual date values (not formulas starting with =)
        if struct.date_row is None:
            for c in range(1, min(ws.max_column + 1, 15)):
                v = ws.cell(r, c).value
                if isinstance(v, date) and not isinstance(v, str):
                    struct.date_row = r
                    break
                if isinstance(v, str) and re.match(r"\d{4}-\d{2}-\d{2}", v.strip()):
                    struct.date_row = r
                    break

        # Title row
        if v2 and "年" in v2 and "月" in v2 and ("保安" in v2 or "出勤" in v2):
            struct.title_row = r

    struct.data_rows = data_rows
    struct.segments = _find_segments(data_rows)

    # Detect day_start_col and work_hours presence from header row
    if struct.header_row:
        has_wh = False
        day_col = None
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(struct.header_row, c).value or "").strip()
            if v == "工作時數":
                has_wh = True
            if v in ("更份", "班次"):
                day_col = c + 1
        if day_col:
            struct.day_start_col = day_col
        struct.has_work_hours_col = has_wh

    # Fallback for templates that don't have a clean "排位 + 員工編號" header
    # row (e.g. an older 大元邨 中 sheet variant that's just 排位+姓名+更份).
    # Only apply when header_row wasn't found — otherwise we'd override a
    # correctly-detected day_start=9 layout (东汇邨保安/中 has 工作時數 col).
    if not struct.header_row and ws.title == "中":
        struct.day_start_col = 8
        struct.has_work_hours_col = False

    # Detect day_end_col by walking the date row: the last column holding a
    # date or `=...+1` formula. Templates with month-specific layouts (e.g. a
    # 28-day Feb template) place aggregation labels right after the day cells,
    # so clearing "excess day columns" blindly past day_start_col + 30 would
    # wipe those aggregations.
    if struct.date_row:
        last_day_col = None
        for c in range(struct.day_start_col, ws.max_column + 1):
            v = ws.cell(struct.date_row, c).value
            if hasattr(v, "year"):
                last_day_col = c
            elif isinstance(v, str) and v.startswith("=") and "+1" in v:
                last_day_col = c
            elif v is not None and isinstance(v, str) and v.strip() and not v.startswith("="):
                break  # hit an aggregation label
        if last_day_col:
            struct.day_end_col = last_day_col

    return struct


def _find_segments(data_rows: List[int]) -> List[SheetSegment]:
    """Find contiguous segments from sorted row numbers."""
    if not data_rows:
        return []
    segments: List[SheetSegment] = []
    start = data_rows[0]
    prev = data_rows[0]
    for r in data_rows[1:]:
        if r == prev + 1:
            prev = r
        else:
            segments.append(SheetSegment(start_row=start, end_row=prev, original_end=prev))
            start = r
            prev = r
    segments.append(SheetSegment(start_row=start, end_row=prev, original_end=prev))
    return segments


# ---------------------------------------------------------------------------
# Data building
# ---------------------------------------------------------------------------

def _build_shift_data(
    shift: str,
    pages: List[Dict[str, Any]],
    entries: List[Dict[str, Any]],
    year: int,
    month_num: int,
) -> List[Dict[str, Any]]:
    """Build row data for a shift with supervisor/security segment cells.

    Part-time GS appear in BOTH segments:
    - Supervisor segment: fulltime GS + parttime GS
    - Security segment: parttime GS + non-GS
    """
    prefix = SHIFT_PREFIX.get(shift, "M")
    days_in_month = calendar.monthrange(year, month_num)[1]

    # Filter pages for this shift
    shift_pages = [p for p in pages if p.get("shift") == shift]

    # Build entry map: page_id -> day -> entry
    entry_map: Dict[int, Dict[int, Dict[str, Any]]] = {}
    for e in entries:
        pid = e.get("page_id")
        if pid not in entry_map:
            entry_map[pid] = {}
        entry_map[pid][e.get("day", 0)] = e

    # Categorise (case-insensitive — defensive read in case any rank_code was
    # inserted via a non-canonical path; writes are normalized in update_page_fields)
    def _gs(p): return (p.get("rank_code") or "").strip().upper() == "GS"
    gs_pages = [p for p in shift_pages if _gs(p)]
    fulltime_gs = [p for p in gs_pages if not p.get("is_part_time")]
    parttime_gs = [p for p in gs_pages if p.get("is_part_time")]
    other_pages = [p for p in shift_pages if not _gs(p)]

    # Build original cells for all
    def _orig_cells(page: Dict[str, Any]) -> Dict[int, str]:
        days_map = entry_map.get(page.get("id", 0), {})
        return {d: (_cell_code(days_map[d]) if d in days_map else "") for d in range(1, days_in_month + 1)}

    for p in fulltime_gs:
        p["_orig_cells"] = _orig_cells(p)
    for p in parttime_gs:
        p["_orig_cells"] = _orig_cells(p)
    for p in other_pages:
        p["_orig_cells"] = _orig_cells(p)

    # Daily call logic
    for d in range(1, days_in_month + 1):
        ft_working = sum(1 for p in fulltime_gs if p["_orig_cells"].get(d) == "8")
        needed = max(0, 2 - ft_working)
        available = sorted(
            [p for p in parttime_gs if p["_orig_cells"].get(d) == "8"],
            key=lambda p: p.get("employee_no") or "",
        )
        called_ids = {p.get("id") for p in available[:needed]}

        # Fulltime GS: only in supervisor segment
        for p in fulltime_gs:
            p.setdefault("_supervisor_cells", {})[d] = p["_orig_cells"].get(d, "")

        # Parttime GS: in both segments
        for p in parttime_gs:
            orig = p["_orig_cells"].get(d, "")
            if orig == "RD":
                p.setdefault("_supervisor_cells", {})[d] = "RD"
                p.setdefault("_security_cells", {})[d] = "RD"
            elif orig == "8":
                if p.get("id") in called_ids:
                    p.setdefault("_supervisor_cells", {})[d] = "8"
                    p.setdefault("_security_cells", {})[d] = "GS"
                else:
                    p.setdefault("_supervisor_cells", {})[d] = "G"
                    p.setdefault("_security_cells", {})[d] = "8"
            else:
                p.setdefault("_supervisor_cells", {})[d] = orig
                p.setdefault("_security_cells", {})[d] = orig

        # Non-GS: only in security segment
        for p in other_pages:
            p.setdefault("_security_cells", {})[d] = p["_orig_cells"].get(d, "")

    # Build ordered segment rows
    supervisor_pages = sorted(fulltime_gs, key=lambda p: p.get("employee_no") or "") + \
                       sorted(parttime_gs, key=lambda p: p.get("employee_no") or "")
    security_pages = sorted(parttime_gs, key=lambda p: p.get("employee_no") or "") + \
                     sorted(other_pages, key=lambda p: p.get("employee_no") or "")

    rows: List[Dict[str, Any]] = []
    for idx, page in enumerate(supervisor_pages, start=1):
        rows.append({
            "rank_seq": f"{prefix}{idx:02d}",
            "employee_no": (page.get("employee_no") or "").zfill(7) if page.get("employee_no") else "",
            "name": page.get("name") or "",
            "position_text": page.get("position_text") or "",
            "rank_code": page.get("rank_code") or "",
            "is_supervisor": True,
            "is_part_time": bool(page.get("is_part_time")),
            "shift_label": SHIFT_SHEETS.get(shift, shift),
            "cells": page.get("_supervisor_cells", {}),
            "segment": "supervisor",
        })

    for idx, page in enumerate(security_pages, start=len(supervisor_pages) + 1):
        rows.append({
            "rank_seq": f"{prefix}{idx:02d}",
            "employee_no": (page.get("employee_no") or "").zfill(7) if page.get("employee_no") else "",
            "name": page.get("name") or "",
            "position_text": page.get("position_text") or "",
            "rank_code": page.get("rank_code") or "",
            "is_supervisor": False,
            "is_part_time": bool(page.get("is_part_time")),
            "shift_label": SHIFT_SHEETS.get(shift, shift),
            "cells": page.get("_security_cells", {}),
            "segment": "security",
        })

    return rows


# ---------------------------------------------------------------------------
# Row adjustment
# ---------------------------------------------------------------------------

def _adjust_rows(
    ws,
    struct: SheetStruct,
    actual_rows: List[Dict[str, Any]],
) -> Tuple[List[Tuple[int, int]], List[Tuple[int, int, int]]]:
    """Insert or delete rows to match actual data count.

    Returns (operations, formula_copies) where:
      - operations: [(insert_at_or_delete_from, amount_signed)] for _fix_formulas
      - formula_copies: [(source_row, insert_at, amount)] for _copy_row_formulas
    """
    operations: List[Tuple[int, int]] = []

    if not struct.segments:
        return operations, []

    # Split by segment appearance
    supervisor_rows = [r for r in actual_rows if r.get("segment") == "supervisor"]
    security_rows = [r for r in actual_rows if r.get("segment") == "security"]

    # Segment 0 = supervisor, Segment 1 = security
    template_counts = [seg.end_row - seg.start_row + 1 for seg in struct.segments]

    assigned = [len(supervisor_rows), len(security_rows)]
    # Pad with zeros if more template segments than we handle
    while len(assigned) < len(template_counts):
        assigned.append(0)

    # Apply adjustments from bottom to top
    formula_copies: List[Tuple[int, int, int]] = []  # (source_row, insert_at, amount)
    for seg, target_count in zip(reversed(struct.segments), reversed(assigned)):
        current_count = seg.end_row - seg.start_row + 1
        diff = target_count - current_count

        if diff > 0:
            insert_at = seg.end_row + 1
            # Unmerge BEFORE insert_rows (sub-cells still exist for openpyxl
            # to del), re-merge at shifted positions AFTER insert (so the
            # value openpyxl moves to the new top-left is preserved).
            from openpyxl.worksheet.cell_range import CellRange
            mr_moves: List[Tuple[str, str]] = []
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row >= insert_at:
                    new = CellRange(min_col=mr.min_col, min_row=mr.min_row + diff,
                                    max_col=mr.max_col, max_row=mr.max_row + diff)
                elif mr.max_row >= insert_at:
                    new = CellRange(min_col=mr.min_col, min_row=mr.min_row,
                                    max_col=mr.max_col, max_row=mr.max_row + diff)
                else:
                    continue
                mr_moves.append((str(mr), str(new)))
            for old, _ in mr_moves:
                ws.unmerge_cells(old)
            ws.insert_rows(insert_at, amount=diff)
            for _, new in mr_moves:
                ws.merge_cells(new)
            _shift_print_area(ws, insert_at, diff)
            operations.append((insert_at, diff))
            _copy_row_style(ws, seg.end_row, insert_at, diff)
            formula_copies.append((seg.end_row, insert_at, diff))
        elif diff < 0:
            # Keep at least 1 row to avoid breaking formulas
            diff = max(diff, -(current_count - 1))
            if diff < 0:
                delete_from = seg.end_row + diff + 1
                # Mirror order for deletion: unmerge, delete, re-merge shifted.
                from openpyxl.worksheet.cell_range import CellRange
                amt = -diff
                mr_moves: List[Tuple[str, str]] = []
                for mr in list(ws.merged_cells.ranges):
                    if mr.min_row >= delete_from + amt:
                        new = CellRange(min_col=mr.min_col, min_row=mr.min_row - amt,
                                        max_col=mr.max_col, max_row=mr.max_row - amt)
                    elif mr.max_row >= delete_from:
                        new_max = max(mr.min_row, mr.max_row - amt)
                        new = CellRange(min_col=mr.min_col, min_row=mr.min_row,
                                        max_col=mr.max_col, max_row=new_max)
                    else:
                        continue
                    mr_moves.append((str(mr), str(new)))
                for old, _ in mr_moves:
                    ws.unmerge_cells(old)
                ws.delete_rows(delete_from, amount=amt)
                for _, new in mr_moves:
                    ws.merge_cells(new)
                _shift_print_area(ws, delete_from, diff)
                operations.append((delete_from, diff))

    return operations, formula_copies


def _shift_merged_ranges(ws, insert_at: int, amount: int) -> None:
    """Shift merged ranges to track row insertion/deletion.

    openpyxl's insert_rows/delete_rows don't update merged ranges, so cells in
    merged blocks below the operation point would otherwise "stick" to the
    original row, leaving labels/values mis-aligned in the output.

    Two-phase: first collect target moves, then unmerge all old ranges, then
    re-merge at new positions. Single-pass fails when shift +1 would collide
    with an already-existing range one row below.
    """
    from openpyxl.worksheet.cell_range import CellRange
    affected: List[Tuple[str, str]] = []
    for mr in list(ws.merged_cells.ranges):
        if amount > 0:
            if mr.min_row >= insert_at:
                new = CellRange(min_col=mr.min_col, min_row=mr.min_row + amount,
                                max_col=mr.max_col, max_row=mr.max_row + amount)
                affected.append((str(mr), str(new)))
            elif mr.max_row >= insert_at:
                new = CellRange(min_col=mr.min_col, min_row=mr.min_row,
                                max_col=mr.max_col, max_row=mr.max_row + amount)
                affected.append((str(mr), str(new)))
        else:  # delete
            n = -amount
            if mr.min_row >= insert_at + n:
                new = CellRange(min_col=mr.min_col, min_row=mr.min_row - n,
                                max_col=mr.max_col, max_row=mr.max_row - n)
                affected.append((str(mr), str(new)))
            elif mr.max_row >= insert_at:
                new_max = max(mr.min_row, mr.max_row - n)
                new = CellRange(min_col=mr.min_col, min_row=mr.min_row,
                                max_col=mr.max_col, max_row=new_max)
                affected.append((str(mr), str(new)))
    for old, _ in affected:
        ws.unmerge_cells(old)
    for _, new in affected:
        ws.merge_cells(new)


_PRINT_AREA_REF_RE = re.compile(r"\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)")
_PRINT_TITLE_ROW_RE = re.compile(r"\$(\d+):\$(\d+)")


def _shift_print_area(ws, insert_at: int, amount: int) -> None:
    """Shift print_area's bottom row + print_title_rows when rows shift."""
    pa = ws.print_area
    if pa:
        def repl(m: re.Match) -> str:
            c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
            if r1 >= insert_at:
                r1 += amount
            if r2 >= insert_at:
                r2 += amount
            return f"${c1}${r1}:${c2}${r2}"
        new_pa = _PRINT_AREA_REF_RE.sub(repl, str(pa))
        if new_pa != pa:
            ws.print_area = new_pa
    titles = ws.print_title_rows
    if titles:
        m = _PRINT_TITLE_ROW_RE.match(titles)
        if m:
            r1, r2 = int(m.group(1)), int(m.group(2))
            if r1 >= insert_at: r1 += amount
            if r2 >= insert_at: r2 += amount
            new_titles = f"${r1}:${r2}"
            if new_titles != titles:
                ws.print_title_rows = new_titles


def _copy_row_style(ws, source_row: int, insert_at: int, amount: int) -> None:
    """Copy cell styles from source_row to newly inserted rows."""
    for offset in range(amount):
        target_row = insert_at + offset
        for col in range(1, ws.max_column + 1):
            src = ws.cell(source_row, col)
            dst = ws.cell(target_row, col)
            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = copy(src.number_format)
                dst.protection = copy(src.protection)
                dst.alignment = copy(src.alignment)


def _copy_row_formulas(ws, source_row: int, insert_at: int, amount: int) -> None:
    """Copy row-anchored formulas from source_row to inserted rows.

    Per-day cells get overwritten by _fill_data_row, but the weekly/monthly
    aggregation columns (e.g. AF: =SUM(D{r}:J{r}); AJ: =SUM(AF{r}:AI{r}))
    only ever exist via this copy. Must run AFTER _fix_formulas so the
    target-row rewrite isn't double-shifted by the global cell-ref pass.
    """
    for offset in range(amount):
        target_row = insert_at + offset
        for col in range(1, ws.max_column + 1):
            sv = ws.cell(source_row, col).value
            if isinstance(sv, str) and sv.startswith("="):
                def _rewrite(m: re.Match) -> str:
                    col_part, ref_row = m.group(1), int(m.group(2))
                    if ref_row == source_row:
                        return f"{col_part}{target_row}"
                    return m.group(0)
                ws.cell(target_row, col).value = _CELL_REF_RE.sub(_rewrite, sv)


# ---------------------------------------------------------------------------
# Formula adjustment
# ---------------------------------------------------------------------------

def _fix_formulas(ws, struct: SheetStruct, operations: List[Tuple[int, int]]) -> None:
    """Fix all formulas after row insertions/deletions.

    Two-pass approach:
    1. Apply global row shift to all cell references (sequential, bottom-to-top).
    2. Expand vertical range references for segments that grew.
    """
    if not operations:
        return

    # Pass 1: global sequential adjustments
    for insert_at, amount in sorted(operations, key=lambda x: x[0], reverse=True):
        _adjust_formulas_for_operation(ws, insert_at, amount)

    # Pass 2: expand segment ranges that grew
    for seg in struct.segments:
        cumulative_above = sum(
            amt for at, amt in operations if at <= seg.start_row
        )
        seg_insertions = sum(
            amt for at, amt in operations
            if seg.start_row < at <= seg.end_row + 1
        )
        if seg_insertions <= 0:
            continue

        shifted_start = seg.start_row + cumulative_above
        shifted_end = seg.end_row + cumulative_above

        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if not (cell.value and isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue

                def replace_range(match: re.Match) -> str:
                    col_part = match.group(1)
                    r1 = int(match.group(2))
                    r2 = int(match.group(3))
                    if r1 == shifted_start and r2 == shifted_end:
                        return f"{col_part}{r1}:{col_part}{r2 + seg_insertions}"
                    if r1 == shifted_end and r2 == shifted_start:
                        return f"{col_part}{r2 + seg_insertions}:{col_part}{r1 + seg_insertions}"
                    return match.group(0)

                new_formula = _RANGE_RE.sub(replace_range, cell.value)
                if new_formula != cell.value:
                    cell.value = new_formula


def _adjust_formulas_for_operation(ws, insert_at: int, amount: int) -> None:
    """Shift row references in all formulas by amount for refs >= insert_at."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if not (cell.value and isinstance(cell.value, str) and cell.value.startswith("=")):
                continue
            formula = cell.value

            def replace_ref(match: re.Match) -> str:
                col_part = match.group(1)
                ref_row = int(match.group(2))
                if ref_row >= insert_at:
                    new_row = ref_row + amount
                    if new_row < 1:
                        new_row = 1
                    return f"{col_part}{new_row}"
                return match.group(0)

            new_formula = _CELL_REF_RE.sub(replace_ref, formula)
            if new_formula != formula:
                cell.value = new_formula


# ---------------------------------------------------------------------------
# Data filling
# ---------------------------------------------------------------------------

def _compute_new_segments(
    struct: SheetStruct,
    operations: List[Tuple[int, int]],
) -> List[SheetSegment]:
    """Compute segment bounds after all row operations."""
    new_segments: List[SheetSegment] = []
    for seg in struct.segments:
        cumulative_above = sum(amt for at, amt in operations if at <= seg.start_row)
        seg_insertions = sum(
            amt for at, amt in operations
            if seg.start_row < at <= seg.end_row + 1
        )
        new_start = seg.start_row + cumulative_above
        new_end = seg.end_row + cumulative_above + seg_insertions
        new_segments.append(
            SheetSegment(start_row=new_start, end_row=new_end, original_end=seg.original_end)
        )
    return new_segments


def _fill_data(
    ws,
    struct: SheetStruct,
    actual_rows: List[Dict[str, Any]],
    operations: List[Tuple[int, int]],
    year: int,
    month_num: int,
) -> None:
    """Fill actual data into the worksheet."""
    days_in_month = calendar.monthrange(year, month_num)[1]

    new_segments = _compute_new_segments(struct, operations)
    if not new_segments:
        return

    # Split by segment appearance
    supervisor_rows = [r for r in actual_rows if r.get("segment") == "supervisor"]
    security_rows = [r for r in actual_rows if r.get("segment") == "security"]

    template_counts = [seg.end_row - seg.start_row + 1 for seg in new_segments]

    seg0_needed = min(len(supervisor_rows), template_counts[0] if len(template_counts) > 0 else 0)
    seg1_needed = min(len(security_rows), template_counts[1] if len(template_counts) > 1 else 0)

    assigned = [seg0_needed, seg1_needed]
    while len(assigned) < len(template_counts):
        assigned.append(0)

    # Build ordered list per segment
    seg0_rows = supervisor_rows[:seg0_needed]
    seg1_rows = security_rows[:seg1_needed]
    segment_rows = [seg0_rows, seg1_rows]
    while len(segment_rows) < len(new_segments):
        segment_rows.append([])

    # Fill each segment
    for seg, seg_count, seg_data in zip(new_segments, assigned, segment_rows):
        for offset, row_data in enumerate(seg_data):
            if offset >= seg_count:
                break
            row_num = seg.start_row + offset
            _fill_data_row(ws, row_num, row_data, struct, days_in_month)

        # Clear any remaining rows in this segment
        for offset in range(len(seg_data), seg.end_row - seg.start_row + 1):
            row_num = seg.start_row + offset
            _clear_data_row(ws, row_num, struct, days_in_month)


def _fill_data_row(
    ws,
    row_num: int,
    row_data: Dict[str, Any],
    struct: SheetStruct,
    days_in_month: int,
) -> None:
    """Fill a single data row.

    Cols 2-4 (排位/員工編號/姓名) are always rewritten from API data.
    Cols 5 (Rank) and 6 (職銜) preserve the template's existing label unless
    the template cell is blank — OCR rank_code/position_text are often missing
    or coarse, and the manually-curated template values are more authoritative.
    """
    ws.cell(row_num, 2, value=row_data["rank_seq"])
    ws.cell(row_num, 3, value=row_data["employee_no"])
    ws.cell(row_num, 4, value=row_data["name"])
    # Rank (col 5): only fill when template cell is blank
    tpl_rank = ws.cell(row_num, 5).value
    if tpl_rank is None or tpl_rank == "":
        rank_val = row_data.get("rank_code") or ""
        if rank_val:
            ws.cell(row_num, 5, value=rank_val)
    # Position (col 6): only fill when template cell is blank
    tpl_pos = ws.cell(row_num, 6).value
    if tpl_pos is None or tpl_pos == "":
        pos_val = row_data.get("position_text") or ""
        if pos_val:
            ws.cell(row_num, 6, value=pos_val)

    # Work hours (col 7) - only if present
    if struct.has_work_hours_col:
        ws.cell(row_num, 7, value=row_data.get("contracted_hours") or 8)

    # Shift label (col before day_start)
    shift_col = struct.day_start_col - 1
    ws.cell(row_num, shift_col, value=row_data["shift_label"])

    # Day cells
    cells = row_data.get("cells", {})
    for d in range(1, days_in_month + 1):
        col = struct.day_start_col + d - 1
        val = cells.get(d, "")
        if val == "8":
            ws.cell(row_num, col, value=8)   # int so COUNT() can count it
        elif val:
            ws.cell(row_num, col, value=val) # str: RD / GS / G / FS etc.
        else:
            ws.cell(row_num, col, value="")

    # Clear excess day columns (29-31 for short months). Cap at day_end_col
    # so we don't wipe the aggregation columns that come right after the day
    # cells in month-specific templates.
    max_day_col = struct.day_end_col if struct.day_end_col else (struct.day_start_col + 30)
    for d in range(days_in_month + 1, 32):
        col = struct.day_start_col + d - 1
        if col > max_day_col:
            break
        ws.cell(row_num, col, value="")

    # Part-time GS blue background for name and employee_no
    if row_data.get("is_part_time"):
        if row_data.get("segment") == "supervisor":
            fill = _SUPERVISOR_PT_FILL
        else:
            fill = _SECURITY_PT_FILL
        ws.cell(row_num, 3).fill = fill
        ws.cell(row_num, 4).fill = fill


def _clear_data_row(
    ws,
    row_num: int,
    struct: SheetStruct,
    days_in_month: int,
) -> None:
    """Clear a template data row that has no corresponding actual data."""
    ws.cell(row_num, 2, value="")
    ws.cell(row_num, 3, value="")
    ws.cell(row_num, 4, value="")
    ws.cell(row_num, 5, value="")
    ws.cell(row_num, 6, value="")
    if struct.has_work_hours_col:
        ws.cell(row_num, 7, value="")
    shift_col = struct.day_start_col - 1
    ws.cell(row_num, shift_col, value="")
    for d in range(1, 32):
        col = struct.day_start_col + d - 1
        ws.cell(row_num, col, value="")


# ---------------------------------------------------------------------------
# Date / title updates
# ---------------------------------------------------------------------------

def _get_holiday_days(year: int, month_num: int) -> set:
    """Return set of day numbers that are public holidays."""
    try:
        from ...models import list_public_holidays
    except Exception:
        return set()
    holidays = list_public_holidays(year=year)
    result: set = set()
    for h in holidays:
        try:
            y, m, d = h["date"].split("-")
            if int(y) == year and int(m) == month_num:
                result.add(int(d))
        except Exception:
            continue
    return result


def _update_dates_and_title(
    ws,
    struct: SheetStruct,
    year: int,
    month_num: int,
) -> None:
    """Update date row, weekday row, and title row for the target month."""
    days_in_month = calendar.monthrange(year, month_num)[1]
    weekday_strip = [WEEKDAY_CH[date(year, month_num, d).weekday()] for d in range(1, days_in_month + 1)]
    holiday_days = _get_holiday_days(year, month_num)

    # Update date row
    if struct.date_row:
        for d in range(1, days_in_month + 1):
            col = struct.day_start_col + d - 1
            cell = ws.cell(struct.date_row, col, value=date(year, month_num, d))
            if d in holiday_days:
                cell.fill = _HOLIDAY_FILL
        # Clear excess date cells (cap at day_end_col; skip merged sub-cells)
        max_clear_col = struct.day_end_col if struct.day_end_col else (struct.day_start_col + 30)
        from openpyxl.cell.cell import MergedCell as _MC
        for d in range(days_in_month + 1, 32):
            col = struct.day_start_col + d - 1
            if col > max_clear_col:
                break
            cell = ws.cell(struct.date_row, col)
            if isinstance(cell, _MC):
                continue
            cell.value = ""

    # Update weekday row (header_row) — col 9..N may be cross-sheet formulas
    # (e.g. =SUBSTITUTE(TEXT(I6,"aaa"),"週","")) that already render the right
    # weekday from the date row. Only overwrite cells that are NOT formulas and
    # NOT merged sub-cells (writing to a merged sub-cell raises AttributeError).
    from openpyxl.cell.cell import MergedCell
    if struct.header_row:
        for d in range(1, days_in_month + 1):
            col = struct.day_start_col + d - 1
            cell = ws.cell(struct.header_row, col)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                # formula already produces the weekday — only colour for holiday
                if d in holiday_days:
                    cell.fill = _HOLIDAY_FILL
                continue
            cell.value = weekday_strip[d - 1]
            if d in holiday_days:
                cell.fill = _HOLIDAY_FILL
        # Clear excess weekday cells (skip merged + formula).
        # Cap at day_end_col so we don't walk into aggregation-label cols on
        # month-specific (e.g. 28-day Feb) templates where col day_start+28+
        # holds "法定假期"/"工作總日數" etc. rather than a weekday slot.
        max_clear_col = struct.day_end_col if struct.day_end_col else (struct.day_start_col + 30)
        for d in range(days_in_month + 1, 32):
            col = struct.day_start_col + d - 1
            if col > max_clear_col:
                break
            cell = ws.cell(struct.header_row, col)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                continue
            cell.value = ""

    # Update title row
    if struct.title_row:
        title_cell = ws.cell(struct.title_row, 2)
        if title_cell.value and isinstance(title_cell.value, str):
            old_title = title_cell.value
            new_title = re.sub(r"\d{4}年\d{1,2}月", f"{year}年{month_num}月", old_title)
            if new_title != old_title:
                title_cell.value = new_title


# ---------------------------------------------------------------------------
# Data sheet update
# ---------------------------------------------------------------------------

def _update_data_sheet(wb: Workbook, year: int, month_num: int) -> None:
    """Update the Data sheet month and year values."""
    if "Data" not in wb.sheetnames:
        return
    ws = wb["Data"]

    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or "").strip()
        if label == "Month":
            ws.cell(r, 2, value=_MONTH_NAMES[month_num - 1])
        elif label == "Year":
            ws.cell(r, 2, value=year)


# ---------------------------------------------------------------------------
# Main entrypoint
# ---------------------------------------------------------------------------

def _build_shift_data_from_api(
    shift: str,
    project_id: int,
    category_id: int,
    month: str,
) -> List[Dict[str, Any]]:
    """Pull row data from /api/shortfall-engine and convert to the format
    _fill_data_row expects (cells as {day: code_string}).

    This unifies the export engine with the displayed Shortfall report —
    same supervisor/security segmentation, part-time GS handling, etc.
    Replaces the legacy _build_shift_data which mis-tagged supervisors when
    OCR rank_code was missing (which is the case for shift=B/C 大元邨 +
    all 东汇邨 excel imports).
    """
    from ...api.shortfall_engine import api_shortfall_engine
    try:
        data = api_shortfall_engine(
            shift=shift, project_id=project_id, category_id=category_id, month=month
        )
    except Exception:
        return []
    if not data.get("has_data"):
        return []
    out: List[Dict[str, Any]] = []
    for seg in data.get("segments", []):
        seg_type = seg.get("type", "security")  # "supervisor" or "security"
        for row in seg.get("rows", []):
            cells_dict: Dict[int, str] = {}
            for c in row.get("cells", []) or []:
                day = c.get("day")
                if day is None:
                    continue
                # Use code for cell display (e.g. '8' for full day, 'RD' for rest).
                # Numeric value overrides happen in cleaning_shortfall.py; here we
                # keep security/supervisor templates' COUNTIF(8) formulas working.
                cells_dict[int(day)] = c.get("code") or ""
            out.append({
                "page_id": row.get("page_id"),
                "rank_seq": row.get("rank_seq") or "",
                "employee_no": row.get("employee_no") or "",
                "name": row.get("name") or "",
                "rank_code": row.get("rank_code") or "",
                "position_text": row.get("position_text") or "",
                "shift_label": row.get("shift_label") or SHIFT_SHEETS.get(shift, shift),
                "workdays": row.get("workdays", 0),
                "total_hours": row.get("total_hours") or 0,
                "is_part_time": bool(row.get("is_part_time")),
                "is_supervisor": bool(row.get("is_supervisor")),
                "segment": seg_type,
                "cells": cells_dict,
                "contracted_hours": row.get("contracted_hours") or 8,
            })
    return out


def run(wb: Workbook, context: Dict[str, Any]) -> None:
    """Fill the Deploy Roster Shortfall template workbook."""
    project_id = context.get("project_id")
    category_id = context.get("category_id")
    month = context.get("month", "")

    # Resolve year/month
    try:
        y, m = month.split("-")
        year, month_num = int(y), int(m)
    except Exception:
        year, month_num = 2026, 3

    # Process each shift sheet
    for shift, sheet_name in SHIFT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        struct = _scan_sheet(ws, shift)

        if not struct.segments:
            continue

        if project_id and category_id and month:
            actual_rows = _build_shift_data_from_api(shift, project_id, category_id, month)
        else:
            # Legacy fallback for callers that don't pass project_id (rare)
            pages = context.get("pages", [])
            entries = context.get("entries", [])
            actual_rows = _build_shift_data(shift, pages, entries, year, month_num)

        # Phase 1: Adjust rows
        operations, formula_copies = _adjust_rows(ws, struct, actual_rows)

        # Phase 2: Fix formulas
        _fix_formulas(ws, struct, operations)

        # Phase 2b: Carry per-row aggregation formulas to inserted rows.
        # Must run AFTER _fix_formulas so the target-row rewrite isn't
        # double-shifted by the global cell-ref shift pass.
        for source_row, insert_at, amount in formula_copies:
            _copy_row_formulas(ws, source_row, insert_at, amount)

        # Phase 3: Fill data
        _fill_data(ws, struct, actual_rows, operations, year, month_num)

        # Phase 4: Update dates and title
        _update_dates_and_title(ws, struct, year, month_num)

    # Update Data sheet
    _update_data_sheet(wb, year, month_num)
