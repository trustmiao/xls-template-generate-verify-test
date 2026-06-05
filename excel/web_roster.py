"""Built-in engine: web roster — adjust template and export to HTML.

The template ships with maximum personnel rows pre-configured.
For each month the engine:

  1. Updates date row to the target month
  2. Normalises number format (integers → 0 decimals, floats → 2 decimals)
  3. Deletes excess day columns from the right edge
  4. Adjusts or deletes the "29-31號時數" summary column
  5. Shrinks each personnel segment to match actual headcount
  6. Exports as HTML
"""
from __future__ import annotations

import calendar
import re
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook


def _html_escape(s: str) -> str:
    """Escape HTML special characters — avoids importing the std-lib 'html'
    module which is shadowed by the local app/engine/html/ package."""
    return (
        s.replace("&", "&amp;")
         .replace("<", "&lt;")
         .replace(">", "&gt;")
         .replace('"', "&quot;")
    )

from ..common.col_adjust import col_num_to_letter
from ..common.shift_info import _get_holidays_for_month, WEEKDAY_CH
from .. import excel_row_ops

# Re-use v2 layout detection
from .cleaning_shortfall_v2 import (
    _find_date_start,
    _update_dates,
    _adjust_day_columns,
)

RANK_COL = 2  # col B
NAME_COL = 3  # col C

_TITLE_RE = re.compile(r"^(\d+)\.\s")
_RANK_RE = re.compile(r"^[A-Z]{1,2}\d+$")


# ── Sheet → shift mapping ──
def _shift_from_sheet_name(sheet_name: str) -> str:
    mapping = {"早": "A", "中": "B", "夜": "C"}
    return mapping.get(sheet_name, "A")


# ── Data fetching ──
def _fetch_data(context: Dict[str, Any], shift: str) -> Dict[str, Any]:
    from ...api.shortfall_engine import api_shortfall_engine
    return api_shortfall_engine(
        shift=shift,
        project_id=context["project_id"],
        category_id=context["category_id"],
        month=context["month"],
    )


# ── Roster region detection ──
def _detect_roster_regions(ws) -> List[Dict[str, Any]]:
    """Detect personnel roster regions in a worksheet.

    A region is defined by:
      - A title row (col B starts with "N. ")
      - A contiguous block of data rows (col B matches rank pattern)
    """
    regions = []
    for row in range(1, ws.max_row + 1):
        b_val = ws.cell(row=row, column=RANK_COL).value
        if b_val and isinstance(b_val, str) and _TITLE_RE.match(b_val):
            title_row = row
            data_start = None
            for r in range(title_row + 1, ws.max_row + 1):
                b = ws.cell(row=r, column=RANK_COL).value
                if b and isinstance(b, str) and _RANK_RE.match(b.strip()):
                    data_start = r
                    break
            if data_start is None:
                continue
            data_end = data_start
            for r in range(data_start + 1, ws.max_row + 1):
                b = ws.cell(row=r, column=RANK_COL).value
                if b and isinstance(b, str) and _RANK_RE.match(b.strip()):
                    data_end = r
                else:
                    break
            regions.append({
                "index": len(regions) + 1,
                "title_row": title_row,
                "data_start": data_start,
                "data_end": data_end,
            })
    return regions


# ── Row shrink ──
def _shrink_region(ws, region: Dict[str, Any], target_count: int) -> int:
    """Delete excess rows from the end of a region to reach target_count."""
    current = region["data_end"] - region["data_start"] + 1
    to_delete = current - target_count
    if to_delete <= 0:
        return 0
    for _ in range(to_delete):
        row_to_delete = region["data_end"]
        excel_row_ops.delete_row_and_fixup(ws, row_to_delete)
        region["data_end"] -= 1
    return to_delete


# ── Data fill ──
def _fill_data_row(
    ws, row: int, row_data: Dict[str, Any], day_start_col: int, days_in_month: int
) -> None:
    ws.cell(row, RANK_COL, value=row_data.get("rank_seq") or row_data.get("employee_no") or "")
    ws.cell(row, NAME_COL, value=row_data.get("name") or "")
    cells = {c["day"]: c for c in row_data.get("cells", [])}
    for d in range(1, days_in_month + 1):
        col = day_start_col + d - 1
        c = cells.get(d)
        if not c:
            ws.cell(row, col, value="")
            continue
        v = c.get("value")
        if v is None:
            v = c.get("code")
        try:
            num = float(v)
            ws.cell(row, col, value=int(num) if num == int(num) else num)
        except (TypeError, ValueError):
            ws.cell(row, col, value=str(v) if v else "")


# ── Number format ──
def _adjust_number_formats(ws, day_start_col: int, days_in_month: int) -> None:
    """Set number format: integers → '0', floats → '0.00'."""
    for row in range(1, ws.max_row + 1):
        for col in range(day_start_col, day_start_col + days_in_month):
            cell = ws.cell(row, col)
            v = cell.value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                if isinstance(v, int) or v == int(v):
                    cell.number_format = "0"
                else:
                    cell.number_format = "0.00"


# ── HTML export helpers ──
def _cell_to_css(cell) -> str:
    """Extract inline CSS from an openpyxl cell style."""
    styles: List[str] = []

    def _rgb_str(rgb) -> str:
        """Normalise openpyxl Color.rgb (str or RGB object) to an 8-char string."""
        s = str(rgb) if rgb is not None else ""
        return s if len(s) == 8 else ""

    if cell.font:
        if cell.font.name:
            styles.append(f"font-family:{cell.font.name}")
        if cell.font.size:
            styles.append(f"font-size:{cell.font.size}pt")
        if cell.font.bold:
            styles.append("font-weight:bold")
        if cell.font.color and cell.font.color.rgb:
            rgb = _rgb_str(cell.font.color.rgb)
            if rgb and rgb != "00000000":
                styles.append(f"color:#{rgb[2:]}")
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = _rgb_str(cell.fill.fgColor.rgb)
        if rgb and rgb != "00000000":
            styles.append(f"background-color:#{rgb[2:]}")
    if cell.border:
        for side_name, side in [
            ("top", cell.border.top),
            ("bottom", cell.border.bottom),
            ("left", cell.border.left),
            ("right", cell.border.right),
        ]:
            if side and side.style:
                color = _rgb_str(side.color.rgb) if side.color and side.color.rgb else "000000"
                if len(color) == 8:
                    color = color[2:]
                styles.append(f"border-{side_name}:1px solid #{color}")
    if cell.alignment:
        if cell.alignment.horizontal:
            styles.append(f"text-align:{cell.alignment.horizontal}")
        if cell.alignment.vertical:
            styles.append(f"vertical-align:{cell.alignment.vertical}")
    return ";".join(styles)


def _format_cell_value(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "year"):
        return str(value.day)
    return _html_escape(str(value))


def _find_weekday_row(ws, date_row):
    """Find the weekday row below the date row."""
    for r in range(date_row + 1, min(ws.max_row, date_row + 5) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "TEXT" in v and "aaa" in v:
                return r
    return None


def workbook_to_html(wb: Workbook, month: str) -> str:
    """Convert an openpyxl workbook to a single HTML string.

    Features:
      - Renders every non-Data sheet as a separate <table>
      - Preserves merged cells (rowspan/colspan)
      - Preserves basic cell styles (font, colour, border, alignment)
      - Date row shows day number + public-holiday badge
      - Weekday row shows Chinese weekday characters
    """
    year, month_num = int(month[:4]), int(month[5:7])
    days_in_month = calendar.monthrange(year, month_num)[1]
    holidays = _get_holidays_for_month(year, month_num)

    html_parts = [
        "<!DOCTYPE html>",
        '<html>',
        '<head>',
        '<meta charset="utf-8">',
        "<title>Web Roster</title>",
        "<style>",
        "body { font-family: Arial, 'Microsoft JhengHei', sans-serif; margin: 20px; }",
        "table { border-collapse: collapse; margin-bottom: 30px; }",
        "td, th { padding: 3px 6px; border: 1px solid #ccc; white-space: nowrap; }",
        '.holiday { color: #c00; font-size: 8pt; font-weight: bold; }',
        "</style>",
        "</head>",
        "<body>",
    ]

    for ws in wb.worksheets:
        if ws.title == "Data":
            continue

        html_parts.append(f'<h2>{_html_escape(ws.title)}</h2>')
        html_parts.append("<table>")

        date_row, day_start_col = _find_date_start(ws)
        weekday_row = _find_weekday_row(ws, date_row) if date_row else None

        # Build merged-cell coverage maps
        merged_covered: set[tuple[int, int]] = set()
        merged_map: Dict[tuple[int, int], tuple[int, int]] = {}
        for mr in ws.merged_cells.ranges:
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    if (r, c) != (mr.min_row, mr.min_col):
                        merged_covered.add((r, c))
            merged_map[(mr.min_row, mr.min_col)] = (
                mr.max_row - mr.min_row + 1,
                mr.max_col - mr.min_col + 1,
            )

        for row in range(1, ws.max_row + 1):
            html_parts.append("<tr>")
            for col in range(1, ws.max_column + 1):
                if (row, col) in merged_covered:
                    continue

                cell = ws.cell(row, col)
                style = _cell_to_css(cell)
                value = cell.value

                # Date row: show day number + holiday badge
                if row == date_row and day_start_col and col >= day_start_col:
                    day = col - day_start_col + 1
                    if 1 <= day <= days_in_month:
                        val_str = str(day)
                        if day in holidays:
                            val_str += (
                                f'<br><span class="holiday">'
                                f'{_html_escape(holidays[day])}</span>'
                            )
                    else:
                        val_str = _format_cell_value(value)
                # Weekday row: show Chinese weekday
                elif row == weekday_row and day_start_col and col >= day_start_col:
                    day = col - day_start_col + 1
                    if 1 <= day <= days_in_month:
                        val_str = WEEKDAY_CH[date(year, month_num, day).weekday()]
                    else:
                        val_str = _format_cell_value(value)
                else:
                    val_str = _format_cell_value(value)

                attrs = ""
                if (row, col) in merged_map:
                    rs, cs = merged_map[(row, col)]
                    if rs > 1:
                        attrs += f' rowspan="{rs}"'
                    if cs > 1:
                        attrs += f' colspan="{cs}"'
                if style:
                    attrs += f' style="{style}"'

                html_parts.append(f"<td{attrs}>{val_str}</td>")
            html_parts.append("</tr>")

        html_parts.append("</table>")

    html_parts.append("</body></html>")
    return "\n".join(html_parts)


# ── Main entry ──
def run(wb: Workbook, context: Dict[str, Any]) -> None:
    """Adjust the workbook in-place for web display.

    Steps per sheet (skips "Data" sheets):
      1. Adjust day columns to match month length
      2. Update date row to target month
      3. Detect roster regions
      4. Fetch API data for the shift
      5. Shrink regions to match actual headcount (delete from bottom)
      6. Fill retained rows with API data
      7. Normalise number formats
    """
    month = context["month"]
    year, month_num = int(month[:4]), int(month[5:7])
    days_in_month = calendar.monthrange(year, month_num)[1]

    for ws in wb.worksheets:
        if ws.title == "Data":
            continue

        # Phase 1: Adjust day columns and dates
        _adjust_day_columns(ws, days_in_month)
        _update_dates(ws, month)

        # Phase 2: Detect roster regions
        regions = _detect_roster_regions(ws)
        if not regions:
            continue

        # Phase 3: Fetch data for this shift
        shift = _shift_from_sheet_name(ws.title)
        try:
            data = _fetch_data(context, shift)
        except Exception:
            # DB / API not available — leave template data as-is
            continue

        if not data.get("has_data"):
            # No data — clear all data rows
            for region in regions:
                for r in range(region["data_start"], region["data_end"] + 1):
                    for col in range(1, ws.max_column + 1):
                        ws.cell(r, col).value = ""
            continue

        # Match API segments to regions by index
        segments = data.get("segments", [])
        api_by_index: Dict[int, Dict[str, Any]] = {}
        for i, seg in enumerate(segments):
            api_by_index[i + 1] = seg

        # Phase 4-6: Process regions bottom-up (shrink → re-detect → fill).
        # Sort by data_start descending so lower regions are processed first.
        # This avoids row-number drift when an upper region is shrunk.
        for region in sorted(regions, key=lambda r: r["data_start"], reverse=True):
            idx = region["index"]
            seg = api_by_index.get(idx)
            if not seg:
                continue
            rows = seg.get("rows", [])
            target = len(rows)
            _shrink_region(ws, region, target)

            # Re-detect to get updated row numbers after this shrink
            fresh_regions = _detect_roster_regions(ws)
            fresh_region = next((r for r in fresh_regions if r["index"] == idx), None)
            if not fresh_region:
                continue

            # Phase 5: Fill data for this region
            date_row, day_start_col = _find_date_start(ws)
            for i, row_data in enumerate(rows):
                row = fresh_region["data_start"] + i
                _fill_data_row(ws, row, row_data, day_start_col, days_in_month)

        # Phase 6: Adjust number formats
        date_row, day_start_col = _find_date_start(ws)
        if day_start_col:
            _adjust_number_formats(ws, day_start_col, days_in_month)
