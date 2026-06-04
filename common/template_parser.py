"""Parse Excel Shortfall template into structural metadata."""
from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from functools import lru_cache
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill


def _normalize_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


@dataclass
class CellStyle:
    fill_color: Optional[str] = None
    font_color: Optional[str] = None
    font_bold: bool = False
    font_size: Optional[float] = None
    border_top: Optional[str] = None
    border_bottom: Optional[str] = None
    border_left: Optional[str] = None
    border_right: Optional[str] = None


@dataclass
class TemplateRow:
    row_index: int
    rank_seq: str
    employee_no: str
    name: str
    rank_code: str
    position_text: str
    hours: Optional[str]
    shift_label: str


@dataclass
class Segment:
    title: str
    start_row: int
    end_row: int
    type: str
    rows: List[TemplateRow] = field(default_factory=list)


@dataclass
class TemplateStructure:
    sheet_name: str
    header_row: int
    data_start_row: int
    data_end_row: int
    col_map: Dict[str, int]
    segments: List[Segment]
    summary_rows: List[int]
    styles: Dict[Tuple[int, int], CellStyle]
    day_start_col: int
    total_days_col: Optional[int]
    total_hours_col: Optional[int]
    required_counts: Dict[str, Any] = field(default_factory=dict)


def _rgb_from_color(color) -> Optional[str]:
    if color is None:
        return None
    if hasattr(color, "rgb") and color.rgb:
        v = color.rgb
        if isinstance(v, str):
            # openpyxl returns 'FFRRGGBB' or 'RRGGBB' or '00000000'
            if v == "00000000":
                return None
            if len(v) == 8:
                return f"#{v[2:]}"
            if len(v) == 6:
                return f"#{v}"
    return None


def _extract_style(cell) -> CellStyle:
    fill_color = None
    if cell.fill and isinstance(cell.fill, PatternFill) and cell.fill.fgColor:
        fill_color = _rgb_from_color(cell.fill.fgColor)

    font_color = None
    font_bold = False
    font_size = None
    if cell.font:
        font_bold = cell.font.bold or False
        font_size = cell.font.size
        if cell.font.color:
            font_color = _rgb_from_color(cell.font.color)

    def _border_color(b):
        if b and b.style and b.color:
            c = _rgb_from_color(b.color)
            if c:
                return c
        return None

    border_top = border_bottom = border_left = border_right = None
    if cell.border:
        border_top = _border_color(cell.border.top)
        border_bottom = _border_color(cell.border.bottom)
        border_left = _border_color(cell.border.left)
        border_right = _border_color(cell.border.right)

    return CellStyle(
        fill_color=fill_color,
        font_color=font_color,
        font_bold=font_bold,
        font_size=font_size,
        border_top=border_top,
        border_bottom=border_bottom,
        border_left=border_left,
        border_right=border_right,
    )


def _detect_header_row(ws, max_row: int = 20) -> int:
    """Find the first row containing 員工編號 or 姓名."""
    for r in range(1, min(max_row, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            val = _normalize_cell(ws.cell(r, c).value)
            if "員工編號" in val or val == "姓名":
                return r
    return 7  # fallback


def _detect_columns(ws, header_row: int) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col)
        raw = cell.value
        val = _normalize_cell(raw)
        low = val.lower()
        if "員工編號" in val or "employee" in low:
            col_map["employee_no"] = col
        elif val == "姓名" or "name" in low:
            col_map["name"] = col
        elif val == "Rank" or "rank" in low:
            col_map["rank"] = col
        elif "職銜" in val or "position" in low:
            col_map["position"] = col
        elif "工作時數" in val:
            col_map["hours"] = col
        elif "更份" in val or "shift" in low:
            col_map["shift"] = col
        elif "排位" in val or "seq" in low:
            col_map["seq"] = col
        elif val.isdigit() and 1 <= int(val) <= 31:
            day = int(val)
            col_map[f"day_{day}"] = col

        # 周标题列检测
        if "1-7" in val or "8-14" in val or "15-21" in val or "22-28" in val:
            col_map["week_title"] = col
        if "每月" in val and "工作時數" in val:
            col_map["monthly_hours"] = col
        if "工作日數" in val:
            col_map["work_days"] = col

        # 扫描 header_row + 1 和 header_row + 2 找日期值（保洁模板日期在列名下方）
        for offset in (1, 2):
            date_cell = ws.cell(header_row + offset, col)
            date_raw = date_cell.value
            if isinstance(date_raw, datetime):
                dt = date_raw
                if hasattr(dt, "day"):
                    col_map[f"day_{dt.day}"] = col
    return col_map


_WEEKDAY_CHARS = set("一二三四五六日")


def _detect_day_start_from_weekday(ws, weekday_row: int) -> Optional[int]:
    """Scan weekday row for day-of-week characters to find date columns."""
    for col in range(1, ws.max_column + 1):
        val = _normalize_cell(ws.cell(weekday_row, col).value)
        if val and any(ch in val for ch in _WEEKDAY_CHARS):
            return col
    return None


def _is_segment_header(val: str) -> Tuple[bool, str]:
    """Check if a cell value is a segment header. Returns (is_header, seg_type)."""
    if not val:
        return False, ""
    if re.match(r"^\d+\.\s*", val):
        if "主管" in val or "保安主任" in val:
            return True, "supervisor"
        if "保安員" in val or "出勤" in val:
            return True, "security"
    if "主管出勤" in val or ("主管" in val and "時數" in val):
        return True, "supervisor"
    if "保安員出勤" in val or ("保安員" in val and "時數" in val):
        return True, "security"
    return False, ""


def _detect_segments(ws, header_row: int, data_end_row: int) -> List[Segment]:
    """Detect segment boundaries from column B text across the entire sheet."""
    segments: List[Segment] = []
    for row in range(1, data_end_row + 1):
        val = _normalize_cell(ws.cell(row, 2).value)
        is_header, seg_type = _is_segment_header(val)
        if is_header:
            segments.append(Segment(title=val, start_row=row, end_row=data_end_row, type=seg_type))

    if not segments:
        return []

    # Assign data row boundaries
    for i, seg in enumerate(segments):
        if i == 0:
            # First segment data starts right after header row
            seg.start_row = header_row + 1
        else:
            # Subsequent segments start after their own header
            seg.start_row = segments[i].start_row + 1

        if i + 1 < len(segments):
            # End before next segment header (minus 1 for blank row if any)
            next_header_row = segments[i + 1].start_row
            seg.end_row = next_header_row - 1
            # Shrink if there are blank rows before the next header
            while seg.end_row >= seg.start_row and not _normalize_cell(ws.cell(seg.end_row, 2).value) and not _normalize_cell(ws.cell(seg.end_row, 3).value):
                seg.end_row -= 1
        else:
            # Last segment: find last data row
            seg.end_row = _find_segment_end(ws, seg.start_row, data_end_row)

    return segments


def _find_segment_end(ws, start_row: int, abs_end: int) -> int:
    """Find the last data row in a segment."""
    for row in range(start_row, abs_end + 1):
        val_b = _normalize_cell(ws.cell(row, 2).value)
        val_c = _normalize_cell(ws.cell(row, 3).value)
        name = _normalize_cell(ws.cell(row, 4).value)
        # Stop at empty row after data, or summary keyword
        if not name and not val_c and re.match(r"^\d+\.\s*", val_b):
            return row - 1
        if "總人數" in val_b or "總時數" in val_b or "合約要求" in val_b:
            return row - 1
        if "Summary" in val_b or "summary" in val_b:
            return row - 1
    return abs_end


def _detect_summary_rows(ws, data_start_row: int, data_end_row: int) -> List[int]:
    summaries: List[int] = []
    for row in range(data_start_row, data_end_row + 1):
        val_b = _normalize_cell(ws.cell(row, 2).value)
        if "總人數" in val_b or "總時數" in val_b or "合約要求" in val_b or "Summary" in val_b:
            summaries.append(row)
    return summaries


def _extract_required_counts(ws_data, summary_rows: List[int], day_start_col: int,
                              ws_formulas=None) -> Dict[str, Any]:
    """Read required counts from '要求' summary rows in the template."""
    result: Dict[str, Any] = {}
    for row in summary_rows:
        val_b = _normalize_cell(ws_data.cell(row, 2).value)
        if not val_b:
            continue

        # Cleaning template: weekly required hours → headcount
        # Handle this first because it scans the whole row and does not rely on day_start_col.
        if "每週所需工作總時數" in val_b:
            weekly_hours = None
            for c in range(1, ws_data.max_column + 1):
                v = ws_data.cell(row, c).value
                if isinstance(v, (int, float)) and v > 0:
                    weekly_hours = v
                    break
            # Fallback: blank templates' formula cells (=8*7, =5.25*7) have no
            # cached value in data_only mode. Read the FORMULA string from the
            # non-evaluated sheet and evaluate simple `=N*M` patterns.
            if weekly_hours is None and ws_formulas is not None:
                import re as _re
                for c in range(1, ws_formulas.max_column + 1):
                    fv = ws_formulas.cell(row, c).value
                    if not isinstance(fv, str) or not fv.startswith("="): continue
                    s = fv[1:].strip()
                    m = _re.fullmatch(r"(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)", s)
                    if m:
                        v = float(m.group(1)) * float(m.group(2))
                        # Multiply by 7 days if this looks like a daily-hours expression
                        # (e.g. =8*7 already encodes a week)
                        weekly_hours = v
                        break
                    # =N (single literal) inside the formula
                    m = _re.fullmatch(r"(\d+(?:\.\d+)?)", s)
                    if m:
                        weekly_hours = float(m.group(1))
                        break
            if weekly_hours:
                # Save raw weekly_hours for hour-based calculations
                if "科文" in val_b:
                    result["supervisor_weekly_hours"] = weekly_hours
                elif "VO" in val_b:
                    result["ordinary_weekly_hours"] = weekly_hours
                # Also compute headcount for backward compatibility
                daily_hours = 5.25 if "VO" in val_b else 8
                headcount = round(weekly_hours / (daily_hours * 7))
                if headcount > 0:
                    if "科文" in val_b:
                        result["supervisor"] = headcount
                    elif "VO" in val_b:
                        result["ordinary"] = headcount
                    else:
                        result["all"] = headcount
            continue

        # "合約要求" banner — the count is embedded in val_b itself
        # ("合約要求每日人數: 31"), not in any day cell. Parse it BEFORE the
        # day_val check below so it still works when the banner row's day
        # cells are blank (which is the canonical / desired shape — the
        # banner row exists only for the contractual headcount label).
        if "合約要求" in val_b:
            m = re.search(r"(\d+)", val_b)
            if m:
                result["contract"] = int(m.group(1))
            continue

        # Read first day value (day_start_col)
        day_val = ws_data.cell(row, day_start_col).value
        try:
            if isinstance(day_val, (int, float)):
                count = int(day_val)
            elif isinstance(day_val, str) and day_val.strip().isdigit():
                count = int(day_val.strip())
            else:
                continue
        except (ValueError, TypeError):
            continue

        if "保安主任" in val_b or "supervisor" in val_b.lower():
            if "要求" in val_b:
                result["supervisor"] = count
        elif "特別保安員" in val_b or "special" in val_b.lower():
            if "要求" in val_b:
                result["special"] = count
        elif "普通保安" in val_b or "ordinary" in val_b.lower():
            if "要求" in val_b:
                result["ordinary"] = count
        elif "保安每日工作" in val_b or ("all" in val_b.lower() and "保安" in val_b):
            if "要求" in val_b:
                result["all"] = count
        # NOTE: "合約要求" was handled at the top of the loop body (before the
        # day_val numeric check) so the banner row's blank day cells don't
        # short-circuit parsing.
    return result


def _is_data_row(ws, row: int, col_map: Dict[str, int]) -> bool:
    name_col = col_map.get("name")
    emp_col = col_map.get("employee_no")
    seq_col = col_map.get("seq", 2)
    name = _normalize_cell(ws.cell(row, name_col).value) if name_col else ""
    emp = _normalize_cell(ws.cell(row, emp_col).value) if emp_col else ""
    seq = _normalize_cell(ws.cell(row, seq_col).value)
    has_name_or_emp = bool(name) or bool(emp)
    if not has_name_or_emp:
        return False
    # For templates without employee_no column (e.g. cleaning roster),
    # require rank_seq to avoid picking up subtitle rows like "中文 / 英文"
    if not emp_col:
        return bool(seq)
    return True


def _collect_data_rows(ws, start_row: int, end_row: int, col_map: Dict[str, int]) -> List[TemplateRow]:
    rows: List[TemplateRow] = []
    for r in range(start_row, end_row + 1):
        if not _is_data_row(ws, r, col_map):
            continue
        # For single-sheet (cleaning) templates, absent columns must NOT fall back
        # to hardcoded defaults that overlap day columns (e.g. default 3 = day1).
        seq_col = col_map.get("seq", 2)
        emp_col = col_map.get("employee_no")
        name_col = col_map.get("name", 4)
        rank_col = col_map.get("rank")
        pos_col = col_map.get("position")
        hours_col = col_map.get("hours")
        shift_col = col_map.get("shift")

        rank_seq = _normalize_cell(ws.cell(r, seq_col).value)
        emp = _normalize_cell(ws.cell(r, emp_col).value) if emp_col else ""
        name = _normalize_cell(ws.cell(r, name_col).value)
        rank = _normalize_cell(ws.cell(r, rank_col).value) if rank_col else ""
        pos = _normalize_cell(ws.cell(r, pos_col).value) if pos_col else ""
        hours = _normalize_cell(ws.cell(r, hours_col).value) if hours_col else None
        shift_lbl = _normalize_cell(ws.cell(r, shift_col).value) if shift_col else ""

        rows.append(TemplateRow(
            row_index=r,
            rank_seq=rank_seq,
            employee_no=emp,
            name=name,
            rank_code=rank,
            position_text=pos,
            hours=hours,
            shift_label=shift_lbl,
        ))
    return rows


def parse_template(excel_path: str, shift: str) -> TemplateStructure:
    """Parse a template sheet into structure metadata.

    Result is cached by (abspath, mtime, shift): the scan is pure and the
    returned TemplateStructure is consumed read-only by render_sheet, so
    repeated views of the same template/shift skip the ~0.7 s openpyxl scan.
    """
    try:
        mtime = os.path.getmtime(excel_path)
    except OSError:
        mtime = 0.0
    return _parse_template_cached(os.path.abspath(excel_path), mtime, shift)


@lru_cache(maxsize=16)
def _parse_template_cached(excel_path: str, mtime: float, shift: str) -> TemplateStructure:
    from .shift_info import load_template_workbook
    wb = load_template_workbook(excel_path, data_only=False)

    # Sheet name from shift label
    from .shift_info import SHIFT_INFO
    label = SHIFT_INFO[shift]["label"]
    if label not in wb.sheetnames:
        # Fallback: single-sheet roster (cleaning)
        roster_sheets = [s for s in wb.sheetnames if s.startswith("Roster-")]
        if roster_sheets:
            label = roster_sheets[0]
        else:
            raise ValueError(f"Sheet '{label}' not found in {excel_path}")

    ws = wb[label]

    header_row = _detect_header_row(ws)
    col_map = _detect_columns(ws, header_row)

    # Determine day start and total columns
    day_cols = sorted([c for k, c in col_map.items() if k.startswith("day_")])
    if day_cols:
        day_start_col = day_cols[0]
    else:
        # Fallback: scan weekday row for day-of-week characters
        day_start_col = _detect_day_start_from_weekday(ws, header_row + 2)
        if not day_start_col:
            day_start_col = 9 if "hours" in col_map else 8

    # total_days_col / total_hours_col only for security templates (has employee_no)
    if "employee_no" in col_map and day_cols:
        total_days_col = col_map.get("day_31", day_start_col + 30) + 1
        total_hours_col = total_days_col + 1
    else:
        total_days_col = None
        total_hours_col = None

    # Scan for data boundaries
    data_start_row = header_row + 1
    data_end_row = ws.max_row
    # Shrink data_end_row to last row with meaningful content
    for r in range(ws.max_row, header_row, -1):
        has_content = False
        for c in range(1, min(ws.max_column, 10) + 1):
            if _normalize_cell(ws.cell(r, c).value):
                has_content = True
                break
        if has_content:
            data_end_row = r
            break
    else:
        data_end_row = header_row

    # Detect segments and summary rows
    segments = _detect_segments(ws, header_row, data_end_row)
    summary_rows = _detect_summary_rows(ws, data_start_row, data_end_row)

    # Load data_only version for evaluated cell values (formulas → results)
    try:
        wb_data = load_workbook(excel_path, data_only=True)
        ws_data = wb_data[label]
    except Exception:
        wb_data = None
        ws_data = None

    # Collect data rows per segment (or all if no segments)
    # Use data_only sheet so formula cells yield evaluated values, not formula strings
    ws_rows = ws_data if ws_data is not None else ws
    if segments:
        for seg in segments:
            seg.rows = _collect_data_rows(ws_rows, seg.start_row, seg.end_row, col_map)
    else:
        # No segments: treat everything as one security segment
        all_rows = _collect_data_rows(ws_rows, data_start_row, data_end_row, col_map)
        if all_rows:
            segments.append(Segment(
                title="",
                start_row=data_start_row,
                end_row=data_end_row,
                type="security",
                rows=all_rows,
            ))

    # Extract styles for all relevant cells (limit scope to avoid huge payload)
    styles: Dict[Tuple[int, int], CellStyle] = {}
    style_rows = {header_row, header_row + 1}
    for seg in segments:
        style_rows.add(seg.start_row - 1)  # segment header
        for tr in seg.rows:
            style_rows.add(tr.row_index)
        style_rows.add(seg.end_row + 1)
    for sr in summary_rows:
        style_rows.add(sr)

    for r in style_rows:
        if r < 1 or r > ws.max_row:
            continue
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            st = _extract_style(cell)
            if st.fill_color or st.font_bold:
                styles[(r, c)] = st

    # Read evaluated required-count values from data_only sheet
    try:
        if ws_data is not None:
            required_counts = _extract_required_counts(ws_data, summary_rows, day_start_col, ws_formulas=ws)
        else:
            required_counts = _extract_required_counts(ws, summary_rows, day_start_col, ws_formulas=ws)
    except Exception:
        required_counts = {}

    return TemplateStructure(
        sheet_name=label,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        col_map=col_map,
        segments=segments,
        summary_rows=summary_rows,
        styles=styles,
        day_start_col=day_start_col,
        total_days_col=total_days_col,
        total_hours_col=total_hours_col,
        required_counts=required_counts,
    )
