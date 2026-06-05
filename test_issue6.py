#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Issue #6 test: Adjust personnel counts for Feb and Apr across all templates.

For each template:
  - Feb (28 days): segment1→1 person, segment2→2 people, segment3→3 people (if exists)
  - Apr (30 days): segment1→1 person, segment2→2 people, segment3→3 people (if exists)

Steps per template/month:
  1. _adjust_day_columns() + _update_dates() to set the month
  2. detect roster regions
  3. delete rows from each region until target count is reached
"""

import re
import calendar
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, r"D:\claude\claude_hk\backend")
from app.engine.excel.cleaning_shortfall_v2 import (
    _adjust_day_columns,
    _update_dates,
)

TEMPLATES = [
    ("东汇-保安", "templates/东汇-保安轮休表template.xlsx", [1, 2]),
    ("东汇-保洁", "templates/东汇-保洁轮休表_template.xlsx", [1, 2, 3]),
    ("TY-大元保安", "templates/TY-2026.03- SG_SEC-Deploy Roster  Shortfall - template.xlsx", [1, 2]),
]

OUT_DIR = Path("test_outputs_issue6")
OUT_DIR.mkdir(exist_ok=True)

# ── Region detection (copied from delete_one_row.py) ──
rank_pattern = re.compile(r'^[A-Z]+\d+$')
title_pattern = re.compile(r'^\d+\.\s+')


def detect_roster_regions(ws):
    regions = []
    for row in range(1, ws.max_row + 1):
        b_val = ws.cell(row=row, column=2).value
        if b_val and isinstance(b_val, str) and title_pattern.match(b_val):
            title_row = row
            data_start = None
            for r in range(title_row + 1, ws.max_row + 1):
                b = ws.cell(row=r, column=2).value
                if b and isinstance(b, str) and rank_pattern.match(b.strip()):
                    data_start = r
                    break
            if data_start is None:
                continue
            data_end = data_start
            for r in range(data_start + 1, ws.max_row + 1):
                b = ws.cell(row=r, column=2).value
                is_data = b and isinstance(b, str) and rank_pattern.match(b.strip())
                if is_data:
                    data_end = r
                else:
                    break
            regions.append({
                'index': len(regions) + 1,
                'title_row': title_row,
                'data_start': data_start,
                'data_end': data_end,
                'title': b_val,
            })
    return regions


# ── Formula adjustment ──
cell_ref = re.compile(r'(?<![A-Z$])(\$?[A-Z]{1,3})(\$?)(\d+)')


def adjust_formula(formula, deleted_row):
    """
    Adjust cell references after deleting a row.

    - Range refs (A1:A5): both bounds use >= rule (shrink range)
    - Single refs (A1):    use > rule only (A1→#REF! if row==deleted)
    """
    # First, identify which refs are part of a range (contain ':' nearby)
    # We'll do a two-pass: first find all ranges, mark their positions
    range_positions = set()
    range_pattern = re.compile(r'(?<![A-Z$])(\$?[A-Z]{1,3})(\$?)(\d+)(:(\$?[A-Z]{1,3})(\$?)(\d+))')
    for m in range_pattern.finditer(formula):
        range_positions.add((m.start(), m.end()))

    result = []
    last_end = 0

    for match in cell_ref.finditer(formula):
        # Skip if already handled as part of a range (we handle ranges below)
        # Actually, cell_ref finditer will also match inside ranges.
        # We need to detect if this match is inside a range match.
        in_range = any(start <= match.start() < end for start, end in range_positions)

        result.append(formula[last_end:match.start()])
        col_part = match.group(1)
        abs_row = match.group(2) == '$'
        row_num = int(match.group(3))

        if not abs_row:
            if in_range:
                # Range bounds: >= deleted_row → shift
                if row_num >= deleted_row:
                    row_num -= 1
            else:
                # Single ref: > deleted_row → shift; == deleted_row → #REF!
                if row_num > deleted_row:
                    row_num -= 1
                elif row_num == deleted_row:
                    result.append("#REF!")
                    last_end = match.end()
                    continue

        result.append(f"{col_part}{'$' if abs_row else ''}{row_num}")
        last_end = match.end()

    result.append(formula[last_end:])
    return ''.join(result)


def adjust_all_formulas(ws, deleted_row):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = cell.value
            if val and isinstance(val, str) and val.startswith('='):
                cell.value = adjust_formula(val, deleted_row)


def adjust_merged_range(range_str, deleted_row):
    ref_pattern = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')

    def parse_ref(ref):
        m = ref_pattern.match(ref)
        if m:
            col_abs, col, row_abs, row = m.groups()
            row = int(row)
            if row > deleted_row:
                row -= 1
            return f"{col_abs}{col}{row_abs}{row}"
        return ref

    if ':' in range_str:
        start, end = range_str.split(':')
        return f"{parse_ref(start)}:{parse_ref(end)}"
    return parse_ref(range_str)


def fix_merged_cells(ws, deleted_row):
    merged_ranges = [str(mc) for mc in ws.merged_cells.ranges]
    for mc_str in merged_ranges:
        try:
            ws.unmerge_cells(mc_str)
        except Exception:
            pass
    for mc_str in merged_ranges:
        adjusted = adjust_merged_range(mc_str, deleted_row)
        try:
            ws.merge_cells(adjusted)
        except Exception:
            pass


def safe_delete_rows(ws, idx, amount=1):
    """
    Delete rows, correctly mimicking Excel behavior.
    Replaces openpyxl's ws.delete_rows() which has an iter_rows bug that
    creates phantom empty cells which then overwrite valid data during shift.
    """
    max_row = max((row for row, col in ws._cells.keys()), default=0)

    for r in range(idx, max_row + 1):
        source_r = r + amount
        if source_r > max_row:
            for key in list(ws._cells.keys()):
                if key[0] == r:
                    del ws._cells[key]
            continue

        target_cols = {col for row, col in ws._cells.keys() if row == r}
        source_cols = {col for row, col in ws._cells.keys() if row == source_r}

        for col in target_cols - source_cols:
            del ws._cells[(r, col)]

        for col in source_cols:
            if (source_r, col) in ws._cells:
                cell = ws._cells[(source_r, col)]
                ws._cells[(r, col)] = cell
                cell.row = r
                del ws._cells[(source_r, col)]

    for r in range(idx, max_row + 1):
        source_r = r + amount
        if source_r in ws.row_dimensions:
            ws.row_dimensions[r].height = ws.row_dimensions[source_r].height
        elif r in ws.row_dimensions:
            ws.row_dimensions[r].height = None

    ws._current_row = max(row for row, col in ws._cells.keys()) if ws._cells else 0


def adjust_print_area(ws, deleted_row):
    """Adjust print-area row numbers after a row is deleted."""
    if not ws.print_area or not ws._print_area:
        return
    for cr in ws._print_area.ranges:
        if cr.min_row > deleted_row:
            cr.min_row -= 1
        if cr.max_row > deleted_row:
            cr.max_row -= 1


# ── Core deletion ──
def delete_second_row_of_region(ws, region):
    """Delete the 2nd data row of a region and fix formulas + merged cells + print area."""
    row_to_delete = region['data_start'] + 1
    if row_to_delete > region['data_end']:
        return False

    merged_ranges = [str(mc) for mc in ws.merged_cells.ranges]
    adjust_all_formulas(ws, row_to_delete)
    safe_delete_rows(ws, row_to_delete, 1)

    for mc_str in merged_ranges:
        try:
            ws.unmerge_cells(mc_str)
        except Exception:
            pass
    for mc_str in merged_ranges:
        adjusted = adjust_merged_range(mc_str, row_to_delete)
        try:
            ws.merge_cells(adjusted)
        except Exception:
            pass
    adjust_print_area(ws, row_to_delete)
    return True


def process_template(name, path, targets, month, out_dir):
    """Generate output for one template + month with specified segment targets."""
    year, mon = map(int, month.split('-'))
    days_in_month = calendar.monthrange(year, mon)[1]

    wb = load_workbook(path, data_only=False)

    # Step 1: Adjust day columns and dates on every sheet
    for ws in wb.worksheets:
        if ws.title == "Data":
            continue
        _adjust_day_columns(ws, days_in_month)
        _update_dates(ws, month)

    # Step 2: Delete rows from each region to reach target counts
    for ws in wb.worksheets:
        if ws.title == "Data":
            continue
        regions = detect_roster_regions(ws)
        print(f"\n  Sheet '{ws.title}': {len(regions)} region(s)")
        for region in regions:
            idx = region['index']
            if idx > len(targets):
                print(f"    Region {idx}: skipped (no target defined)")
                continue
            target = targets[idx - 1]
            current = region['data_end'] - region['data_start'] + 1
            to_delete = current - target
            if to_delete <= 0:
                print(f"    Region {idx}: already has {current} person(s), no deletion needed")
                continue

            print(f"    Region {idx}: {current} → {target} (deleting {to_delete} row(s))")
            for _ in range(to_delete):
                # Re-detect because row numbers shift after each deletion
                regions_now = detect_roster_regions(ws)
                r = next((x for x in regions_now if x['index'] == idx), None)
                if r is None:
                    break
                ok = delete_second_row_of_region(ws, r)
                if not ok:
                    break

    out_name = f"{month}_{name}_adjusted.xlsx"
    out_path = out_dir / out_name
    wb.save(str(out_path))
    print(f"  💾 Saved: {out_path}")
    return out_path


def inspect_file(path, label):
    """Print final segment structure."""
    wb = load_workbook(path, data_only=False)
    print(f"\n{'='*60}")
    print(f"INSPECT: {label}")
    print(f"{'='*60}")
    for ws in wb.worksheets:
        if ws.title == "Data":
            continue
        regions = detect_roster_regions(ws)
        print(f"\n  Sheet '{ws.title}':")
        for r in regions:
            count = r['data_end'] - r['data_start'] + 1
            names = []
            for row in range(r['data_start'], r['data_end'] + 1):
                rank = ws.cell(row, 2).value or ""
                name = ws.cell(row, 3).value or ""
                names.append(f"{rank}/{name}")
            print(f"    Region {r['index']}: {count} person(s) — {names}")


def main():
    for name, path, targets in TEMPLATES:
        for month in ["2026-02", "2026-04"]:
            print(f"\n{'='*60}")
            print(f"Processing: {name} | {month} | targets={targets}")
            print(f"{'='*60}")
            out = process_template(name, path, targets, month, OUT_DIR)
            inspect_file(out, f"{name} {month}")

    print(f"\n{'='*60}")
    print(f"All outputs saved to: {OUT_DIR}")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
